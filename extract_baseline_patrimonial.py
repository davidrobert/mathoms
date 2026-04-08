#!/usr/bin/env python3
"""
Extract patrimônio baseline from IRPF declarations, receipts, QuintoAndar rent reports, and real estate XLSX.
Step 6b: E1.5 - Baseline Patrimonial extraction
"""

import json
import re
from datetime import datetime
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    print("Installing pdfplumber...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'pdfplumber', '--break-system-packages'])
    import pdfplumber

try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl', '--break-system-packages'])
    from openpyxl import load_workbook

BASE_DIR = Path("/sessions/sharp-cool-shannon/mnt/Financas Familia/financas-familia")
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "processed" / "E2_extracts"

# Ensure output directory exists
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def extract_irpf_declaration(pdf_path):
    """Extract patrimônio data from IRPF declaration PDF."""
    result = {
        "source_file": str(pdf_path),
        "extraction_type": "irpf_declaracao",
        "bens_direitos": [],
        "total_bens": None,
        "rendimentos": {
            "pj": None,
            "clt": None,
            "alugueis": None,
            "financeiros": None
        },
        "pagamentos_deductiveis": None,
        "extraction_date": datetime.now().isoformat()
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            for page in pdf.pages:
                full_text += page.extract_text() or ""

            # Extract total de bens
            match = re.search(r'Total de Bens e Direitos\s*[:\-]?\s*([\d.,]+)', full_text, re.IGNORECASE)
            if match:
                valor_str = match.group(1).replace('.', '').replace(',', '.')
                result["total_bens"] = float(valor_str)

            # Extract bens_direitos - look for line patterns like "01 - Imóvel"
            # Pattern: grupo codigo - descricao ... valor_anterior ... valor_atual
            bens_section = re.search(r'BENS E DIREITOS(.*?)(?:LIABILITIES|DÍVIDAS E OBRIGAÇÕES|$)',
                                    full_text, re.IGNORECASE | re.DOTALL)

            if bens_section:
                bens_text = bens_section.group(1)
                # Look for patterns with grupo, codigo
                lines = bens_text.split('\n')

                for i, line in enumerate(lines):
                    # Pattern: "01 - Imóvel" or "01-Imóvel"
                    match = re.match(r'^(\d{2})\s*[\-]?\s*(.+)$', line.strip())
                    if match:
                        grupo = match.group(1)
                        descricao = match.group(2).strip()

                        bem = {
                            "grupo": grupo,
                            "codigo": grupo,  # In many IRPF docs, codigo = grupo
                            "descricao": descricao,
                            "valor_31_12_anterior": None,
                            "valor_31_12_atual": None
                        }

                        # Look for values in following lines
                        for j in range(i+1, min(i+5, len(lines))):
                            next_line = lines[j].strip()
                            # Extract numeric values
                            values = re.findall(r'([\d.,]+)', next_line)
                            if len(values) >= 2:
                                try:
                                    val1 = float(values[0].replace('.', '').replace(',', '.'))
                                    val2 = float(values[1].replace('.', '').replace(',', '.'))
                                    bem["valor_31_12_anterior"] = val1
                                    bem["valor_31_12_atual"] = val2
                                    break
                                except:
                                    pass

                        result["bens_direitos"].append(bem)

            # Extract rendimentos
            # PJ - Rendimentos de PJ
            match = re.search(r'Rendimentos.*?PJ\s*[:\-]?\s*([\d.,]+)', full_text, re.IGNORECASE)
            if match:
                result["rendimentos"]["pj"] = float(match.group(1).replace('.', '').replace(',', '.'))

            # CLT - Salários/Rendimentos
            match = re.search(r'Rendimentos.*?CLT\s*[:\-]?\s*([\d.,]+)|Total de.*?Rendimentos\s*[:\-]?\s*([\d.,]+)',
                            full_text, re.IGNORECASE)
            if match:
                valor_str = match.group(1) or match.group(2)
                if valor_str:
                    result["rendimentos"]["clt"] = float(valor_str.replace('.', '').replace(',', '.'))

            # Aluguéis
            match = re.search(r'Aluguéis?\s*[:\-]?\s*([\d.,]+)', full_text, re.IGNORECASE)
            if match:
                result["rendimentos"]["alugueis"] = float(match.group(1).replace('.', '').replace(',', '.'))

            # Rendimentos financeiros
            match = re.search(r'Rendimentos?\s+Financeiros?\s*[:\-]?\s*([\d.,]+)', full_text, re.IGNORECASE)
            if match:
                result["rendimentos"]["financeiros"] = float(match.group(1).replace('.', '').replace(',', '.'))

            # Pagamentos deductiveis
            match = re.search(r'Pagamentos?.*?Dedutíveis?\s*[:\-]?\s*([\d.,]+)', full_text, re.IGNORECASE)
            if match:
                result["pagamentos_deductiveis"] = float(match.group(1).replace('.', '').replace(',', '.'))

    except Exception as e:
        result["error"] = str(e)

    return result

def extract_irpf_receipt(pdf_path):
    """Extract data from IRPF receipt PDF."""
    result = {
        "source_file": str(pdf_path),
        "extraction_type": "irpf_recibo",
        "imposto_total": None,
        "imposto_devido": None,
        "data_envio": None,
        "extraction_date": datetime.now().isoformat()
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            for page in pdf.pages:
                full_text += page.extract_text() or ""

            # Extract imposto total
            match = re.search(r'Imposto Total\s*[:\-]?\s*([\d.,]+)', full_text, re.IGNORECASE)
            if match:
                result["imposto_total"] = float(match.group(1).replace('.', '').replace(',', '.'))

            # Extract imposto devido
            match = re.search(r'Imposto\s+(?:Devido|a Pagar)\s*[:\-]?\s*([\d.,]+)', full_text, re.IGNORECASE)
            if match:
                result["imposto_devido"] = float(match.group(1).replace('.', '').replace(',', '.'))

            # Extract data de envio
            match = re.search(r'Data\s+(?:de\s+)?Envio\s*[:\-]?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{4})', full_text, re.IGNORECASE)
            if match:
                result["data_envio"] = match.group(1)

    except Exception as e:
        result["error"] = str(e)

    return result

def extract_quintoandar_rent(pdf_path):
    """Extract data from QuintoAndar rent report PDF."""
    result = {
        "source_file": str(pdf_path),
        "extraction_type": "quintoandar_aluguel",
        "properties": [],
        "extraction_date": datetime.now().isoformat()
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            tables = []
            for page in pdf.pages:
                full_text += page.extract_text() or ""
                # Try to extract tables
                if page.tables:
                    tables.extend(page.tables)

            # Try to extract from tables first
            for table in tables:
                for row in table:
                    if row and len(row) >= 3:
                        prop_info = {
                            "descricao": None,
                            "renda_bruta": None,
                            "descontos": None,
                            "renda_liquida": None,
                            "ano": None
                        }
                        # Try to parse row values
                        if any(str(cell).strip() for cell in row):
                            prop_info["descricao"] = str(row[0]) if row[0] else None
                            try:
                                if len(row) > 1 and row[1]:
                                    prop_info["renda_bruta"] = float(str(row[1]).replace('.', '').replace(',', '.'))
                                if len(row) > 2 and row[2]:
                                    prop_info["descontos"] = float(str(row[2]).replace('.', '').replace(',', '.'))
                                if len(row) > 3 and row[3]:
                                    prop_info["renda_liquida"] = float(str(row[3]).replace('.', '').replace(',', '.'))
                            except:
                                pass

                        if prop_info["descricao"]:
                            result["properties"].append(prop_info)

            # Extract from text patterns
            # Look for renda_bruta_aluguel, descontos, renda_liquida
            match = re.search(r'Renda\s+(?:Bruta|Mensal).*?[:\-]?\s*([\d.,]+)', full_text, re.IGNORECASE)
            if match and not result["properties"]:
                renda_bruta = float(match.group(1).replace('.', '').replace(',', '.'))

                match_desc = re.search(r'Descontos?\s*[:\-]?\s*([\d.,]+)', full_text, re.IGNORECASE)
                descontos = None
                if match_desc:
                    descontos = float(match_desc.group(1).replace('.', '').replace(',', '.'))

                match_liq = re.search(r'Renda\s+Líquida\s*[:\-]?\s*([\d.,]+)', full_text, re.IGNORECASE)
                renda_liquida = None
                if match_liq:
                    renda_liquida = float(match_liq.group(1).replace('.', '').replace(',', '.'))

                result["properties"].append({
                    "descricao": "Imóvel não identificado",
                    "renda_bruta": renda_bruta,
                    "descontos": descontos,
                    "renda_liquida": renda_liquida,
                    "ano": 2025
                })

    except Exception as e:
        result["error"] = str(e)

    return result

def extract_real_estate_xlsx(xlsx_path):
    """Extract data from real estate XLSX file."""
    result = {
        "source_file": str(xlsx_path),
        "extraction_type": "real_estate",
        "properties": [],
        "extraction_date": datetime.now().isoformat()
    }

    try:
        wb = load_workbook(xlsx_path)

        # Try to find the main data sheet
        sheet = None
        for sheet_name in wb.sheetnames:
            if any(keyword in sheet_name.lower() for keyword in ['dados', 'imovel', 'property', 'properties', 'real estate']):
                sheet = wb[sheet_name]
                break

        if not sheet:
            sheet = wb.active

        # Extract headers
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower())

        # Extract rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if not any(cell.value for cell in row[:3]):  # Skip empty rows
                continue

            prop = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    header = headers[col_idx]
                    value = cell.value

                    # Convert datetime objects to string
                    if isinstance(value, datetime):
                        value = value.isoformat()

                    # Map common headers
                    if 'endereco' in header or 'address' in header or 'local' in header:
                        prop["endereco"] = value
                    elif 'data' in header and 'aquisicao' in header:
                        prop["data_aquisicao"] = str(value) if value else None
                    elif 'valor' in header and 'compra' in header:
                        try:
                            prop["valor_compra"] = float(value) if value else None
                        except:
                            prop["valor_compra"] = value
                    elif 'financiamento' in header:
                        try:
                            prop["financiamento"] = float(value) if value else None
                        except:
                            prop["financiamento"] = value
                    elif 'status' in header:
                        prop["status"] = value
                    else:
                        # Store all other columns
                        if header:
                            prop[header] = value

            if prop:
                result["properties"].append(prop)

    except Exception as e:
        result["error"] = str(e)

    return result

def main():
    """Main extraction function."""

    print("=" * 80)
    print("STEP 6b: E1.5 - Baseline Patrimonial Extraction")
    print("=" * 80)

    # Define input files and their output names
    files_to_process = [
        (DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracao_2023-0_original.pdf",
         "receitafederal_irpfdeclaracao_2023-2_extract.json",
         "irpf_declaracao"),

        (DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracao_2024-0_original.pdf",
         "receitafederal_irpfdeclaracao_2024-2_extract.json",
         "irpf_declaracao"),

        (DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracaomariana_2024-0_original.pdf",
         "receitafederal_irpfdeclaracaomariana_2024-2_extract.json",
         "irpf_declaracao"),

        (DATA_DIR / "income_tax_br" / "receitafederal_irpfrecibo_2024-0_original.pdf",
         "receitafederal_irpfrecibo_2024-2_extract.json",
         "irpf_recibo"),

        (DATA_DIR / "income_tax_br" / "receitafederal_irpfrecibomariana_2024-0_original.pdf",
         "receitafederal_irpfrecibomariana_2024-2_extract.json",
         "irpf_recibo"),

        (DATA_DIR / "financial_statements" / "quintoandar_informerendimentosaluguel_2025-0_original.pdf",
         "quintoandar_informerendimentosaluguel_2025-2_extract.json",
         "quintoandar_aluguel"),

        (DATA_DIR / "financial_statements" / "quintoandar_informerendimentosaluguelmariana_2025-0_original.pdf",
         "quintoandar_informerendimentosaluguelmariana_2025-2_extract.json",
         "quintoandar_aluguel"),

        (DATA_DIR / "real_estate" / "dados_imoveis-0_original.xlsx",
         "dados_imoveis-2_extract.json",
         "real_estate"),
    ]

    extracted_data = {}
    declarations = []
    receipts = []
    properties = []

    # Process each file
    for input_path, output_name, file_type in files_to_process:
        print(f"\nProcessing: {input_path.name}")
        print(f"  Type: {file_type}")

        if not input_path.exists():
            print(f"  ERROR: File not found!")
            continue

        try:
            if file_type == "irpf_declaracao":
                data = extract_irpf_declaration(input_path)
                declarations.append(data)
            elif file_type == "irpf_recibo":
                data = extract_irpf_receipt(input_path)
                receipts.append(data)
            elif file_type == "quintoandar_aluguel":
                data = extract_quintoandar_rent(input_path)
                if data["properties"]:
                    properties.extend(data["properties"])
            elif file_type == "real_estate":
                data = extract_real_estate_xlsx(input_path)
                if data["properties"]:
                    properties.extend(data["properties"])

            # Save individual extract
            output_path = OUTPUT_DIR / output_name
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            print(f"  Saved: {output_path.name}")
            extracted_data[output_name] = data

        except Exception as e:
            print(f"  ERROR during extraction: {e}")

    # Create consolidated baseline
    consolidated = {
        "pipeline_stage": "E1.5_Baseline_Patrimonial",
        "data_processamento": datetime.now().isoformat(),
        "membros": ["David", "Mariana"],
        "anos_base": [2023, 2024, 2025],
        "declarations": declarations,
        "receipts": receipts,
        "properties": properties,
        "summary": {
            "total_declarations": len(declarations),
            "total_receipts": len(receipts),
            "total_properties": len(properties),
            "bens_direitos_count": sum(len(d.get("bens_direitos", [])) for d in declarations),
            "total_bens_value": sum(d.get("total_bens") or 0 for d in declarations)
        }
    }

    # Save consolidated baseline
    consolidated_path = OUTPUT_DIR / "baseline_patrimonial-1.5_consolidated.json"
    with open(consolidated_path, 'w', encoding='utf-8') as f:
        json.dump(consolidated, f, ensure_ascii=False, indent=2)

    print("\n" + "=" * 80)
    print("CONSOLIDATION SUMMARY")
    print("=" * 80)
    print(f"Total IRPF Declarations: {len(declarations)}")
    print(f"Total IRPF Receipts: {len(receipts)}")
    print(f"Total Properties: {len(properties)}")
    print(f"Total Bens e Direitos: {consolidated['summary']['bens_direitos_count']}")
    print(f"Total Bens Value: R$ {consolidated['summary']['total_bens_value']:,.2f}")
    print(f"\nConsolidated output: {consolidated_path.name}")
    print("=" * 80)

if __name__ == "__main__":
    main()
