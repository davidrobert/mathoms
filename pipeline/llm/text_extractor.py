"""DocumentTextExtractor — extract text from PDFs and spreadsheets for LLM input.

Supports: PDF (via pdfplumber), XLSX/XLS (via openpyxl), CSV (stdlib).
Images (JPG, JPEG, PNG) are returned as raw bytes via extract_image_bytes().
"""

from __future__ import annotations

import csv
import enum
import io
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png"}

_MEDIA_TYPE_MAP = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
}


class ReaderOutcome(str, enum.Enum):
    """Por que um documento não virou texto. Só ``documento_vazio`` não é defeito."""

    ok = "ok"
    leitor_ausente = "leitor_ausente"
    leitor_indisponivel = "leitor_indisponivel"
    leitura_falhou = "leitura_falhou"
    documento_vazio = "documento_vazio"


# O ``str`` cru não distinguia "não existe leitor para este formato" de "o
# documento está vazio" — as duas produziam ``""``, e só a primeira é defeito.
# Quem decide o balanço do fan-out precisa da diferença (ADR-393 D2).
@dataclass(frozen=True)
class TextExtraction:
    """Resultado da extração de texto: texto + motivo tipado."""

    outcome: ReaderOutcome
    text: str = ""
    detalhe: str = ""

    @property
    def is_defect(self) -> bool:
        return self.outcome not in (ReaderOutcome.ok, ReaderOutcome.documento_vazio)


def _sheet_lines(sheet_name: str, rows, *, budget: int) -> tuple[str, int]:
    """Uma aba → texto, parando quando o orçamento de chars acaba."""
    lines: list[str] = [f"=== Sheet: {sheet_name} ==="]
    usados = 0
    for row in rows:
        cells = ["" if c is None else str(c) for c in row]
        if any(cells):
            row_text = " | ".join(cells)
            lines.append(row_text)
            usados += len(row_text) + 1
        if usados > budget:
            lines.append("[... truncated ...]")
            break
    return "\n".join(lines), usados


# Fonte única de "este formato tem leitor" (ADR-393 D5). O E0 consulta a mesma
# constante para recusar na ENTRADA o que ninguém consegue ler — descobrir no
# meio do fan-out é tarde: o documento já foi aceito, contado e prometido.
_READER_BY_SUFFIX: dict[str, str] = {
    ".pdf": "_extract_pdf",
    ".xlsx": "_extract_xlsx",
    ".xls": "_extract_xls",
    ".csv": "_extract_csv",
    ".json": "_extract_plain",
    ".txt": "_extract_plain",
    ".md": "_extract_plain",
}

# Imagens não passam por `_reader_for` — vão como conteúdo multimodal via
# `extract_image_bytes`, e por isso contam como legíveis.
READABLE_SUFFIXES: frozenset[str] = frozenset(_READER_BY_SUFFIX) | frozenset(IMAGE_EXTENSIONS)


def _xls_cell_text(cell, datemode: int, xlrd) -> str:
    """Célula BIFF → texto. Data em serial float é ilegível para o LLM."""
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate.xldate_as_datetime(cell.value, datemode).isoformat(sep=" ")
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return ""
    if cell.ctype == xlrd.XL_CELL_NUMBER and float(cell.value).is_integer():
        return str(int(cell.value))
    return str(cell.value)


class DocumentTextExtractor:
    """Extracts text content from documents for use as LLM prompt input.

    Usage:
        extractor = DocumentTextExtractor(max_chars=50_000)
        text = extractor.extract(Path("extrato.pdf"))
    """

    def __init__(self, max_chars: int = 100_000, max_pages: int = 50):
        self.max_chars = max_chars
        self.max_pages = max_pages

    @staticmethod
    def is_image(path: Path) -> bool:
        return path.suffix.lower() in IMAGE_EXTENSIONS

    def extract_image_bytes(self, path: Path) -> tuple[bytes, str]:
        """Return raw bytes and MIME type for an image file."""
        media_type = _MEDIA_TYPE_MAP.get(path.suffix.lower(), "image/jpeg")
        return path.read_bytes(), media_type

    # Wrapper de compat: seis stages fora do escopo da A40.l68 ainda consomem
    # esta forma e herdam a cegueira que a ADR-393 §D2 declara (não conserta).
    def extract(self, path: Path) -> str:
        """Texto do documento, ou ``""`` — prefira ``extract_result``."""
        return self.extract_result(path).text

    def extract_result(self, path: Path) -> TextExtraction:
        """Texto + motivo tipado (ADR-393 D2)."""
        reader = self._reader_for(path.suffix.lower())
        if reader is None:
            return self._no_reader(path)
        try:
            text = reader(path)
        except Exception as exc:
            return self._read_failure(path, exc)
        return (
            TextExtraction(ReaderOutcome.ok, text=text)
            if text.strip()
            else TextExtraction(ReaderOutcome.documento_vazio, text=text)
        )

    def _no_reader(self, path: Path) -> TextExtraction:
        """Formato sem leitor — era o `""` mudo indistinguível de doc vazio."""
        suffix = path.suffix.lower()
        detalhe = f"imagem ({suffix})" if suffix in IMAGE_EXTENSIONS else (suffix or "sem extensão")
        logger.warning("text_extractor.reader_missing", extra={"suffix": suffix})
        return TextExtraction(ReaderOutcome.leitor_ausente, detalhe=detalhe)

    def _read_failure(self, path: Path, exc: Exception) -> TextExtraction:
        """Lib ausente vs. leitor que levantou — motivos distintos, nunca `""`."""
        detalhe = str(exc) if isinstance(exc, ImportError) else f"{type(exc).__name__}: {exc}"
        outcome = (
            ReaderOutcome.leitor_indisponivel
            if isinstance(exc, ImportError)
            else ReaderOutcome.leitura_falhou
        )
        logger.error("text_extractor.read_failed", extra={"file": path.name, "erro": detalhe})
        return TextExtraction(outcome, detalhe=detalhe)

    def _reader_for(self, suffix: str):
        """Leitor do formato, ou ``None`` quando não existe — o que era `""` mudo."""
        nome = _READER_BY_SUFFIX.get(suffix)
        return getattr(self, nome) if nome else None

    def _extract_plain(self, path: Path) -> str:
        return path.read_text(encoding="utf-8")[: self.max_chars]

    def _extract_pdf(self, path: Path) -> str:
        """Extract text from PDF using pdfplumber."""
        try:
            import pdfplumber
        except ImportError as exc:  # vira leitor_indisponivel em extract_result
            raise ImportError("pdfplumber não instalado — PDF ilegível") from exc

        pages_text: list[str] = []
        total_chars = 0

        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages):
                if i >= self.max_pages:
                    pages_text.append(f"\n[... truncated at {self.max_pages} pages ...]")
                    break

                text = page.extract_text() or ""

                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        table_text = self._format_table(table)
                        if table_text and table_text not in text:
                            text += "\n" + table_text

                if total_chars + len(text) > self.max_chars:
                    remaining = self.max_chars - total_chars
                    pages_text.append(text[:remaining])
                    pages_text.append(f"\n[... truncated at {total_chars + remaining} chars ...]")
                    break

                pages_text.append(text)
                total_chars += len(text)

        return "\n\n".join(pages_text)

    def _extract_xlsx(self, path: Path) -> str:
        """XLSX via openpyxl. O `.xls` legado tem leitor próprio — ver `_extract_xls`."""
        try:
            import openpyxl
        except ImportError as exc:  # vira leitor_indisponivel em extract_result
            raise ImportError("openpyxl não instalado — planilha ilegível") from exc

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            return self._sheets_to_text(
                (name, wb[name].iter_rows(values_only=True)) for name in wb.sheetnames
            )
        finally:
            wb.close()

    # openpyxl levanta `InvalidFileException` em BIFF e o motivo virava
    # `leitura_falhou` — 168/168 dos `.xls` do corpus (A40.l68 §Ataque C).
    # `xlrd` é o mesmo leitor que os parsers determinísticos já usam.
    def _extract_xls(self, path: Path) -> str:
        """`.xls` legado (BIFF) via xlrd — openpyxl não lê este formato."""
        try:
            import xlrd
        except ImportError as exc:  # vira leitor_indisponivel em extract_result
            raise ImportError("xlrd não instalado — .xls legado ilegível") from exc

        wb = xlrd.open_workbook(path)
        return self._sheets_to_text(
            (sheet.name, self._xls_rows(sheet, wb.datemode)) for sheet in wb.sheets()
        )

    @staticmethod
    def _xls_rows(sheet, datemode: int):
        """Linhas do BIFF já em texto — serial de data vira ISO, não `45678.0`."""
        import xlrd

        for i in range(sheet.nrows):
            yield [_xls_cell_text(c, datemode, xlrd) for c in sheet.row(i)]

    def _sheets_to_text(self, sheets) -> str:
        """Planilha → texto: cabeçalho por aba + linhas `a | b`, cortadas em max_chars."""
        sheets_text: list[str] = []
        total_chars = 0

        for sheet_name, rows in sheets:
            texto, usados = _sheet_lines(sheet_name, rows, budget=self.max_chars - total_chars)
            sheets_text.append(texto)
            total_chars += usados
            if total_chars > self.max_chars:
                break

        return "\n\n".join(sheets_text)

    def _extract_csv(self, path: Path) -> str:
        """Extract text from CSV file."""
        text = path.read_text(encoding="utf-8", errors="replace")
        if len(text) > self.max_chars:
            text = text[: self.max_chars] + "\n[... truncated ...]"
        return text

    @staticmethod
    def _format_table(table: list[list[Optional[str]]]) -> str:
        """Format a pdfplumber table as pipe-delimited text."""
        if not table:
            return ""
        rows = []
        for row in table:
            cells = [str(c).strip() if c else "" for c in row]
            rows.append(" | ".join(cells))
        return "\n".join(rows)

    def extract_multiple(self, paths: list[Path]) -> dict[str, str]:
        """Extract text from multiple files. Returns {filename: text}."""
        results = {}
        for path in paths:
            results[path.name] = self.extract(path)
        return results
