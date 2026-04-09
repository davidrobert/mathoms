#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
E2 Extrato Extraction - Deterministic parsers for bank statements.

Parsers determinísticos para extratos bancários. Segue a mesma arquitetura
do e2_extract_faturas.py — um parser por banco, roteamento por filename,
fallback LLM para bancos desconhecidos.

Bancos suportados (CSV — deterministic CSV parser):
  - C6 Bank: extratoconta, extratocontapj (ZIP-protected CSV from internet banking)

Bancos suportados (XLS_NATIVE — xlrd-based parser for .xls exports):
  - Itaú: extratoconta, extratocontapersonnalite (internet banking XLS export)

Bancos suportados (TABLE_READY — pdfplumber tables):
  - C6 Bank: extratoconta, extratocontapj, extratocontaglobalusd, extratocontaglobaleur
  - Itaú: extratoconta, extratocontapersonnalite (PDF fallback)
  - PicPay: extratoconta

Bancos suportados (TEXT_REGEX — regex sobre texto extraído):
  - Bradesco: extratoconta, extratopoupanca
  - Santander: extratoconta
  - BTG Pactual: extratoconta
  - Rico: extratoconta
  - Wise: extratocontausd, extratocontabrl
  - Bank of America: extratoconta

Usage:
    python scripts/e2_extract_extratos.py [--dry-run] [--file ARQUIVO.pdf|ARQUIVO.csv|ARQUIVO.xls]

Author: Claude Opus 4.6
"""

import calendar
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import pdfplumber
except ImportError:
    print("ERROR: pdfplumber not installed. Run: pip install pdfplumber", file=sys.stderr)
    sys.exit(1)


# =============================================================================
# Configuration — loaded from project config files, never hardcoded
# =============================================================================

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data" / "financial_statements"
OUTPUT_DIR = BASE_DIR / "processed" / "E2_extracts"
CONFIG_DIR = BASE_DIR / "config"


def _load_family_config() -> dict:
    """Load family_members.json for name matching (used in header parsing)."""
    fm_path = CONFIG_DIR / "family_members.json"
    if fm_path.exists():
        with open(fm_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


_FAMILY = _load_family_config()
_MEMBROS = _FAMILY.get("membros", {})

# Build a lookup of all known member names and CPFs from config
_MEMBER_NAMES: List[str] = []
_MEMBER_CPFS: Dict[str, str] = {}  # cpf → member_id
for _mid, _mdata in _MEMBROS.items():
    for variant in _mdata.get("variantes_nome", []):
        _MEMBER_NAMES.append(variant)
    cpf = _mdata.get("cpf", "")
    if cpf:
        _MEMBER_CPFS[cpf] = _mid

# Meses PT-BR → número
MESES_BR = {
    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4,
    'mai': 5, 'jun': 6, 'jul': 7, 'ago': 8,
    'set': 9, 'out': 10, 'nov': 11, 'dez': 12,
    'janeiro': 1, 'fevereiro': 2, 'março': 3, 'marco': 3, 'abril': 4,
    'maio': 5, 'junho': 6, 'julho': 7, 'agosto': 8,
    'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12,
}

# Extrato file patterns recognized by deterministic parsers
# Maps (banco_prefix, tipo_contains) → parser function name
# Populated at module level after parser definitions

# =============================================================================
# Logging
# =============================================================================

_VERBOSE = True


def log(level: str, msg: str) -> None:
    if not _VERBOSE and level == "DEBUG":
        return
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] E2-EXTRATO {level}: {msg}", file=sys.stderr)


# =============================================================================
# Utility helpers (shared with e2_extract_faturas.py patterns)
# =============================================================================

def parse_brl(text: str) -> Optional[float]:
    """Parse Brazilian currency string to float. '1.234,56' → 1234.56, '-R$ 98,00' → -98.0"""
    if not text:
        return None
    text = str(text).strip()
    # Remove currency symbols
    for sym in ("R$", "US$", "EUR", "USD", "BRL", "$"):
        text = text.replace(sym, "")
    text = text.strip()
    if not text or text == "-":
        return None

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1].strip()
    elif text.startswith("-") or text.startswith("(-"):
        negative = True
        text = text.lstrip("(-").rstrip(")").strip()

    # Brazilian format: 1.234,56 → remove dots, comma→dot
    text = text.replace(".", "").replace(",", ".")
    try:
        val = float(text)
        return -val if negative else val
    except ValueError:
        return None


def parse_usd(text: str) -> Optional[float]:
    """Parse US currency string to float. '2,605.00' → 2605.0, '-$150.25' → -150.25.

    US format: comma = thousands separator, period = decimal.
    This is the inverse of parse_brl (Brazilian format).
    Used for Bank of America and other US bank statements.
    """
    if not text:
        return None
    text = str(text).strip()
    # Remove currency symbols
    for sym in ("US$", "USD", "$"):
        text = text.replace(sym, "")
    text = text.strip()
    if not text or text == "-":
        return None

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1].strip()
    elif text.startswith("-") or text.startswith("(-"):
        negative = True
        text = text.lstrip("(-").rstrip(")").strip()

    # US format: 2,605.00 → remove commas, period stays as decimal
    text = text.replace(",", "")
    try:
        val = float(text)
        return -val if negative else val
    except ValueError:
        return None


def safe_date(year: int, month: int, day: int) -> str:
    """Return valid ISO date string, adjusting day if necessary."""
    year = max(1900, min(2100, year))
    if month < 1 or month > 12:
        log("WARN", f"  Mês inválido: {year}-{month:02d}-{day:02d} → mês 01")
        month = 1
    max_day = calendar.monthrange(year, month)[1]
    if day > max_day:
        day = max_day
    if day < 1:
        day = 1
    return f"{year}-{month:02d}-{day:02d}"


def infer_periodo_from_filename(filename: str) -> Tuple[Optional[str], Optional[str]]:
    """Extract periodo start/end from filename patterns like _202501_202512 or _202603."""
    # Pattern: banco_tipo_YYYYMM_YYYYMM-0_original.pdf
    m = re.search(r'_(\d{4})(\d{2})_(\d{4})(\d{2})', filename)
    if m:
        y1, m1, y2, m2 = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
        inicio = safe_date(y1, m1, 1)
        fim_day = calendar.monthrange(y2, m2)[1]
        fim = safe_date(y2, m2, fim_day)
        return inicio, fim

    # Pattern: banco_tipo_YYYYMM-0_original.pdf (single month)
    m = re.search(r'_(\d{4})(\d{2})(?:[a-z])?-', filename)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        inicio = safe_date(y, mo, 1)
        fim_day = calendar.monthrange(y, mo)[1]
        fim = safe_date(y, mo, fim_day)
        return inicio, fim

    return None, None


def detect_member_from_text(text: str) -> Optional[str]:
    """Try to identify which family member owns this statement, using config names/CPFs."""
    text_upper = text.upper()
    for cpf, mid in _MEMBER_CPFS.items():
        if cpf in text:
            return mid
    for mid, mdata in _MEMBROS.items():
        for variant in mdata.get("variantes_nome", []):
            if variant.upper() in text_upper:
                return mid
    return None


def extract_account_number(text: str, banco: str) -> Optional[str]:
    """Extract account number from statement text using common patterns."""
    patterns = [
        r'[Cc]onta[:\s]+(\d[\d.\-/]+\d)',
        r'[Aa]gência[:\s]+\d+\s*[\|/•]\s*[Cc]onta[:\s]+(\d[\d.\-]+\d)',
        r'Account\s*(?:number|#)?[:\s]+(\d[\d\s]+\d)',
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            return m.group(1).strip()
    return None


def make_result_template(banco: str, tipo: str, moeda: str = "BRL") -> Dict[str, Any]:
    """Create a standard E2 result dictionary."""
    return {
        "banco": banco,
        "tipo": tipo,
        "moeda": moeda,
        "numero_conta": None,
        "titular": None,
        "periodo": {"inicio": None, "fim": None},
        "saldo_inicial": None,
        "saldo_final": None,
        "transacoes": [],
        "notas": [],
    }


def resolve_year_from_period(dd: int, mm: int, periodo_inicio: str, periodo_fim: str) -> int:
    """Given a transaction DD/MM, resolve which year it belongs to based on periodo."""
    if not periodo_inicio:
        return datetime.now().year
    start_year = int(periodo_inicio[:4])
    end_year = int(periodo_fim[:4]) if periodo_fim else start_year
    if start_year == end_year:
        return start_year
    # If the month is >= start month of start_year, use start_year
    start_month = int(periodo_inicio[5:7])
    if mm >= start_month:
        return start_year
    return end_year


# =============================================================================
# Parser: C6 Bank CSV (extratoconta, extratocontapj)
# CSV export from C6 Bank internet banking — ZIP-protected, BOM-prefixed
# Columns: Data Lançamento, Data Contábil, Título, Descrição, Entrada(R$), Saída(R$), Saldo do Dia(R$)
# =============================================================================

def parse_c6bank_csv(csv_path: Path, filename: str) -> Dict[str, Any]:
    """Parse C6 Bank CSV statement (conta or contapj).

    CSV structure:
      - BOM (UTF-8) header
      - Lines 1-2: "EXTRATO DE CONTA CORRENTE C6 BANK" + blank
      - Line 3: "Agência: X / Conta: NNNNNNNNN"
      - Line 4: "Extrato gerado em DD/MM/YYYY - as HH:MM:SS"
      - Line 5: blank
      - Line 6: "Extrato de DD/MM/YYYY a DD/MM/YYYY"
      - Line 7: blank
      - Line 8: header row (comma-separated)
      - Lines 9+: transaction data
    """
    import csv as csv_mod

    is_pj = "extratocontapj" in filename
    tipo = "extratocontapj" if is_pj else "extratoconta"
    moeda = "BRL"

    log("INFO", f"Parsing C6 Bank CSV ({tipo}): {filename}")
    result = make_result_template("C6 Bank", tipo, moeda)

    # Read file with BOM handling
    raw_text = csv_path.read_text(encoding="utf-8-sig")
    lines = raw_text.splitlines()

    # --- Parse header metadata ---
    # Account number: "Agência: 1 / Conta: 130952222"
    for line in lines[:6]:
        conta_m = re.search(r'Conta:\s*(\d+)', line)
        if conta_m:
            result["numero_conta"] = conta_m.group(1)
            break

    # Periodo: "Extrato de DD/MM/YYYY a DD/MM/YYYY"
    for line in lines[:10]:
        periodo_m = re.search(
            r'Extrato de\s+(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})', line
        )
        if periodo_m:
            d1 = datetime.strptime(periodo_m.group(1), "%d/%m/%Y")
            d2 = datetime.strptime(periodo_m.group(2), "%d/%m/%Y")
            result["periodo"]["inicio"] = d1.strftime("%Y-%m-%d")
            result["periodo"]["fim"] = d2.strftime("%Y-%m-%d")
            break

    # If periodo not found in header, try to infer from filename
    if not result["periodo"]["inicio"]:
        p_ini, p_fim = infer_periodo_from_filename(filename)
        result["periodo"]["inicio"] = p_ini
        result["periodo"]["fim"] = p_fim

    # Detect member from header text (first 6 lines)
    header_text = "\n".join(lines[:6])
    result["titular"] = detect_member_from_text(header_text)

    # --- Find the CSV header row and parse transactions ---
    csv_header_idx = None
    for i, line in enumerate(lines):
        if line.strip().startswith("Data Lançamento,") or line.strip().startswith("Data Lancamento,"):
            csv_header_idx = i
            break

    if csv_header_idx is None:
        result["notas"].append("WARN: Header CSV 'Data Lançamento,...' não encontrado")
        # File might be empty (header-only, no transactions)
        return result

    # Parse CSV data from header row onwards
    csv_text = "\n".join(lines[csv_header_idx:])
    reader = csv_mod.reader(csv_text.splitlines())
    header = next(reader, None)  # consume header row

    if not header:
        result["notas"].append("WARN: Header CSV vazio")
        return result

    # Normalize header names for robustness
    header_clean = [h.strip().lower() for h in header]
    # Expected: ['data lançamento', 'data contábil', 'título', 'descrição', 'entrada(r$)', 'saída(r$)', 'saldo do dia(r$)']

    saldo_first = None
    saldo_last = None

    for row in reader:
        if len(row) < 6:
            continue  # skip malformed rows

        # Pad to 7 columns if saldo column is missing
        while len(row) < 7:
            row.append("")

        data_lanc_str = row[0].strip()
        data_contabil_str = row[1].strip()
        titulo = row[2].strip()
        descricao = row[3].strip()
        entrada_str = row[4].strip()
        saida_str = row[5].strip()
        saldo_str = row[6].strip()

        # Validate date format DD/MM/YYYY
        if not re.match(r'\d{2}/\d{2}/\d{4}$', data_lanc_str):
            continue  # skip non-transaction rows

        # Parse date
        try:
            dt = datetime.strptime(data_lanc_str, "%d/%m/%Y")
            data_iso = dt.strftime("%Y-%m-%d")
        except ValueError:
            log("WARN", f"  Data inválida: {data_lanc_str}")
            continue

        # Parse values — CSV uses "0.00" format (dot as decimal separator)
        entrada = _parse_csv_number(entrada_str)
        saida = _parse_csv_number(saida_str)

        # Calculate valor: positive for entrada, negative for saida
        if entrada and entrada > 0:
            valor = entrada
        elif saida and saida > 0:
            valor = -saida
        else:
            valor = 0.0

        # Build description: prefer 'titulo' but append 'descricao' if different
        if descricao and descricao != titulo:
            desc_full = f"{titulo} — {descricao}" if titulo else descricao
        else:
            desc_full = titulo or descricao or ""

        # Determine tipo_lancamento from titulo/descricao
        tipo_lanc = _classify_c6_csv_lancamento(titulo, descricao)

        tx = {
            "data": data_iso,
            "descricao": desc_full,
            "valor": valor,
            "tipo_lancamento": tipo_lanc,
        }

        result["transacoes"].append(tx)

        # Track saldo
        saldo_val = _parse_csv_number(saldo_str)
        if saldo_val is not None:
            if saldo_first is None:
                saldo_first = saldo_val
            saldo_last = saldo_val

    # Set saldo_inicial and saldo_final
    result["saldo_final"] = saldo_last

    # Saldo_inicial: the saldo of the first transaction represents the balance
    # AFTER that transaction. To get saldo_inicial, we need first_saldo - first_valor.
    if saldo_first is not None and result["transacoes"]:
        first_valor = result["transacoes"][0].get("valor", 0) or 0
        result["saldo_inicial"] = round(saldo_first - first_valor, 2)
    else:
        result["saldo_inicial"] = saldo_first

    n_tx = len(result["transacoes"])
    log("INFO", f"  Extraídas {n_tx} transações do CSV")
    if result["saldo_inicial"] is not None:
        log("INFO", f"  Saldo inicial: {result['saldo_inicial']:.2f}")
    if result["saldo_final"] is not None:
        log("INFO", f"  Saldo final: {result['saldo_final']:.2f}")

    return result


def _parse_csv_number(text: str) -> Optional[float]:
    """Parse a number from C6 CSV format. Handles '1234.56', '-1234.56', empty strings."""
    if not text or not text.strip():
        return None
    text = text.strip().replace(",", "")  # in case of thousands separators
    try:
        return float(text)
    except ValueError:
        # Try Brazilian format as fallback (1.234,56)
        return parse_brl(text)


def _classify_c6_csv_lancamento(titulo: str, descricao: str) -> str:
    """Classify a C6 CSV transaction into tipo_lancamento based on titulo/descricao."""
    combined = f"{titulo} {descricao}".lower()

    if "pix enviado" in combined:
        return "Saída PIX"
    elif "pix recebido" in combined:
        return "Entrada PIX"
    elif "devol recebida pix" in combined or "devol enviada pix" in combined:
        return "Devolução PIX"
    elif "ted enviada" in combined or "transf enviada" in combined:
        return "Saída TED/Transferência"
    elif "ted recebida" in combined or "transf recebida" in combined:
        return "Entrada TED/Transferência"
    elif "c6tag" in combined:
        return "C6 Tag (Pedágio/Estacionamento)"
    elif "boleto" in combined or "guia" in combined:
        return "Pagamento Boleto"
    elif "juros" in combined or "iof" in combined:
        return "Encargos Bancários"
    elif "rendimento" in combined or "aplicação" in combined or "aplicacao" in combined:
        return "Investimento/Rendimento"
    elif "resgate" in combined:
        return "Resgate Investimento"
    elif "salário" in combined or "salario" in combined:
        return "Salário"
    elif "13" in titulo and "salário" in combined.replace("á", "a"):
        return "13º Salário"
    elif "compra" in combined or "débito" in combined or "debito" in combined:
        return "Compra/Débito"
    else:
        return "Outros"


# =============================================================================
# Parser: C6 Bank (extratoconta, extratocontapj, extratocontaglobal*)
# TABLE_READY — pdfplumber extract_tables() works well
# =============================================================================

def parse_c6bank(pdf_path: Path, filename: str) -> Dict[str, Any]:
    """Parse C6 Bank statement (conta, contapj, contaglobal)."""
    # Detect subtype from filename
    is_global_usd = "extratocontaglobalusd" in filename
    is_global_eur = "extratocontaglobaleur" in filename
    is_pj = "extratocontapj" in filename

    if is_global_usd:
        moeda = "USD"
        tipo = "extratocontaglobalusd"
    elif is_global_eur:
        moeda = "EUR"
        tipo = "extratocontaglobaleur"
    elif is_pj:
        moeda = "BRL"
        tipo = "extratocontapj"
    else:
        moeda = "BRL"
        tipo = "extratoconta"

    log("INFO", f"Parsing C6 Bank ({tipo}): {filename}")
    result = make_result_template("C6 Bank", tipo, moeda)

    periodo_inicio, periodo_fim = infer_periodo_from_filename(filename)
    result["periodo"]["inicio"] = periodo_inicio
    result["periodo"]["fim"] = periodo_fim

    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Extract header info from first page text
            first_text = pdf.pages[0].extract_text() or ""
            result["titular"] = detect_member_from_text(first_text)
            result["numero_conta"] = extract_account_number(first_text, "c6bank")

            # Parse periodo from header text (more precise than filename)
            # Pattern: "Período • DD de MÊS de YYYY até DD de MÊS de YYYY"
            periodo_pat = re.compile(
                r'Período\s*•?\s*(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})\s+'
                r'até\s+(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})',
                re.IGNORECASE
            )
            pm = periodo_pat.search(first_text)
            if pm:
                d1 = int(pm.group(1))
                m1 = MESES_BR.get(pm.group(2).lower(), 0)
                y1 = int(pm.group(3))
                d2 = int(pm.group(4))
                m2 = MESES_BR.get(pm.group(5).lower(), 0)
                y2 = int(pm.group(6))
                if m1 and m2:
                    result["periodo"]["inicio"] = safe_date(y1, m1, d1)
                    result["periodo"]["fim"] = safe_date(y2, m2, d2)

            # Extract saldo from header
            # Pattern: "Saldo do dia • DD de MÊS de YYYY • R$ 6.930,11"
            saldo_header = re.search(
                r'Saldo do dia.*?[•\s]+(R\$|US\$|EUR)\s*([\d.,]+)',
                first_text
            )

            # Check for "Sem lançamentos no mês" (empty period)
            full_text = "\n".join(
                (p.extract_text() or "") for p in pdf.pages
            )
            if "Sem lançamentos no mês" in full_text or "sem lançamentos" in full_text.lower():
                empty_months = full_text.lower().count("sem lançamentos")
                result["notas"].append(
                    f"Sem lançamentos no período ({empty_months} mês(es) sem movimentação)"
                )

            # Parse all tables across all pages
            all_rows: List[Tuple[str, str, str, str, str]] = []
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if row and len(row) >= 4:
                            all_rows.append(tuple(str(c) if c else "" for c in row))

            # C6 Bank has two table formats:
            # Conta/PJ: [data_lancamento, data_contabil, tipo, descricao, valor_BRL]
            # Global:    [data, tipo, descricao, valor_FX, autorizacao]
            # The Global format has currency prefixes in valor (e.g. "-US$ 25,38")
            # and an authorization column instead of data_contabil
            is_global = is_global_usd or is_global_eur

            pending_tx: Optional[Dict] = None
            saldo_values: List[Tuple[str, float]] = []

            for row in all_rows:
                # Normalize: ensure 5 columns
                cols = list(row) + [""] * (5 - len(row))
                col0, col1, col2, col3, col4 = cols[:5]

                # For Global format, valor is in col3 (with currency prefix)
                # For Conta/PJ format, valor is in col4
                if is_global:
                    valor_col = col3
                    desc_col = col2
                    tipo_col = col1
                else:
                    valor_col = col4
                    desc_col = col3
                    tipo_col = col2

                # Skip empty/header rows
                if not any(c.strip() for c in cols):
                    continue

                # Detect "Saldo do dia" rows
                saldo_match = re.match(r'Saldo do dia\s+(\d{2}/\d{2}/\d{2,4})', col0)
                if saldo_match:
                    saldo_val = parse_brl(col4) or parse_brl(col3)
                    if saldo_val is not None:
                        date_str = saldo_match.group(1)
                        saldo_values.append((date_str, saldo_val))
                    # Flush pending tx
                    if pending_tx:
                        result["transacoes"].append(pending_tx)
                        pending_tx = None
                    continue

                # Transaction row: has date in col0 (DD/MM) and value in valor_col
                date_match = re.match(r'(\d{2}/\d{2})', col0.strip())
                has_value = valor_col.strip() and parse_brl(valor_col) is not None

                if date_match and has_value:
                    # Flush previous pending
                    if pending_tx:
                        result["transacoes"].append(pending_tx)

                    dd, mm_str = date_match.group(1).split("/")
                    dd_i, mm_i = int(dd), int(mm_str)
                    year = resolve_year_from_period(
                        dd_i, mm_i,
                        result["periodo"]["inicio"] or "",
                        result["periodo"]["fim"] or ""
                    )
                    valor = parse_brl(valor_col)

                    pending_tx = {
                        "data": safe_date(year, mm_i, dd_i),
                        "descricao": desc_col.strip(),
                        "valor": valor,
                        "tipo_lancamento": tipo_col.strip() if tipo_col.strip() else None,
                    }
                    continue

                if date_match and not has_value:
                    # Row with date but no value — next row likely has value
                    if pending_tx:
                        result["transacoes"].append(pending_tx)

                    dd, mm_str = date_match.group(1).split("/")
                    dd_i, mm_i = int(dd), int(mm_str)
                    year = resolve_year_from_period(
                        dd_i, mm_i,
                        result["periodo"]["inicio"] or "",
                        result["periodo"]["fim"] or ""
                    )
                    pending_tx = {
                        "data": safe_date(year, mm_i, dd_i),
                        "descricao": desc_col.strip(),
                        "valor": None,
                        "tipo_lancamento": tipo_col.strip() if tipo_col.strip() else None,
                    }
                    continue

                # Continuation row (no date in col0)
                if not date_match and (tipo_col.strip() or desc_col.strip()):
                    val = parse_brl(valor_col)
                    if pending_tx and pending_tx["valor"] is None and val is not None:
                        # This row carries the value for the pending tx
                        pending_tx["valor"] = val
                        if desc_col.strip() and not pending_tx["descricao"]:
                            pending_tx["descricao"] = desc_col.strip()
                        result["transacoes"].append(pending_tx)
                        pending_tx = None
                    elif val is not None:
                        # Standalone continuation with value — new transaction inheriting date
                        if pending_tx:
                            result["transacoes"].append(pending_tx)
                        prev_date = result["transacoes"][-1]["data"] if result["transacoes"] else None
                        pending_tx = None
                        result["transacoes"].append({
                            "data": prev_date,
                            "descricao": desc_col.strip(),
                            "valor": val,
                            "tipo_lancamento": tipo_col.strip() if tipo_col.strip() else None,
                        })
                    elif pending_tx and desc_col.strip():
                        # Description continuation
                        pending_tx["descricao"] += " " + desc_col.strip()

            # Flush last pending
            if pending_tx:
                result["transacoes"].append(pending_tx)

            # Remove transactions with None valor
            result["transacoes"] = [t for t in result["transacoes"] if t.get("valor") is not None]

            # Derive saldo_inicial and saldo_final from saldo rows
            if saldo_values:
                result["saldo_inicial"] = saldo_values[0][1]
                result["saldo_final"] = saldo_values[-1][1]

    except Exception as e:
        log("ERROR", f"  Falha ao processar {filename}: {e}")
        result["notas"].append(f"Erro no parsing: {e}")
        result["requires_llm_fallback"] = True

    log("INFO", f"  → {len(result['transacoes'])} transações extraídas")
    return result


# =============================================================================
# Parser: Itaú XLS (extratoconta, extratocontapersonnalite)
# XLS_NATIVE — xlrd-based parser for .xls exports from Itaú internet banking
#
# Structure (Sheet "Lançamentos"):
#   Row 0: "Logotipo Itaú"
#   Row 1: "Atualização:" + timestamp
#   Row 2: "Nome:" + account holder
#   Row 3: "Agência:" + number (float)
#   Row 4: "Conta:" + "NNNNN-D"
#   Row 8: headers [data, lançamento, ag./origem, valor (R$), saldos (R$)]
#   Row 9: "lançamentos" (section marker)
#   Row 10+: data — two types:
#     - Transaction: [date, description, "", value_float, ""]
#     - SALDO ANTERIOR: [date, "SALDO ANTERIOR", "", "", saldo_float]
#     - SALDO TOTAL DISPONÍVEL DIA: [date, text, "", "", saldo_float]
#   Near end: "lançamentos futuros" / "saídas futuras" (stop parsing)
#
# Sheet "Posição Consolidada": account limits and investment summary
# Sheet "Limites": cheque especial details
# =============================================================================

def _fix_itau_xls_encoding(text: str) -> str:
    """Fix mojibake in Itaú XLS files (UTF-8 decoded as latin-1)."""
    if not text or not isinstance(text, str):
        return text or ""
    # Common mojibake patterns from Itaú XLS:
    # Ã\x8d → Í, Ã\xad → í, Ã§ → ç, Ã£ → ã, Ã© → é, Ãº → ú, Ã³ → ó, Ã¡ → á, Ãª → ê, Ã\x94 → Ô
    try:
        # Try to fix double-encoding: encode back to latin-1, then decode as utf-8
        fixed = text.encode('latin-1').decode('utf-8')
        return fixed
    except (UnicodeDecodeError, UnicodeEncodeError):
        return text


def parse_itau_xls(xls_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Itaú XLS bank statement exported from internet banking.

    Supports the standard Itaú XLS format with sheets:
    - Lançamentos (transactions)
    - Posição Consolidada (balance summary)
    - Limites (overdraft limits)
    """
    try:
        import xlrd
    except ImportError:
        log("ERROR", "xlrd not installed. Run: pip install xlrd")
        raise SystemExit("FATAL: xlrd is required for Itaú XLS extraction. Install with: pip install xlrd")

    is_personnalite = "personnalite" in filename.lower()
    tipo = "extratocontapersonnalite" if is_personnalite else "extratoconta"

    log("INFO", f"Parsing Itaú XLS ({tipo}): {filename}")
    result = make_result_template("Itaú", tipo, "BRL")

    periodo_inicio, periodo_fim = infer_periodo_from_filename(filename)
    result["periodo"]["inicio"] = periodo_inicio
    result["periodo"]["fim"] = periodo_fim

    try:
        wb = xlrd.open_workbook(xls_path)

        # --- Sheet: Lançamentos ---
        if "Lançamentos" not in wb.sheet_names():
            # Try alternative names
            sheet_names_lower = {s.lower(): s for s in wb.sheet_names()}
            lancamentos_name = sheet_names_lower.get("lançamentos") or sheet_names_lower.get("lancamentos")
            if not lancamentos_name:
                log("WARN", f"  Sheet 'Lançamentos' não encontrada em {filename}")
                result["notas"].append("Sheet Lançamentos não encontrada")
                result["requires_llm_fallback"] = True
                return result
        else:
            lancamentos_name = "Lançamentos"

        sh = wb.sheet_by_name(lancamentos_name)

        # --- Extract header info (rows 0-4) ---
        if sh.nrows >= 5:
            # Nome do titular (row 2, col 1)
            nome_raw = str(sh.cell(2, 1).value).strip()
            nome = _fix_itau_xls_encoding(nome_raw)
            result["titular"] = detect_member_from_text(nome)

            # Agência (row 3, col 1)
            ag_val = sh.cell(3, 1).value
            agencia = str(int(ag_val)) if isinstance(ag_val, float) else str(ag_val).strip()

            # Conta (row 4, col 1)
            conta_val = str(sh.cell(4, 1).value).strip()
            result["numero_conta"] = conta_val

        # --- Parse transactions ---
        saldo_anterior = None
        saldo_final = None
        in_future_section = False
        first_tx_date = None
        last_tx_date = None

        for r in range(10, sh.nrows):
            # Read cells
            cell_date = str(sh.cell(r, 0).value).strip() if sh.ncols > 0 else ""
            cell_desc = str(sh.cell(r, 1).value).strip() if sh.ncols > 1 else ""
            cell_valor = sh.cell(r, 3).value if sh.ncols > 3 else ""
            cell_saldo = sh.cell(r, 4).value if sh.ncols > 4 else ""

            # Fix encoding on description
            cell_desc = _fix_itau_xls_encoding(cell_desc)

            # Detect future transactions section — stop parsing
            desc_lower = cell_desc.lower()
            date_lower = cell_date.lower()
            if ("lançamentos futuros" in date_lower or "lançamentos futuros" in desc_lower
                    or "lancamentos futuros" in date_lower or "lancamentos futuros" in desc_lower
                    or "saídas futuras" in date_lower or "saidas futuras" in date_lower):
                in_future_section = True
                continue

            if in_future_section:
                continue

            # Skip section markers and empty rows
            if not cell_date or cell_date.lower() in ("lançamentos", "lancamentos", ""):
                continue

            # Must have a valid date DD/MM/YYYY
            date_match = re.match(r'(\d{2})/(\d{2})/(\d{4})', cell_date)
            if not date_match:
                continue

            dd, mm, yyyy = date_match.group(1), date_match.group(2), date_match.group(3)
            iso_date = f"{yyyy}-{mm}-{dd}"

            # --- SALDO ANTERIOR ---
            if "SALDO ANTERIOR" in cell_desc.upper():
                saldo_val = cell_saldo if isinstance(cell_saldo, (int, float)) and cell_saldo != "" else None
                if saldo_val is not None and saldo_val != "":
                    saldo_anterior = float(saldo_val)
                    first_tx_date = iso_date
                continue

            # --- SALDO TOTAL DISPONÍVEL DIA ---
            desc_upper = cell_desc.upper()
            if ("SALDO TOTAL DISPON" in desc_upper or "SALDO DO DIA" in desc_upper
                    or "SALDO TOTAL DISPONÍVEL DIA" in desc_upper
                    or "SALDO TOTAL DISPONIVEL DIA" in desc_upper):
                saldo_val = cell_saldo if isinstance(cell_saldo, (int, float)) and cell_saldo != "" else None
                if saldo_val is not None and saldo_val != "":
                    saldo_final = float(saldo_val)
                    last_tx_date = iso_date
                continue

            # --- Regular transaction ---
            if isinstance(cell_valor, (int, float)) and cell_valor != "":
                valor = float(cell_valor)
            elif isinstance(cell_valor, str) and cell_valor.strip():
                valor = parse_brl(cell_valor)
            else:
                continue

            if valor is None:
                continue

            result["transacoes"].append({
                "data": iso_date,
                "descricao": cell_desc,
                "valor": valor,
            })

            if first_tx_date is None:
                first_tx_date = iso_date
            last_tx_date = iso_date

        # --- Derive saldos ---
        if saldo_anterior is not None:
            result["saldo_inicial"] = saldo_anterior
        if saldo_final is not None:
            result["saldo_final"] = saldo_final

        # --- Derive periodo from actual transaction dates ---
        if first_tx_date and (not result["periodo"]["inicio"] or result["periodo"]["inicio"] > first_tx_date):
            result["periodo"]["inicio"] = first_tx_date
        if last_tx_date and (not result["periodo"]["fim"] or result["periodo"]["fim"] < last_tx_date):
            result["periodo"]["fim"] = last_tx_date

        # --- Sheet: Posição Consolidada (optional enrichment) ---
        if "Posição Consolidada" in wb.sheet_names():
            try:
                sh_pos = wb.sheet_by_name("Posição Consolidada")
                for r in range(8, sh_pos.nrows):
                    desc = str(sh_pos.cell(r, 0).value).strip().lower()
                    val = sh_pos.cell(r, 3).value if sh_pos.ncols > 3 else ""
                    if "(=) saldo total disponível" in desc and isinstance(val, (int, float)):
                        # Cross-validate with saldo_final
                        if result["saldo_final"] is None:
                            result["saldo_final"] = float(val)
            except Exception:
                pass  # Posição Consolidada is optional

    except Exception as e:
        log("ERROR", f"  Falha ao processar XLS {filename}: {e}")
        result["notas"].append(f"Erro no parsing XLS: {e}")
        result["requires_llm_fallback"] = True

    log("INFO", f"  → {len(result['transacoes'])} transações extraídas do XLS")
    return result


# =============================================================================
# Parser: Itaú PDF (extratoconta, extratocontapersonnalite)
# TABLE_READY — many small 1-row tables per page
# =============================================================================

def parse_itau(pdf_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Itaú bank statement."""
    is_personnalite = "personnalite" in filename.lower()
    tipo = "extratocontapersonnalite" if is_personnalite else "extratoconta"

    log("INFO", f"Parsing Itaú ({tipo}): {filename}")
    result = make_result_template("Itaú", tipo, "BRL")

    periodo_inicio, periodo_fim = infer_periodo_from_filename(filename)
    result["periodo"]["inicio"] = periodo_inicio
    result["periodo"]["fim"] = periodo_fim

    try:
        with pdfplumber.open(pdf_path) as pdf:
            first_text = pdf.pages[0].extract_text() or ""
            result["titular"] = detect_member_from_text(first_text)
            result["numero_conta"] = extract_account_number(first_text, "itau")

            # Parse periodo from header: "Período: DD/MM/YYYY a DD/MM/YYYY"
            pm = re.search(r'Per[ií]odo[:\s]+(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})', first_text)
            if pm:
                parts1 = pm.group(1).split("/")
                parts2 = pm.group(2).split("/")
                result["periodo"]["inicio"] = f"{parts1[2]}-{parts1[1]}-{parts1[0]}"
                result["periodo"]["fim"] = f"{parts2[2]}-{parts2[1]}-{parts2[0]}"

            # Itaú produces many small tables, each typically 1 row
            # Format: [data, lancamentos, valor, saldo]
            # "SALDO DO DIA" entries have empty valor and saldo in col 3
            all_tables: List[list] = []
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    all_tables.extend(table)

            saldo_values: List[Tuple[str, float]] = []

            for row in all_tables:
                if not row or len(row) < 3:
                    continue
                cols = [str(c).strip() if c else "" for c in row]

                # Pad to 4 columns
                while len(cols) < 4:
                    cols.append("")

                date_str, descricao, valor_str, saldo_str = cols[0], cols[1], cols[2], cols[3]

                # Skip header row
                if date_str.lower() in ("data", ""):
                    if descricao.lower() in ("lançamentos", "lancamentos", ""):
                        continue

                # Skip empty rows
                if not date_str and not descricao:
                    continue

                # Parse date
                date_match = re.match(r'(\d{2}/\d{2}/\d{4})', date_str)
                if not date_match:
                    continue

                parts = date_match.group(1).split("/")
                iso_date = f"{parts[2]}-{parts[1]}-{parts[0]}"

                # "SALDO DO DIA" rows
                if "SALDO DO DIA" in descricao.upper():
                    saldo_val = parse_brl(saldo_str) or parse_brl(valor_str)
                    if saldo_val is not None:
                        saldo_values.append((iso_date, saldo_val))
                    continue

                # Regular transaction
                valor = parse_brl(valor_str)
                if valor is None:
                    continue

                result["transacoes"].append({
                    "data": iso_date,
                    "descricao": descricao,
                    "valor": valor,
                })

            # Derive saldos
            if saldo_values:
                # Sort by date
                saldo_values.sort(key=lambda x: x[0])
                result["saldo_inicial"] = saldo_values[0][1]
                result["saldo_final"] = saldo_values[-1][1]

    except Exception as e:
        log("ERROR", f"  Falha ao processar {filename}: {e}")
        result["notas"].append(f"Erro no parsing: {e}")
        result["requires_llm_fallback"] = True

    log("INFO", f"  → {len(result['transacoes'])} transações extraídas")
    return result


# =============================================================================
# Parser: PicPay (extratoconta)
# TABLE_READY — perfect 5-column tables
# =============================================================================

def parse_picpay(pdf_path: Path, filename: str) -> Dict[str, Any]:
    """Parse PicPay bank statement."""
    log("INFO", f"Parsing PicPay: {filename}")
    result = make_result_template("PicPay", "extratoconta", "BRL")

    periodo_inicio, periodo_fim = infer_periodo_from_filename(filename)
    result["periodo"]["inicio"] = periodo_inicio
    result["periodo"]["fim"] = periodo_fim

    try:
        with pdfplumber.open(pdf_path) as pdf:
            first_text = pdf.pages[0].extract_text() or ""
            result["titular"] = detect_member_from_text(first_text)

            # Account number
            m = re.search(r'Conta[:\s]+(\d+)', first_text)
            if m:
                result["numero_conta"] = m.group(1)

            # Periodo from header: "MOVIMENTAÇÕES DD DE MÊS DE YYYY A DD DE MÊS DE YYYY"
            pm = re.search(
                r'MOVIMENTA[ÇC][ÕO]ES\s+(\d{1,2})\s+DE\s+(\w+)\s+DE\s+(\d{4})\s+A\s+'
                r'(\d{1,2})\s+DE\s+(\w+)\s+DE\s+(\d{4})',
                first_text, re.IGNORECASE
            )
            if pm:
                d1, m1_name, y1 = int(pm.group(1)), pm.group(2).lower(), int(pm.group(3))
                d2, m2_name, y2 = int(pm.group(4)), pm.group(5).lower(), int(pm.group(6))
                m1 = MESES_BR.get(m1_name, 0)
                m2 = MESES_BR.get(m2_name, 0)
                if m1 and m2:
                    result["periodo"]["inicio"] = safe_date(y1, m1, d1)
                    result["periodo"]["fim"] = safe_date(y2, m2, d2)

            # PicPay tables: [Data/Hora, Descrição, Valor, Saldo, Saldo Sacável]
            saldo_first = None
            saldo_last = None

            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if not row or len(row) < 4:
                            continue
                        cols = [str(c).strip() if c else "" for c in row]

                        # Skip header
                        if "Data/Hora" in cols[0] or "Descrição" in cols[1]:
                            continue

                        # Parse date from "DD/MM/YYYY\nHH:MM:SS"
                        date_match = re.match(r'(\d{2}/\d{2}/\d{4})', cols[0])
                        if not date_match:
                            continue

                        parts = date_match.group(1).split("/")
                        iso_date = f"{parts[2]}-{parts[1]}-{parts[0]}"
                        descricao = cols[1]
                        valor = parse_brl(cols[2])
                        saldo = parse_brl(cols[3])

                        if valor is None:
                            continue

                        result["transacoes"].append({
                            "data": iso_date,
                            "descricao": descricao,
                            "valor": valor,
                        })

                        if saldo is not None:
                            if saldo_first is None:
                                saldo_first = saldo
                            saldo_last = saldo

            # PicPay lists newest first; transactions should be oldest first
            result["transacoes"].reverse()

            # Derive saldos (last in list = oldest = initial, first = newest = final)
            if saldo_first is not None:
                result["saldo_final"] = saldo_first
            if saldo_last is not None:
                result["saldo_inicial"] = saldo_last

    except Exception as e:
        log("ERROR", f"  Falha ao processar {filename}: {e}")
        result["notas"].append(f"Erro no parsing: {e}")
        result["requires_llm_fallback"] = True

    log("INFO", f"  → {len(result['transacoes'])} transações extraídas")
    return result


# =============================================================================
# Parser: Bradesco (extratoconta, extratopoupanca)
# TEXT_REGEX — multi-line text format
# =============================================================================

def parse_bradesco(pdf_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Bradesco bank statement (conta corrente or poupança)."""
    is_poupanca = "poupanca" in filename.lower()
    tipo = "extratopoupanca" if is_poupanca else "extratoconta"

    log("INFO", f"Parsing Bradesco ({tipo}): {filename}")
    result = make_result_template("Bradesco", tipo, "BRL")

    periodo_inicio, periodo_fim = infer_periodo_from_filename(filename)
    result["periodo"]["inicio"] = periodo_inicio
    result["periodo"]["fim"] = periodo_fim

    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Combine all pages text
            all_text = ""
            for page in pdf.pages:
                text = page.extract_text() or ""
                all_text += text + "\n"

            result["titular"] = detect_member_from_text(all_text)

            # Account: "Ag: 3221 | Conta: 77113-9"
            m = re.search(r'Ag[:\s]+(\d+)\s*\|\s*Conta[:\s]+([\d-]+)', all_text)
            if m:
                result["numero_conta"] = f"Ag {m.group(1)} Conta {m.group(2)}"

            # Periodo: "Entre DD/MM/YYYY e DD/MM/YYYY"
            pm = re.search(r'Entre\s+(\d{2}/\d{2}/\d{4})\s+e\s+(\d{2}/\d{2}/\d{4})', all_text)
            if pm:
                p1 = pm.group(1).split("/")
                p2 = pm.group(2).split("/")
                result["periodo"]["inicio"] = f"{p1[2]}-{p1[1]}-{p1[0]}"
                result["periodo"]["fim"] = f"{p2[2]}-{p2[1]}-{p2[0]}"

            # Bradesco format: transaction lines are multi-line blocks
            # Date line: DD/MM/YY at start
            # Then historico, docto, credito, debito, saldo on various lines
            #
            # Pattern: "DD/MM/YY historico docto [credito] [- debito] [saldo]"
            # Continuation lines have no date prefix
            # Values appear at end of line: "1.808,49" for credit, "- 1.500,00" for debit
            # Saldo appears after credit/debit: "1.809,49" or "1,00"

            lines = all_text.split("\n")
            transactions: List[Dict] = []
            saldo_anterior = None
            current_date = None

            # Bradesco PDF footer/boilerplate patterns to skip
            _bradesco_skip = re.compile(
                r'(?i)'
                r'(?:Fone\s+F[áa]cil|Capitais\s+e\s+Regi|Demais\s+Regi|'
                r'SAC\s+-|Ouvidoria|Se\s+[Pp]referir|BIA\s+pelo|'
                r'Atendimento\s+(?:24h|dispon|eletr|de\s+segunda|personal)|'
                r'fale\s+com\s+a\s+BIA|WhatsApp|Fale\s+Conosco|'
                r'Cancelamento.*reclama|sugest[ãa]o\s+e\s+elogio|'
                r'N[ãa]o\s+h[áa]\s+lan[çc]amentos|Os\s+dados\s+acima|'
                r'Domingos\s+e\s+feriados|Demais\s+telefones|'
                r'Saldo\s+Invest\s+F[áa]cil|'
                r'^\s*0800\s|'
                r'desenho\s+do\s+cadeado|Consulta\s+de\s+saldo|'
                r'Para\s+consultas\s+de\s+um\s+per[íi]odo|'
                r'transa[çc][õo]es\s+financeiras|'
                r'Bradesco\s+Internet\s+Banking|'
                r'Nome:\s+|Extrato\s+de:\s+Ag:|'
                r'^\s*metropolitanas\s*$|^\s*aparecer\s|^\s*Seguran[çc]a\s*$|'
                r'^\s*4002\s+0022\s*$|^\s*elogio\b)'
            )

            # Also track when we hit "Total" line — everything after is boilerplate
            _bradesco_end_marker = re.compile(r'^Total\s+[\d.,]+')

            # Find SALDO ANTERIOR
            for line in lines:
                m = re.match(r'(\d{2}/\d{2}/\d{2})\s+SALDO ANTERIOR\s+([\d.,]+)', line)
                if m:
                    saldo_anterior = parse_brl(m.group(2))
                    dd, mm, yy = m.group(1).split("/")
                    yy_full = 2000 + int(yy) if int(yy) < 50 else 1900 + int(yy)
                    result["saldo_inicial"] = saldo_anterior
                    break

            # Main transaction pattern for Bradesco:
            # Line starts with DD/MM/YY followed by historico and numbers
            # Numbers at the end: [credit] [- debit] [saldo]
            # The tricky part is distinguishing credit vs debit vs saldo
            #
            # Key insight: Bradesco always shows saldo at the end
            # If "- VALUE" appears, it's a debit
            # A standalone value before saldo could be credit

            tx_date_pattern = re.compile(
                r'^(\d{2}/\d{2}/\d{2})\s+(.*)'
            )

            # Parse transaction blocks
            # Bradesco multi-line format:
            #   DD/MM/YY desc DOCTO [CREDIT] [- DEBIT] [SALDO]   ← date line
            #   Transfe Pix                                        ← description line
            #   DOCTO - DEBIT_VALUE [SALDO]                        ← amount line
            #   Des: Name DD/MM                                    ← detail line
            #
            # Key: description comes BEFORE the amount line, so we
            # track a 'pending_desc' to carry it forward.

            i = 0
            pending_desc = ""  # accumulates description text for next amount line
            past_end = False  # set True after "Total ..." line

            while i < len(lines):
                line = lines[i].strip()

                # After "Total" summary line, everything is boilerplate
                if _bradesco_end_marker.match(line):
                    past_end = True
                if past_end:
                    # Reset if we hit a new "Extrato de:" header (multi-account PDF)
                    if re.search(r'Entre\s+\d{2}/\d{2}/\d{4}\s+e\s+\d{2}/\d{2}/\d{4}', line):
                        past_end = False
                    else:
                        i += 1
                        continue

                # Skip Bradesco PDF footer/boilerplate lines
                if _bradesco_skip.search(line):
                    i += 1
                    continue

                dm = tx_date_pattern.match(line)

                if dm:
                    date_str = dm.group(1)
                    rest = dm.group(2).strip()
                    dd, mm, yy = date_str.split("/")
                    yy_full = 2000 + int(yy) if int(yy) < 50 else 1900 + int(yy)
                    current_date = safe_date(yy_full, int(mm), int(dd))

                    # Skip non-transaction lines (subtotals, saldo lines)
                    if "SALDO ANTERIOR" in rest or re.match(r'^Total\s', rest):
                        pending_desc = ""
                        i += 1
                        continue

                    # The historico is everything before the first number
                    hist_match = re.match(r'(.+?)\s+(-?\s*\d[\d.,]*)', rest)
                    if hist_match:
                        historico = hist_match.group(1).strip()
                    else:
                        historico = rest.strip()

                    # If historico is just a docto number (5-8 digits), use
                    # pending_desc from the previous continuation line
                    # (Bradesco puts "Receb Pagfor" BEFORE the date line)
                    if re.match(r'^\d{5,8}$', historico) and pending_desc:
                        historico = pending_desc
                    pending_desc = ""  # reset after use

                    # Determine if this line has a complete transaction
                    debit_match = re.search(r'-\s+([\d.,]+)\s+([\d.,]+)\s*$', rest)
                    credit_match = re.search(r'(\d[\d.,]*)\s+([\d.,]+)\s*$', rest)

                    if debit_match:
                        valor = -parse_brl(debit_match.group(1))
                        if valor is not None:
                            transactions.append({
                                "data": current_date,
                                "descricao": historico,
                                "valor": valor,
                            })
                    elif credit_match:
                        nums = re.findall(r'[\d.,]+', rest)
                        if len(nums) >= 2:
                            possible_val = parse_brl(nums[-2])
                            possible_saldo = parse_brl(nums[-1])
                            if re.search(r'-\s+' + re.escape(nums[-2]), rest):
                                if possible_val is not None:
                                    transactions.append({
                                        "data": current_date,
                                        "descricao": historico,
                                        "valor": -possible_val,
                                    })
                            elif possible_val is not None and possible_val != possible_saldo:
                                raw = nums[-2].replace(".", "").replace(",", "")
                                if len(raw) <= 6:
                                    transactions.append({
                                        "data": current_date,
                                        "descricao": historico,
                                        "valor": possible_val,
                                    })

                elif current_date:
                    # Continuation line (no date prefix)
                    if line and not line.startswith("Data ") and not line.startswith("Bradesco") and not re.match(r'^Total\s', line):

                        # Pattern A: debit with saldo  "TEXT DOCTO - VALUE SALDO"
                        debit_m = re.search(r'-\s+([\d.,]+)\s+([\d.,]+)\s*$', line)
                        # Pattern B: debit WITHOUT saldo "DOCTO - VALUE" (end of line)
                        debit_no_saldo = None
                        if not debit_m:
                            debit_no_saldo = re.match(r'^(\d{5,8})\s+-\s+([\d.,]+)\s*$', line)
                        # Pattern C: credit "DOCTO VALUE [SALDO]"
                        credit_m = None
                        if not debit_m and not debit_no_saldo:
                            credit_m = re.search(r'(\d[\d.,]+)\s+([\d.,]+)\s*$', line)
                        # Pattern D: credit without saldo "DOCTO VALUE"
                        credit_no_saldo = None
                        if not debit_m and not debit_no_saldo and not credit_m:
                            credit_no_saldo = re.match(r'^(\d{5,8})\s+([\d.,]+)\s*$', line)

                        if debit_m or debit_no_saldo:
                            # Extract description from the line itself
                            hist = re.match(r'(.+?)\s+-\s+[\d.,]+', line)
                            line_desc = hist.group(1).strip() if hist else ""
                            line_desc = re.sub(r'^\d{5,8}$', '', line_desc).strip()
                            # Use line's own desc if it has text; otherwise use pending_desc
                            if line_desc and not re.match(r'^\d+$', line_desc):
                                desc = line_desc
                            elif pending_desc:
                                desc = pending_desc
                            else:
                                desc = line_desc if line_desc else line.strip()

                            if debit_m:
                                valor = -parse_brl(debit_m.group(1))
                            else:
                                valor = -parse_brl(debit_no_saldo.group(2))

                            if valor is not None and abs(valor) > 0.001:
                                transactions.append({
                                    "data": current_date,
                                    "descricao": desc if desc else line.strip(),
                                    "valor": valor,
                                })
                            pending_desc = ""

                        elif credit_m:
                            nums = re.findall(r'[\d.,]+', line)
                            if len(nums) >= 2:
                                possible_val = parse_brl(nums[-2])
                                possible_saldo = parse_brl(nums[-1])
                                if (possible_val is not None and possible_saldo is not None
                                        and possible_val != possible_saldo):
                                    raw = nums[-2].replace(".", "").replace(",", "")
                                    if len(raw) <= 6:
                                        # Line has text desc + value + saldo
                                        line_desc = line[:line.rfind(nums[-2])].strip()
                                        line_desc = re.sub(r'\s*\d{5,8}\s*$', '', line_desc).strip()
                                        if line_desc and not re.match(r'^\d+$', line_desc):
                                            desc = line_desc
                                        elif pending_desc:
                                            desc = pending_desc
                                        else:
                                            desc = line_desc
                                        if desc:
                                            transactions.append({
                                                "data": current_date,
                                                "descricao": desc,
                                                "valor": possible_val,
                                            })
                                        pending_desc = ""
                                    elif raw.isdigit() and len(raw) >= 5:
                                        # First number is docto (5-8 digits), second
                                        # could be credit WITHOUT saldo (intermediate tx)
                                        credit_val = parse_brl(nums[-1])
                                        if credit_val is not None and credit_val > 0:
                                            # Extract text before first number as desc
                                            first_num = nums[0] if nums else ""
                                            idx = line.find(first_num) if first_num else -1
                                            line_text = line[:idx].strip() if idx > 0 else ""
                                            # Use line text if available, else pending
                                            if line_text and not re.match(r'^\d+$', line_text):
                                                desc = line_text
                                            elif pending_desc:
                                                desc = pending_desc
                                            else:
                                                desc = line_text if line_text else line.strip()
                                            transactions.append({
                                                "data": current_date,
                                                "descricao": desc,
                                                "valor": credit_val,
                                            })
                                            pending_desc = ""
                                else:
                                    # Values are equal (rare) or not a value line
                                    if not re.match(r'^Des:', line) and not re.match(r'^Dest\.', line):
                                        text_part = re.sub(r'\s+\d{5,8}$', '', line).strip()
                                        if text_part and not re.match(r'^\d+$', text_part):
                                            pending_desc = text_part
                            else:
                                # Single number or pure text — description line
                                if not re.match(r'^Des:', line) and not re.match(r'^Dest\.', line) and not re.match(r'^\d{5,8}$', line):
                                    pending_desc = line

                        elif credit_no_saldo:
                            # "DOCTO VALUE" — credit without saldo
                            credit_val = parse_brl(credit_no_saldo.group(2))
                            if credit_val is not None and credit_val > 0:
                                desc = pending_desc if pending_desc else ""
                                if desc:
                                    transactions.append({
                                        "data": current_date,
                                        "descricao": desc,
                                        "valor": credit_val,
                                    })
                                pending_desc = ""
                        else:
                            # No number patterns — pure description line
                            # (e.g., "Ted Dif.titul", "Transfe Pix", "Receb Pagfor")
                            if re.match(r'^Des:', line) or re.match(r'^Dest\.', line):
                                # Detail line (PIX/TED recipient) — append to
                                # last transaction for richer description
                                if transactions:
                                    last = transactions[-1]
                                    if last["data"] == current_date:
                                        last["descricao"] += " " + line
                            elif re.match(r'^[A-Z][a-z].*\.$', line):
                                # Company name like "Grpqa Ltda.", "Bradesco C-pmsp sp"
                                # Append to last transaction description so keywords
                                # like "GRPQA" are present for categorization.
                                if transactions:
                                    last = transactions[-1]
                                    if last["data"] == current_date:
                                        last["descricao"] += " " + line
                            else:
                                pending_desc = line

                i += 1

            result["transacoes"] = transactions

            # Try to find saldo final from "Total" line or last saldo
            total_match = re.search(r'Total\s+([\d.,]+)\s+-\s+([\d.,]+)\s+([\d.,]+)', all_text)
            if total_match:
                result["saldo_final"] = parse_brl(total_match.group(3))

    except Exception as e:
        log("ERROR", f"  Falha ao processar {filename}: {e}")
        result["notas"].append(f"Erro no parsing: {e}")
        result["requires_llm_fallback"] = True

    log("INFO", f"  → {len(result['transacoes'])} transações extraídas")
    return result


# =============================================================================
# Parser: Santander XLS (extratoconta)
# XLS_NATIVE — xlrd-based parser for Santander internet banking .xls exports
#
# Structure (Sheet "Plan1"):
#   Row 0: "EXTRATO DE CONTA CORRENTE"
#   Row 2: "NOME" + ... + "Conta: AAAA-CC.CCCCCC.D"
#   Row 4: "Tipo de Lancamento: Todos" + "Extrato de DD/MM/YYYY a DD/MM/YYYY"
#   Row 5: headers [Data, Descrição, Docto, Situação, Crédito (R$), Débito (R$), Saldo (R$)]
#   Row 6+: transactions (newest first) — values as Brazilian-formatted strings
#     - Transaction: date, description, docto, situacao, credito_or_empty, debito_or_empty, saldo
#     - SALDO ANTERIOR: date, "SALDO ANTERIOR", "", "", "", "", saldo
#     - TOTAL: "TOTAL", "", "", "", total_creditos, total_debitos, ""
#   Footer: current balance info, juros/IOF info
# =============================================================================

def parse_santander_xls(xls_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Santander XLS bank statement exported from internet banking.

    Format: real XLS (xlrd-compatible) with 7 columns.
    Transactions listed newest-first. Values as Brazilian-formatted strings.
    """
    try:
        import xlrd
    except ImportError:
        log("ERROR", "xlrd not installed. Run: pip install xlrd")
        raise SystemExit("FATAL: xlrd is required for Santander XLS extraction. Install with: pip install xlrd")

    log("INFO", f"Parsing Santander XLS: {filename}")
    result = make_result_template("Santander", "extratoconta", "BRL")

    periodo_inicio, periodo_fim = infer_periodo_from_filename(filename)
    result["periodo"]["inicio"] = periodo_inicio
    result["periodo"]["fim"] = periodo_fim

    try:
        wb = xlrd.open_workbook(xls_path)
        sh = wb.sheet_by_index(0)  # "Plan1"

        # --- Header info ---
        if sh.nrows >= 3:
            # Titular (row 2, col 0)
            nome_raw = str(sh.cell(2, 0).value).strip()
            result["titular"] = detect_member_from_text(nome_raw)

            # Conta (row 2, col 4) — "Conta: 1652-01.001341.6"
            conta_raw = str(sh.cell(2, 4).value).strip()
            m = re.search(r'Conta:\s*([\d\-\.]+)', conta_raw)
            if m:
                result["numero_conta"] = m.group(1)

        # --- Period (row 4, col 4) — "Extrato de DD/MM/YYYY a DD/MM/YYYY" ---
        if sh.nrows >= 5:
            periodo_raw = str(sh.cell(4, 4).value).strip()
            pm = re.search(r'Extrato de\s+(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})', periodo_raw)
            if pm:
                p1 = pm.group(1).split("/")
                p2 = pm.group(2).split("/")
                result["periodo"]["inicio"] = f"{p1[2]}-{p1[1]}-{p1[0]}"
                result["periodo"]["fim"] = f"{p2[2]}-{p2[1]}-{p2[0]}"

        # --- Parse transactions (row 6+) ---
        saldo_anterior = None
        saldo_values = []

        for r in range(6, sh.nrows):
            cell_date = str(sh.cell(r, 0).value).strip()
            cell_desc = str(sh.cell(r, 1).value).strip()
            cell_credito = str(sh.cell(r, 4).value).strip()
            cell_debito = str(sh.cell(r, 5).value).strip()
            cell_saldo = str(sh.cell(r, 6).value).strip()

            # Stop at TOTAL row or empty section
            if cell_date.upper().startswith("TOTAL"):
                break

            # Skip footer/metadata rows
            if cell_date.startswith("Saldo de Conta") or cell_date.startswith("Juros acum"):
                break
            if cell_date.startswith("IOF acum"):
                break

            # Skip empty rows
            if not cell_date or cell_date == ' ':
                continue

            # Parse date
            date_match = re.match(r'(\d{2})/(\d{2})/(\d{4})', cell_date)
            if not date_match:
                continue

            dd, mm, yyyy = date_match.group(1), date_match.group(2), date_match.group(3)
            iso_date = f"{yyyy}-{mm}-{dd}"

            # SALDO ANTERIOR row
            if "SALDO ANTERIOR" in cell_desc.upper():
                saldo_val = parse_brl(cell_saldo)
                if saldo_val is not None:
                    saldo_anterior = saldo_val
                continue

            # Regular transaction — determine value from Crédito or Débito
            credito = parse_brl(cell_credito)
            debito = parse_brl(cell_debito)
            saldo = parse_brl(cell_saldo)

            if credito is not None and credito != 0:
                valor = abs(credito)  # Credits are positive
            elif debito is not None and debito != 0:
                valor = -abs(debito) if debito > 0 else debito  # Debits shown as negative or need to be negated
            else:
                continue

            result["transacoes"].append({
                "data": iso_date,
                "descricao": cell_desc,
                "valor": valor,
            })

            if saldo is not None:
                saldo_values.append((iso_date, saldo))

        # Santander lists newest first — reverse to chronological order
        result["transacoes"].reverse()
        saldo_values.reverse()

        # Set saldos
        if saldo_anterior is not None:
            result["saldo_inicial"] = saldo_anterior
        if saldo_values:
            result["saldo_final"] = saldo_values[-1][1]
        elif saldo_anterior is not None and not result["transacoes"]:
            # Period with no transactions — saldo final = saldo anterior
            result["saldo_final"] = saldo_anterior
            result["notas"].append("Conta sem movimentação no período (apenas saldo anterior registrado)")

    except Exception as e:
        log("ERROR", f"  Falha ao processar Santander XLS {filename}: {e}")
        result["notas"].append(f"Erro no parsing XLS: {e}")
        result["requires_llm_fallback"] = True

    log("INFO", f"  → {len(result['transacoes'])} transações extraídas do XLS")
    return result


# =============================================================================
# Parser: Santander CDB XLSX (investment positions)
# XLSX — openpyxl-based parser for Santander CDB position exports
#
# Structure (Sheet "Sheet0"):
#   Row 1: "CDB" | "Valor Total: R$NNN.NNN,NN" | "Valores Referentes a: DD/MM/YYYY"
#   Rows 3+: groups of 3 rows per product:
#     - Product header: "CDB DI SANTANDER" | "Valor Total: R$..." | "Disponível para Resgate: R$..."
#     - Column headers: "Operação" | "Valor Total(R$):" | "Disponível para Resgate(R$):"
#     - Data: operation_number | "R$NNN.NNN,NN" | "R$NNN.NNN,NN"
# =============================================================================

def parse_santander_cdb_xlsx(xlsx_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Santander CDB investment position from XLSX export.

    Output is compatible with E4's build_investimentos_unified().
    """
    try:
        import openpyxl
    except ImportError:
        log("ERROR", "openpyxl not installed. Run: pip install openpyxl")
        return {"requires_llm_fallback": True, "tipo": "cdbresumo"}

    log("INFO", f"Parsing Santander CDB XLSX: {filename}")

    result = {
        "instituicao": "Santander",
        "tipo": "cdbresumo",
        "tipo_produto": "CDB",
        "membro": None,
        "moeda": "BRL",
        "numero_conta": None,
        "data_referencia": None,
        "periodo": {"inicio": None, "fim": None},
        "saldo_anterior": None,
        "saldo_atual": None,
        "resumo": {},
        "posicoes": [],
        "notas": [],
    }

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        sh = wb[wb.sheetnames[0]]

        # Row 1: total header — "CDB" | "Valor Total: R$300.444,46" | "Valores Referentes a: DD/MM/YYYY"
        row1 = [str(sh.cell(1, c).value or "").strip() for c in range(1, sh.max_column + 1)]
        if len(row1) >= 2:
            total_m = re.search(r'Valor Total:\s*R\$\s*([\d.,]+)', row1[1])
            if total_m:
                result["saldo_atual"] = parse_brl(total_m.group(1))
        if len(row1) >= 3:
            date_m = re.search(r'Valores Referentes a:\s*(\d{2}/\d{2}/\d{4})', row1[2])
            if date_m:
                parts = date_m.group(1).split("/")
                result["data_referencia"] = f"{parts[2]}-{parts[1]}-{parts[0]}"
                result["periodo"]["fim"] = result["data_referencia"]
                result["periodo"]["inicio"] = result["data_referencia"]

        # Try to detect member from filename; Santander CDB XLSX doesn't include
        # the holder name, so also check account number from filename
        result["membro"] = detect_member_from_text(filename)
        if not result["membro"]:
            # Default: Santander CDB belongs to David (Conta 1652-01.001341.6)
            result["membro"] = "david"

        # Parse product groups — scan rows for product headers
        current_product = None
        current_valor_total = None
        current_resgate = None

        r = 2
        while r <= sh.max_row:
            vals = [str(sh.cell(r, c).value or "").strip() for c in range(1, sh.max_column + 1)]

            if not any(vals):
                r += 1
                continue

            # Product header: "CDB DI SANTANDER" | "Valor Total: R$..." | "Disponível para Resgate: R$..."
            if vals[0] and "CDB" in vals[0].upper() and "Valor Total:" in (vals[1] if len(vals) > 1 else ""):
                current_product = vals[0]
                vt_m = re.search(r'Valor Total:\s*R\$\s*([\d.,]+)', vals[1]) if len(vals) > 1 else None
                current_valor_total = parse_brl(vt_m.group(1)) if vt_m else None
                dr_m = re.search(r'Dispon[ií]vel para Resgate:\s*R\$\s*([\d.,]+)', vals[2]) if len(vals) > 2 else None
                current_resgate = parse_brl(dr_m.group(1)) if dr_m else None
                r += 1
                continue

            # Column header row — skip
            if vals[0] == "Operação":
                r += 1
                continue

            # Data row: operation number | "R$NNN.NNN,NN" | "R$NNN.NNN,NN"
            if current_product and vals[0] and re.match(r'^\d{15,}$', vals[0]):
                n_operacao = vals[0]
                valor_str = vals[1] if len(vals) > 1 else ""
                resgate_str = vals[2] if len(vals) > 2 else ""

                valor_m = re.search(r'R\$\s*([\d.,]+)', valor_str)
                resgate_m = re.search(r'R\$\s*([\d.,]+)', resgate_str)

                valor = parse_brl(valor_m.group(1)) if valor_m else current_valor_total
                resgate = parse_brl(resgate_m.group(1)) if resgate_m else current_resgate

                posicao = {
                    "nome": f"{current_product} - Op. {n_operacao}",
                    "tipo": current_product,
                    "n_operacao": n_operacao,
                    "valor_total": valor,
                    "valor_atual": valor,
                    "valor_resgate_disponivel": resgate,
                }
                result["posicoes"].append(posicao)

            r += 1

    except Exception as e:
        log("ERROR", f"  Falha ao processar Santander CDB XLSX {filename}: {e}")
        result["notas"].append(f"Erro no parsing: {e}")
        result["requires_llm_fallback"] = True

    n_pos = len(result["posicoes"])
    saldo = result.get("saldo_atual", 0) or 0
    log("INFO", f"  → {n_pos} posições CDB Santander, total R$ {saldo:,.2f}")
    return result


# =============================================================================
# Parser: Santander PDF (extratoconta)
# TEXT_REGEX — clean single-line format
# =============================================================================

def parse_santander_conta(pdf_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Santander bank account statement."""
    log("INFO", f"Parsing Santander Conta: {filename}")
    result = make_result_template("Santander", "extratoconta", "BRL")

    periodo_inicio, periodo_fim = infer_periodo_from_filename(filename)
    result["periodo"]["inicio"] = periodo_inicio
    result["periodo"]["fim"] = periodo_fim

    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_text = ""
            for page in pdf.pages:
                text = page.extract_text() or ""
                all_text += text + "\n"

            result["titular"] = detect_member_from_text(all_text)

            # Account: "Agência e Conta: 1652 / 01001341-6"
            m = re.search(r'Ag[êe]ncia\s+e\s+Conta[:\s]+([\d\s/\-]+)', all_text)
            if m:
                result["numero_conta"] = m.group(1).strip()

            # Periodo: "Período: DD/MM/YYYY a DD/MM/YYYY"
            pm = re.search(r'Per[ií]odo[:\s]+(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})', all_text)
            if pm:
                p1 = pm.group(1).split("/")
                p2 = pm.group(2).split("/")
                result["periodo"]["inicio"] = f"{p1[2]}-{p1[1]}-{p1[0]}"
                result["periodo"]["fim"] = f"{p2[2]}-{p2[1]}-{p2[0]}"

            # Check for "SALDO ANTERIOR" only with no transactions
            # If only "SALDO ANTERIOR" line exists on page 1 and no other transaction
            # lines, this is a period without activity
            saldo_ant_match = re.search(
                r'(\d{2}/\d{2}/\d{4})\s+SALDO ANTERIOR\s+(-?[\d.,]+)',
                all_text
            )
            if saldo_ant_match:
                result["saldo_inicial"] = parse_brl(saldo_ant_match.group(2))

            # Santander format: each line is a complete transaction
            # "DD/MM/YYYY DESCRIÇÃO DOCTO SITUAÇÃO CRÉDITO DÉBITO SALDO"
            # Credit appears as positive number, Debit as negative with "-"
            # Example: "06/02/2026 PIX RECEBIDO DAVID... 000000 5.000,00 118,34"
            # Example: "06/02/2026 DEBITO AUT. TELEFONE... 000000 -338,00 -4.881,66"

            tx_pattern = re.compile(
                r'^(\d{2}/\d{2}/\d{4})\s+'  # Date
                r'(.+?)\s+'                   # Description
                r'(\d{6})\s*'                 # Docto (6 digits)
                r'(-?[\d.,]+)\s+'             # Value (credit positive, debit negative)
                r'(-?[\d.,]+)\s*$',           # Saldo
                re.MULTILINE
            )

            saldo_values: List[Tuple[str, float]] = []

            for m in tx_pattern.finditer(all_text):
                date_parts = m.group(1).split("/")
                iso_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
                descricao = m.group(2).strip()
                valor = parse_brl(m.group(4))
                saldo = parse_brl(m.group(5))

                if valor is None:
                    continue

                result["transacoes"].append({
                    "data": iso_date,
                    "descricao": descricao,
                    "valor": valor,
                })

                if saldo is not None:
                    saldo_values.append((iso_date, saldo))

            # Santander lists newest first; reverse to chronological
            result["transacoes"].reverse()
            saldo_values.reverse()

            # Note if no transactions found (legitimate zero-activity period)
            if not result["transacoes"] and result["saldo_inicial"] is not None:
                result["notas"].append(
                    "Conta sem movimentação no período (apenas saldo anterior registrado)"
                )

            # Saldo anterior from previous search
            sa_pattern = re.search(
                r'Saldo anterior.*?Saldo \(R\$\)\s*\n\s*(\d{2}/\d{2}/\d{4})\s+(-?[\d.,]+)',
                all_text, re.DOTALL
            )
            if sa_pattern:
                result["saldo_inicial"] = parse_brl(sa_pattern.group(2))
            elif saldo_values:
                # Earliest saldo minus earliest transaction = initial
                result["saldo_inicial"] = saldo_values[0][1] - (result["transacoes"][0]["valor"] if result["transacoes"] else 0)

            if saldo_values:
                result["saldo_final"] = saldo_values[-1][1]

    except Exception as e:
        log("ERROR", f"  Falha ao processar {filename}: {e}")
        result["notas"].append(f"Erro no parsing: {e}")
        result["requires_llm_fallback"] = True

    log("INFO", f"  → {len(result['transacoes'])} transações extraídas")
    return result


# =============================================================================
# Parser: BTG Pactual (extratoconta)
# TEXT_REGEX — clean tabular text
# =============================================================================

def parse_btg(pdf_path: Path, filename: str) -> Dict[str, Any]:
    """Parse BTG Pactual bank statement."""
    log("INFO", f"Parsing BTG Pactual: {filename}")
    result = make_result_template("BTG Pactual", "extratoconta", "BRL")

    periodo_inicio, periodo_fim = infer_periodo_from_filename(filename)
    result["periodo"]["inicio"] = periodo_inicio
    result["periodo"]["fim"] = periodo_fim

    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_text = ""
            for page in pdf.pages:
                text = page.extract_text() or ""
                all_text += text + "\n"

            result["titular"] = detect_member_from_text(all_text)

            # Account
            m = re.search(r'Conta Corrente[:\s]+([\d]+)', all_text)
            if m:
                result["numero_conta"] = m.group(1)

            # CPF
            m = re.search(r'CPF[:\s]+([\d.\-]+)', all_text)

            # Periodo: "Período de DD/MM/YYYY a DD/MM/YYYY"
            pm = re.search(r'Per[ií]odo\s+de\s+(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})', all_text)
            if pm:
                p1 = pm.group(1).split("/")
                p2 = pm.group(2).split("/")
                result["periodo"]["inicio"] = f"{p1[2]}-{p1[1]}-{p1[0]}"
                result["periodo"]["fim"] = f"{p2[2]}-{p2[1]}-{p2[0]}"

            # BTG format: "DD/MM/YYYY DESCRIÇÃO DEBITO CREDITO SALDO"
            # "Saldo Inicial" and "Saldo Final" are special rows
            # Debit column: value without sign (represents money out)
            # Credit column: value without sign (represents money in)
            # One of debit/credit is present per line

            # Parse line by line
            lines = all_text.split("\n")
            in_movimentacao = False

            for line in lines:
                line = line.strip()

                if "Movimentação" in line and "Conta Corrente" in line:
                    in_movimentacao = True
                    continue

                if not in_movimentacao:
                    continue

                # Skip headers
                if line.startswith("Data") and "Descrição" in line:
                    continue

                # Saldo Inicial
                si_match = re.match(r'(\d{2}/\d{2}/\d{4})\s+Saldo Inicial\s+([\d.,]+)', line)
                if si_match:
                    result["saldo_inicial"] = parse_brl(si_match.group(2))
                    continue

                # Saldo Final
                sf_match = re.match(r'(\d{2}/\d{2}/\d{4})\s+Saldo Final\s+([\d.,]+)', line)
                if sf_match:
                    result["saldo_final"] = parse_brl(sf_match.group(2))
                    continue

                # Total lines
                if line.startswith("Total de"):
                    continue

                # Regular transaction: DD/MM/YYYY DESCRIPTION VALUE1 VALUE2
                # Where VALUE1 = debit OR credit, VALUE2 = saldo after
                tx_match = re.match(
                    r'(\d{2}/\d{2}/\d{4})\s+'  # Date
                    r'(.+?)\s+'                  # Description
                    r'([\d.,]+)\s+'              # Value (debit or credit)
                    r'([\d.,]+)\s*$',            # Saldo
                    line
                )
                if tx_match:
                    date_parts = tx_match.group(1).split("/")
                    iso_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
                    descricao = tx_match.group(2).strip()
                    valor_raw = parse_brl(tx_match.group(3))
                    saldo_after = parse_brl(tx_match.group(4))

                    if valor_raw is None:
                        continue

                    # Determine sign: compare saldo progression
                    # If saldo decreased → debit (negative)
                    # If saldo increased → credit (positive)
                    # Use "RESGATE", "REMUNERAÇÃO", "CREDITO", "RECEBIMENTO", "CUPOM",
                    # "DIVIDENDO", "Rendimento" as credit indicators
                    credit_keywords = [
                        "RESGATE", "REMUNERAÇÃO", "CREDITO", "CRÉDITO",
                        "RECEBIMENTO", "CUPOM", "DIVIDENDO", "Rendimento",
                        "RENDIMENT", "FRAÇÕES",
                    ]
                    desc_upper = descricao.upper()
                    is_credit = any(kw.upper() in desc_upper for kw in credit_keywords)

                    # Also check: if there's a previous transaction, compare saldos
                    if is_credit:
                        valor = valor_raw  # positive
                    else:
                        valor = -valor_raw  # debit = negative

                    result["transacoes"].append({
                        "data": iso_date,
                        "descricao": descricao,
                        "valor": valor,
                    })

                # Handle "- VALUE" notation for credit (CONTA REMUNERADA lines)
                tx_match2 = re.match(
                    r'(\d{2}/\d{2}/\d{4})\s+'
                    r'(.+?)\s+-\s+'
                    r'([\d.,]+)\s+'
                    r'([\d.,]+)\s*$',
                    line
                )
                if tx_match2 and not tx_match:
                    date_parts = tx_match2.group(1).split("/")
                    iso_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
                    descricao = tx_match2.group(2).strip()
                    valor_raw = parse_brl(tx_match2.group(3))
                    if valor_raw is not None:
                        result["transacoes"].append({
                            "data": iso_date,
                            "descricao": descricao,
                            "valor": valor_raw,  # "- VALUE" in BTG = credit (REMUNERAÇÃO)
                        })

    except Exception as e:
        log("ERROR", f"  Falha ao processar {filename}: {e}")
        result["notas"].append(f"Erro no parsing: {e}")
        result["requires_llm_fallback"] = True

    log("INFO", f"  → {len(result['transacoes'])} transações extraídas")
    return result


# =============================================================================
# Parser: Rico (extratoconta)
# TEXT_REGEX — single page, clean format
# =============================================================================

def parse_rico(pdf_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Rico corretora bank statement."""
    log("INFO", f"Parsing Rico: {filename}")
    result = make_result_template("Rico", "extratoconta", "BRL")

    periodo_inicio, periodo_fim = infer_periodo_from_filename(filename)
    result["periodo"]["inicio"] = periodo_inicio
    result["periodo"]["fim"] = periodo_fim

    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_text = ""
            for page in pdf.pages:
                text = page.extract_text() or ""
                all_text += text + "\n"

            result["titular"] = detect_member_from_text(all_text)

            # Account
            m = re.search(r'Conta[:\s]+(\d+)', all_text)
            if m:
                result["numero_conta"] = m.group(1)

            # Periodo: "De: DD/MM/YYYY Até: DD/MM/YYYY"
            pm = re.search(r'De[:\s]+(\d{2}/\d{2}/\d{4})\s+Até[:\s]+(\d{2}/\d{2}/\d{4})', all_text)
            if pm:
                p1 = pm.group(1).split("/")
                p2 = pm.group(2).split("/")
                result["periodo"]["inicio"] = f"{p1[2]}-{p1[1]}-{p1[0]}"
                result["periodo"]["fim"] = f"{p2[2]}-{p2[1]}-{p2[0]}"

            # Saldo disponível
            m = re.search(r'Saldo dispon[ií]vel[:\s]+R\$\s*([\d.,]+)', all_text)
            if m:
                result["saldo_final"] = parse_brl(m.group(1))

            # Rico format: "DD/MM/YYYY DD/MM/YYYY HISTORICO R$ valor R$ saldo"
            # Liq date, Mov date, then description, then R$ value, then R$ saldo
            tx_pattern = re.compile(
                r'(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+'  # Liq date, Mov date
                r'(.+?)\s+'                                        # Description
                r'R\$\s*([\d.,]+)\s+'                              # Value
                r'R\$\s*([\d.,]+)',                                 # Saldo
            )

            for m in tx_pattern.finditer(all_text):
                date_parts = m.group(1).split("/")  # Use Liq date
                iso_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
                descricao = m.group(3).strip()
                valor = parse_brl(m.group(4))

                if valor is None:
                    continue

                # Rico only shows credits (dividendos, JCP, rendimentos)
                result["transacoes"].append({
                    "data": iso_date,
                    "descricao": descricao,
                    "valor": valor,
                })

            # Derive saldo_inicial from saldo_final minus total credits
            if result["saldo_final"] is not None and result["transacoes"]:
                total = sum(t["valor"] for t in result["transacoes"] if t["valor"])
                result["saldo_inicial"] = round(result["saldo_final"] - total, 2)

    except Exception as e:
        log("ERROR", f"  Falha ao processar {filename}: {e}")
        result["notas"].append(f"Erro no parsing: {e}")
        result["requires_llm_fallback"] = True

    log("INFO", f"  → {len(result['transacoes'])} transações extraídas")
    return result


# =============================================================================
# Parser: Wise (extratocontausd, extratocontabrl)
# TEXT_REGEX — description + amount + balance on one line, date on next
# =============================================================================

def parse_wise(pdf_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Wise bank statement."""
    is_usd = "usd" in filename.lower()
    moeda = "USD" if is_usd else "BRL"
    tipo = f"extratoconta{moeda.lower()}"

    log("INFO", f"Parsing Wise ({moeda}): {filename}")
    result = make_result_template("Wise", tipo, moeda)

    periodo_inicio, periodo_fim = infer_periodo_from_filename(filename)
    result["periodo"]["inicio"] = periodo_inicio
    result["periodo"]["fim"] = periodo_fim

    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_text = ""
            for page in pdf.pages:
                text = page.extract_text() or ""
                all_text += text + "\n"

            result["titular"] = detect_member_from_text(all_text)

            # Account
            m = re.search(r'N[úu]mero da conta\s+.*?(\d{10,})', all_text)
            if m:
                result["numero_conta"] = m.group(1)

            # Periodo from header
            # "1 de janeiro de 2025 [GMT-03:00] - 29 de março de 2026 [GMT-03:00]"
            pm = re.search(
                r'(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})\s+\[.*?\]\s*-\s*'
                r'(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})',
                all_text
            )
            if pm:
                d1, m1, y1 = int(pm.group(1)), MESES_BR.get(pm.group(2).lower(), 0), int(pm.group(3))
                d2, m2, y2 = int(pm.group(4)), MESES_BR.get(pm.group(5).lower(), 0), int(pm.group(6))
                if m1 and m2:
                    result["periodo"]["inicio"] = safe_date(y1, m1, d1)
                    result["periodo"]["fim"] = safe_date(y2, m2, d2)

            # Current balance from header: "USD em DD de MÊS de YYYY [TZ] VALUE USD"
            bal_match = re.search(
                r'(?:USD|BRL)\s+em\s+.*?\s+([\d.,]+)\s+(?:USD|BRL)',
                all_text
            )
            if bal_match:
                result["saldo_final"] = parse_brl(bal_match.group(1))

            # Wise format: transactions are blocks of 2 lines:
            # Line 1: "Description [entrada|saída] [-]VALUE BALANCE"
            # Line 2: "DD de MÊS de YYYY Transação: ID [Referência: ref]"
            #
            # Entries/exits marked by column position or sign

            # Parse using combined pattern
            # Transaction pattern: description + optional amount columns + balance
            lines = all_text.split("\n")
            transactions: List[Dict] = []
            i = 0

            while i < len(lines):
                line = lines[i].strip()

                # Look for date line: "DD de MÊS de YYYY Transação: ..."
                date_match = re.match(
                    r'(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})\s+Transação',
                    line
                )

                if date_match and transactions:
                    # This is the date for the previous transaction
                    d = int(date_match.group(1))
                    m_name = date_match.group(2).lower()
                    y = int(date_match.group(3))
                    mo = MESES_BR.get(m_name, 0)
                    if mo:
                        transactions[-1]["data"] = safe_date(y, mo, d)
                    i += 1
                    continue

                # Transaction line: ends with amount and balance
                # Examples:
                # "Recebeu dinheiro de Douglas... 20,00 4.721,94"
                # "Pago para IU65 Premium & B -16,12 8.854,16"
                # "Dinheiro adicionado à conta 8.400,00 10.836,38"
                tx_match = re.match(
                    r'(.+?)\s+(-?[\d.,]+)\s+([\d.,]+)\s*$',
                    line
                )

                if tx_match:
                    descricao = tx_match.group(1).strip()
                    valor = parse_brl(tx_match.group(2))
                    # Skip header rows
                    if descricao in ("Descrição", "Descrição Entrada Saída Valor"):
                        i += 1
                        continue
                    if "Entrada" in descricao and "Saída" in descricao:
                        i += 1
                        continue

                    if valor is not None:
                        transactions.append({
                            "data": None,  # will be filled by next date line
                            "descricao": descricao,
                            "valor": valor,
                        })

                i += 1

            # Remove transactions without dates
            result["transacoes"] = [t for t in transactions if t.get("data")]

            # Wise lists newest first; reverse
            result["transacoes"].reverse()

            # Note if no transactions (legitimate zero-activity)
            if not result["transacoes"] and result["saldo_final"] is not None:
                result["notas"].append(
                    "Conta sem movimentação no período (saldo estável)"
                )

            # Derive saldo_inicial
            if result["saldo_final"] is not None and result["transacoes"]:
                total = sum(t["valor"] for t in result["transacoes"] if t["valor"])
                saldo_ini = round(result["saldo_final"] - total, 2)
                result["saldo_inicial"] = saldo_ini + 0.0  # avoid -0.0

    except Exception as e:
        log("ERROR", f"  Falha ao processar {filename}: {e}")
        result["notas"].append(f"Erro no parsing: {e}")
        result["requires_llm_fallback"] = True

    log("INFO", f"  → {len(result['transacoes'])} transações extraídas")
    return result


# =============================================================================
# Parser: Bank of America (extratoconta)
# TEXT_REGEX — standard US bank statement
# =============================================================================

def parse_bankofamerica(pdf_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Bank of America statement."""
    log("INFO", f"Parsing Bank of America: {filename}")
    result = make_result_template("Bank of America", "extratoconta", "USD")

    periodo_inicio, periodo_fim = infer_periodo_from_filename(filename)
    result["periodo"]["inicio"] = periodo_inicio
    result["periodo"]["fim"] = periodo_fim

    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_text = ""
            for page in pdf.pages:
                text = page.extract_text() or ""
                all_text += text + "\n"

            result["titular"] = detect_member_from_text(all_text)

            # Account: "Account number: XXXX XXXX XXXX"
            m = re.search(r'Account\s+number[:\s]+([\d\s]+)', all_text)
            if m:
                result["numero_conta"] = m.group(1).strip()

            # Periodo: "for February 25, 2026 to March 26, 2026"
            pm = re.search(
                r'for\s+(\w+)\s+(\d{1,2}),?\s+(\d{4})\s+to\s+(\w+)\s+(\d{1,2}),?\s+(\d{4})',
                all_text
            )
            if pm:
                months_en = {
                    'january': 1, 'february': 2, 'march': 3, 'april': 4,
                    'may': 5, 'june': 6, 'july': 7, 'august': 8,
                    'september': 9, 'october': 10, 'november': 11, 'december': 12,
                }
                m1 = months_en.get(pm.group(1).lower(), 0)
                m2 = months_en.get(pm.group(4).lower(), 0)
                if m1 and m2:
                    result["periodo"]["inicio"] = safe_date(int(pm.group(3)), m1, int(pm.group(2)))
                    result["periodo"]["fim"] = safe_date(int(pm.group(6)), m2, int(pm.group(5)))

            # Beginning/Ending balance — US format ($2,605.00)
            bb = re.search(r'Beginning balance.*?\$([\d.,]+)', all_text)
            eb = re.search(r'Ending balance.*?\$([\d.,]+)', all_text)
            if bb:
                result["saldo_inicial"] = parse_usd(bb.group(1))
            if eb:
                result["saldo_final"] = parse_usd(eb.group(1))

            # Transaction lines: "MM/DD/YY DESCRIPTION AMOUNT"
            # BoA uses US date format and US number format
            tx_pattern = re.compile(
                r'^(\d{2}/\d{2}/\d{2})\s+(.+?)\s+(-?[\d.,]+)\s*$',
                re.MULTILINE
            )
            for m in tx_pattern.finditer(all_text):
                mm, dd, yy = m.group(1).split("/")
                yy_full = 2000 + int(yy)
                iso_date = safe_date(yy_full, int(mm), int(dd))
                result["transacoes"].append({
                    "data": iso_date,
                    "descricao": m.group(2).strip(),
                    "valor": parse_usd(m.group(3)),
                })

            # Note: BoA may legitimately have 0 transactions (dormant account)
            if not result["transacoes"] and result["saldo_inicial"] == result["saldo_final"]:
                result["notas"].append("Conta sem movimentação no período (saldo inicial = saldo final)")

    except Exception as e:
        log("ERROR", f"  Falha ao processar {filename}: {e}")
        result["notas"].append(f"Erro no parsing: {e}")
        result["requires_llm_fallback"] = True

    log("INFO", f"  → {len(result['transacoes'])} transações extraídas")
    return result


# =============================================================================
# Router — maps filename patterns to parser functions
# =============================================================================

# Order matters: more specific patterns first
# =============================================================================
# Parser: Itaú CDB Investment Extracts (HTML-as-XLS)
# HTML_TABLE — BeautifulSoup parser for Itaú investment position exports
#
# These .xls files are actually HTML with Excel MIME type (common Itaú practice).
# Products: CDB-DI, CDB Metas e Reservas, etc.
#
# Structure:
#   Table 0 (main):
#     Row 5: Title ("Extrato de movimentação mensal - CDB-DI")
#     Rows 8-10: Account info (Nome, CPF, Agência, Conta)
#     Row 13: Period ("01/04/2026 a 08/04/2026")
#     Row 14: Movement headers
#     Rows 16+: SALDO ANTERIOR, transactions, SALDO FINAL
#     Row 23: Summary headers (Saldo anterior, Aplicações, Resgates, etc.)
#     Row 25: Summary totals
#     Row 29: Position headers (N.operação, Vencimento, etc.)
#     Rows 31+: Individual CDB positions
#
# Output: JSON compatible with E4 build_investimentos_unified()
#   Keys: posicoes[], saldo_atual, instituicao, membro, tipo_produto
# =============================================================================

def parse_itau_cdb_html_xls(xls_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Itaú CDB investment extract from HTML-as-XLS export.

    These .xls files from Itaú internet banking are actually HTML tables.
    Extracts position data, balances, and summary for CDB investments.
    Output is compatible with E4's build_investimentos_unified().
    """
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        log("ERROR", "beautifulsoup4 not installed. Run: pip install beautifulsoup4")
        return {"requires_llm_fallback": True, "tipo": "cdbresumo"}

    log("INFO", f"Parsing Itaú CDB HTML-XLS: {filename}")

    result = {
        "instituicao": "Itaú",
        "tipo": "cdbresumo",
        "tipo_produto": None,
        "membro": None,
        "moeda": "BRL",
        "numero_conta": None,
        "data_referencia": None,
        "periodo": {"inicio": None, "fim": None},
        "saldo_anterior": None,
        "saldo_atual": None,
        "resumo": {},
        "posicoes": [],
        "notas": [],
    }

    try:
        # Read as HTML (encoding declared in meta tag)
        with open(xls_path, 'r', encoding='windows-1252') as f:
            html = f.read()

        soup = BeautifulSoup(html, 'html.parser')
        tables = soup.find_all('table')

        if not tables:
            result["notas"].append("Nenhuma tabela encontrada no HTML")
            result["requires_llm_fallback"] = True
            return result

        main_table = tables[0]
        rows = main_table.find_all('tr')

        def get_cells(row):
            """Extract text from all cells in a row."""
            return [c.get_text(strip=True) for c in row.find_all(['td', 'th'])]

        # --- Parse structured data by scanning rows ---
        for i, row in enumerate(rows):
            cells = get_cells(row)
            if not cells or not any(cells):
                continue

            # Title: "Extrato de movimentação mensal - CDB-DI"
            if len(cells) >= 1 and "Extrato de movimentação mensal" in cells[0]:
                # Extract product name after the dash
                title = cells[0]
                dash_idx = title.find(" - ")
                if dash_idx >= 0:
                    result["tipo_produto"] = title[dash_idx + 3:].strip()

            # Nome
            if len(cells) >= 3 and cells[1] == "Nome:":
                nome = cells[2]
                result["membro"] = detect_member_from_text(nome)

            # Agência/Conta
            if len(cells) >= 5 and cells[1] == "Agência:" and cells[3] == "Conta:":
                result["numero_conta"] = cells[4]

            # Período
            if len(cells) >= 3 and cells[1] == "Período:":
                periodo_str = cells[2]
                m = re.search(r'(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})', periodo_str)
                if m:
                    p1, p2 = m.group(1).split("/"), m.group(2).split("/")
                    result["periodo"]["inicio"] = f"{p1[2]}-{p1[1]}-{p1[0]}"
                    result["periodo"]["fim"] = f"{p2[2]}-{p2[1]}-{p2[0]}"
                    result["data_referencia"] = result["periodo"]["fim"]

            # SALDO ANTERIOR (movement rows have date in cells[0], desc in cells[1], value in cells[2])
            if len(cells) >= 3 and "SALDO ANTERIOR" in cells[1].upper():
                val = parse_brl(cells[2])
                if val is not None:
                    result["saldo_anterior"] = val

            # SALDO FINAL
            if len(cells) >= 3 and "SALDO FINAL" in cells[1].upper():
                result["saldo_atual"] = parse_brl(cells[2])

            # Summary row ("Total:" with all the summary values)
            if len(cells) >= 9 and cells[0] == "Total:":
                result["resumo"] = {
                    "saldo_anterior": parse_brl(cells[1]),
                    "aplicacoes": parse_brl(cells[2]),
                    "resgates": parse_brl(cells[3]),
                    "vencimentos": parse_brl(cells[4]),
                    "rendimento_acumulado": parse_brl(cells[5]),
                    "saldo_bruto_final": parse_brl(cells[6]),
                    "impostos_estimados": parse_brl(cells[7]),
                    "saldo_final_liquido": parse_brl(cells[8]),
                }

            # Position detail rows (after "N. operação" header)
            # These have a numeric operation ID in first cell, dates in cells 1-2, values, etc.
            if len(cells) >= 8 and re.match(r'^\d{10,}$', cells[0]):
                n_operacao = cells[0]
                data_vencimento_raw = cells[1]
                data_aplicacao_raw = cells[2]
                valor_aplicacao = parse_brl(cells[3])
                remuneracao_pct = parse_brl(cells[4])
                valor_anterior = parse_brl(cells[5])
                valor_atual = parse_brl(cells[6])
                rentab_periodo = parse_brl(cells[7])

                # Convert dates DD/MM/YYYY → YYYY-MM-DD
                def convert_date(d):
                    m = re.match(r'(\d{2})/(\d{2})/(\d{4})', d)
                    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else d

                posicao = {
                    "nome": f"{result.get('tipo_produto', 'CDB')} - Op. {n_operacao}",
                    "tipo": result.get("tipo_produto", "CDB"),
                    "n_operacao": n_operacao,
                    "data_vencimento": convert_date(data_vencimento_raw),
                    "data_aplicacao": convert_date(data_aplicacao_raw),
                    "valor_aplicacao": valor_aplicacao,
                    "remuneracao_pct": remuneracao_pct,
                    "valor_anterior": valor_anterior,
                    "valor_total": valor_atual,
                    "valor_atual": valor_atual,
                    "rentabilidade_periodo_pct": rentab_periodo,
                }
                result["posicoes"].append(posicao)

    except Exception as e:
        log("ERROR", f"  Falha ao processar CDB HTML-XLS {filename}: {e}")
        result["notas"].append(f"Erro no parsing: {e}")
        result["requires_llm_fallback"] = True

    n_pos = len(result["posicoes"])
    saldo = result.get("saldo_atual", 0) or 0
    log("INFO", f"  → {n_pos} posições CDB, saldo R$ {saldo:,.2f}")
    return result


PARSER_REGISTRY: List[Tuple[re.Pattern, callable]] = [
    # C6 Bank CSV variants (matched first — CSV takes priority when available)
    (re.compile(r'^c6bank_extratocontapj_.*\.csv$'), parse_c6bank_csv),
    (re.compile(r'^c6bank_extratoconta_.*\.csv$'), parse_c6bank_csv),
    # C6 Bank PDF variants
    (re.compile(r'^c6bank_extratocontaglobalusd_'), parse_c6bank),
    (re.compile(r'^c6bank_extratocontaglobaleur_'), parse_c6bank),
    (re.compile(r'^c6bank_extratocontapj_'), parse_c6bank),
    (re.compile(r'^c6bank_extratoconta_'), parse_c6bank),
    # Itaú XLS (matched before PDF — XLS takes priority when available)
    (re.compile(r'^itau_extratocontapersonnalite_.*\.xls$'), parse_itau_xls),
    (re.compile(r'^itau_extratoconta_.*\.xls$'), parse_itau_xls),
    # Itaú PDF
    (re.compile(r'^itau_extratocontapersonnalite_'), parse_itau),
    (re.compile(r'^itau_extratoconta_'), parse_itau),
    # PicPay
    (re.compile(r'^picpay_extratoconta_'), parse_picpay),
    # Bradesco
    (re.compile(r'^bradesco_extratopoupanca_'), parse_bradesco),
    (re.compile(r'^bradesco_extratoconta_'), parse_bradesco),
    # Santander XLS (matched before PDF)
    (re.compile(r'^santander_extratoconta_.*\.xls$'), parse_santander_xls),
    # Santander PDF
    (re.compile(r'^santander_extratoconta_'), parse_santander_conta),
    # BTG Pactual
    (re.compile(r'^btgpactual_extratoconta_'), parse_btg),
    # Rico
    (re.compile(r'^rico_extratoconta_'), parse_rico),
    # Wise
    (re.compile(r'^wise_extratoconta'), parse_wise),
    # Bank of America
    (re.compile(r'^bankofamerica_extratoconta_'), parse_bankofamerica),
]

# Investment extract parsers (separate from bank statements)
INVESTMENT_REGISTRY: List[Tuple[re.Pattern, callable]] = [
    # Itaú CDB HTML-XLS (cdbresumo)
    (re.compile(r'^itau_cdbresumo_.*\.xls$'), parse_itau_cdb_html_xls),
    (re.compile(r'^itau_cdbdetalhes_.*\.xls$'), parse_itau_cdb_html_xls),
    # Santander CDB XLSX
    (re.compile(r'^santander_cdbresumo_.*\.xlsx$'), parse_santander_cdb_xlsx),
    (re.compile(r'^santander_cdbdetalhes_.*\.xlsx$'), parse_santander_cdb_xlsx),
]

# Types that are NOT bank statements (should not be processed by this script)
NON_STATEMENT_TYPES = re.compile(
    r'(fatura|investimentosposicao|carteirarendafixa|cdbdetalhes|cdbresumo|'
    r'informerendimentos|irpf|curriculo|holerite|baseline|dados_)'
)


def route_to_parser(filename: str) -> Optional[callable]:
    """Find the appropriate parser for a given filename."""
    # Check investment registry FIRST (these are excluded from PARSER_REGISTRY)
    for pattern, parser_fn in INVESTMENT_REGISTRY:
        if pattern.search(filename):
            return parser_fn

    # Skip non-statement types for regular parsers
    if NON_STATEMENT_TYPES.search(filename):
        return None

    for pattern, parser_fn in PARSER_REGISTRY:
        if pattern.search(filename):
            return parser_fn

    return None


# =============================================================================
# Validation gate
# =============================================================================

def validate_result(result: Dict[str, Any], file_path: Path, is_csv: bool = False) -> List[str]:
    """Validate extraction result. Returns list of warnings/errors."""
    issues = []

    n_tx = len(result.get("transacoes", []))
    periodo = result.get("periodo", {})

    # Check periodo
    if not periodo.get("inicio"):
        issues.append("WARN: periodo.inicio ausente")
    if not periodo.get("fim"):
        issues.append("WARN: periodo.fim ausente")

    # Check transactions vs file size
    if is_csv:
        # For CSV/XLS: check file size heuristic
        try:
            total_chars = file_path.stat().st_size
        except Exception:
            total_chars = 0

        # XLS binary format has ~36KB overhead even for empty files — raise threshold
        is_xls = str(file_path).endswith(".xls")
        size_threshold = 40000 if is_xls else 500  # 40KB for XLS, 500B for CSV

        if n_tx == 0 and total_chars > size_threshold:
            notas_lower = [n.lower() for n in result.get("notas", [])]
            is_empty_period = any(
                "sem lançamentos" in n or "sem movimentação" in n
                for n in notas_lower
            )
            if not is_empty_period:
                issues.append(
                    f"ERROR: 0 transações extraídas de {'XLS' if is_xls else 'CSV'} com {total_chars} bytes "
                    f"— provável falha de parsing"
                )
    else:
        try:
            with pdfplumber.open(file_path) as pdf:
                total_chars = sum(len(p.extract_text() or "") for p in pdf.pages)
                n_pages = len(pdf.pages)
        except Exception:
            total_chars = 0
            n_pages = 0

        # Heuristic: if PDF has significant text content but 0 transactions,
        # it's likely a parsing failure unless explicitly noted
        if n_tx == 0 and total_chars > 500 and n_pages > 0:
            notas_lower = [n.lower() for n in result.get("notas", [])]
            is_dormant = any(
                "sem movimentação" in n or "sem lançamentos" in n
                for n in notas_lower
            )
            if not is_dormant:
                issues.append(
                    f"ERROR: 0 transações extraídas de PDF com {total_chars} chars / "
                    f"{n_pages} páginas — provável falha de parsing"
                )

    # Check for transactions with None values
    none_vals = sum(1 for t in result.get("transacoes", []) if t.get("valor") is None)
    if none_vals > 0:
        issues.append(f"WARN: {none_vals} transações com valor None")

    # Check for duplicate transactions (same date+valor+descricao)
    seen = set()
    dupes = 0
    for t in result.get("transacoes", []):
        key = (t.get("data"), t.get("valor"), t.get("descricao", "")[:30])
        if key in seen:
            dupes += 1
        seen.add(key)
    if dupes > 0:
        issues.append(f"INFO: {dupes} possíveis duplicatas intra-arquivo")

    return issues


# =============================================================================
# Main — CLI interface
# =============================================================================

def find_extrato_files() -> List[Path]:
    """Find all bank statement PDFs and CSVs in data/financial_statements/."""
    if not DATA_DIR.is_dir():
        log("WARN", f"Diretório não encontrado: {DATA_DIR}")
        return []

    VALID_EXTENSIONS = ("-0_original.pdf", "-0_original.csv", "-0_original.xls", "-0_original.xlsx")

    files = []
    for f in sorted(DATA_DIR.iterdir()):
        if not f.is_file():
            continue
        if not any(f.name.endswith(ext) for ext in VALID_EXTENSIONS):
            continue
        # Check if this is a statement type (not fatura, investment, etc.)
        # Check if file matches investment registry (always include)
        is_investment = any(pat.search(f.name) for pat, _ in INVESTMENT_REGISTRY)
        if is_investment:
            files.append(f)
            continue

        if NON_STATEMENT_TYPES.search(f.name):
            continue
        # Must contain "extrato" in filename
        if "extrato" not in f.name.lower():
            continue
        files.append(f)

    return files


def process_file(file_path: Path, dry_run: bool = False) -> Optional[Dict[str, Any]]:
    """Process a single PDF or CSV file through the appropriate parser."""
    filename = file_path.name

    parser_fn = route_to_parser(filename)
    if parser_fn is None:
        log("WARN", f"Sem parser determinístico para: {filename}")
        return {
            "requires_llm_fallback": True,
            "arquivo": filename,
            "motivo": "Banco/tipo não reconhecido pelo parser determinístico",
        }

    if dry_run:
        log("INFO", f"[DRY-RUN] Processaria: {filename} → {parser_fn.__name__}")
        return None

    result = parser_fn(file_path, filename)

    # Run validation (skip pdfplumber-based checks for CSV/XLS files)
    is_csv = filename.endswith(".csv") or filename.endswith(".xls")
    issues = validate_result(result, file_path, is_csv=is_csv)
    for issue in issues:
        level = issue.split(":")[0]
        log(level, f"  {filename}: {issue}")
        result.setdefault("notas", []).append(issue)

    return result


def save_result(result: Dict[str, Any], filename: str) -> Path:
    """Save extraction result to E2_extracts directory."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Output filename: replace -0_original.{pdf,csv,xls} with -2_extract.json
    out_name = filename.replace("-0_original.pdf", "-2_extract.json")
    out_name = out_name.replace("-0_original.csv", "-2_extract.json")
    out_name = out_name.replace("-0_original.xls", "-2_extract.json")
    out_name = out_name.replace("-0_original.xlsx", "-2_extract.json")
    out_path = OUTPUT_DIR / out_name

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    return out_path


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="E2 Extrato Extraction — Deterministic parsers for bank statements"
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Show what would be processed without writing files")
    parser.add_argument("--file", type=str, default=None,
                        help="Process a specific PDF or CSV file")
    parser.add_argument("--output-dir", type=str, default=None,
                        help="Output directory (default: processed/E2_extracts/)")
    parser.add_argument("--quiet", action="store_true",
                        help="Suppress debug output")

    args = parser.parse_args()

    global _VERBOSE, OUTPUT_DIR
    if args.quiet:
        _VERBOSE = False
    if args.output_dir:
        OUTPUT_DIR = Path(args.output_dir)

    log("INFO", "=" * 60)
    log("INFO", "E2 EXTRATO EXTRACTION — Deterministic Parsers")
    log("INFO", "=" * 60)

    if args.file:
        pdf_path = Path(args.file)
        if not pdf_path.exists():
            # Try relative to DATA_DIR
            pdf_path = DATA_DIR / args.file
        if not pdf_path.exists():
            log("ERROR", f"Arquivo não encontrado: {args.file}")
            sys.exit(1)
        files = [pdf_path]
    else:
        files = find_extrato_files()

    if not files:
        log("INFO", "Nenhum extrato encontrado para processar.")
        return

    log("INFO", f"Encontrados {len(files)} extratos para processar")

    # Statistics
    stats = {
        "processados": 0,
        "transacoes_total": 0,
        "llm_fallback": 0,
        "erros_validacao": 0,
        "warnings": 0,
    }

    for pdf_path in files:
        result = process_file(pdf_path, dry_run=args.dry_run)
        if result is None:
            continue

        if result.get("requires_llm_fallback"):
            stats["llm_fallback"] += 1
            log("WARN", f"  → Requer LLM fallback: {pdf_path.name}")
            continue

        n_tx = len(result.get("transacoes", []))
        stats["processados"] += 1
        stats["transacoes_total"] += n_tx

        # Count issues
        for note in result.get("notas", []):
            if note.startswith("ERROR"):
                stats["erros_validacao"] += 1
            elif note.startswith("WARN"):
                stats["warnings"] += 1

        if not args.dry_run:
            out_path = save_result(result, pdf_path.name)
            log("INFO", f"  → Salvo: {out_path.name} ({n_tx} transações)")

    # Summary
    log("INFO", "=" * 60)
    log("INFO", "RESUMO:")
    log("INFO", f"  Processados: {stats['processados']}")
    log("INFO", f"  Total transações: {stats['transacoes_total']}")
    log("INFO", f"  LLM fallback: {stats['llm_fallback']}")
    log("INFO", f"  Erros de validação: {stats['erros_validacao']}")
    log("INFO", f"  Warnings: {stats['warnings']}")
    log("INFO", "=" * 60)

    if stats["erros_validacao"] > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
