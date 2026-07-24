#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Rico Corretora — extrato conta (PDF com regex)."""

import re
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

from scripts.e2.common import (
    BANCO_RICO,
    detect_member_from_text,
    infer_periodo_from_filename,
    log,
    make_result_template,
    new_investment_position_result,
    parse_brl,
)
from scripts.e2.validation import apply_rv_carteira_checksum

LOG_PREFIX = "E2-EXTRATO"

PARSERS = [
    # Anchor subtipo-agnóstico (sem terminador) — ver bankofamerica.py.
    (r"^rico_extratoconta", "parse_rico"),
    # Carteira consolidada (XLSX export): posições valoradas por classe.
    (r"^rico_investimentosposicao_.*\.xlsx$", "parse_rico_carteira"),
]

_RICO_TICKER_RE = re.compile(r"^[A-Z]{4}\d{1,2}$")
_RICO_PROVENTOS_RE = re.compile(r"Dividendos|Proventos", re.I)
_RICO_QTD_RE = re.compile(r"\d{1,3}(\.\d{3})*$")


def parse_rico(pdf_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Rico corretora bank statement."""
    log(LOG_PREFIX, "INFO", f"Parsing Rico: {filename}")
    result = make_result_template(BANCO_RICO, "extratoconta", "BRL")
    result["tipo_conta"] = "investimento"

    periodo_inicio, periodo_fim = infer_periodo_from_filename(filename)
    result["periodo"]["inicio"] = periodo_inicio
    result["periodo"]["fim"] = periodo_fim

    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_text = ""
            for page in pdf.pages:
                text = page.extract_text() or ""
                all_text += text + "\n"

            result["titular"] = detect_member_from_text(all_text)

            m = re.search(r"Conta[:\s]+(\d+)", all_text)
            if m:
                result["numero_conta"] = m.group(1)

            pm = re.search(r"De[:\s]+(\d{2}/\d{2}/\d{4})\s+Até[:\s]+(\d{2}/\d{2}/\d{4})", all_text)
            if pm:
                p1 = pm.group(1).split("/")
                p2 = pm.group(2).split("/")
                result["periodo"]["inicio"] = f"{p1[2]}-{p1[1]}-{p1[0]}"
                result["periodo"]["fim"] = f"{p2[2]}-{p2[1]}-{p2[0]}"

            m = re.search(r"Saldo dispon[ií]vel[:\s]+R\$\s*([\d.,]+)", all_text)
            if m:
                result["saldo_final"] = parse_brl(m.group(1))

            tx_pattern = re.compile(
                r"(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+"
                r"(.+?)\s+"
                r"R\$\s*([\d.,]+)\s+"
                r"R\$\s*([\d.,]+)",
            )

            for m in tx_pattern.finditer(all_text):
                date_parts = m.group(1).split("/")
                iso_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
                descricao = m.group(3).strip()
                valor = parse_brl(m.group(4))

                if valor is None:
                    continue

                result["transacoes"].append(
                    {
                        "data": iso_date,
                        "descricao": descricao,
                        "valor": valor,
                    }
                )

            if result["saldo_final"] is not None and result["transacoes"]:
                total = sum(t["valor"] for t in result["transacoes"] if t["valor"])
                result["saldo_inicial"] = round(result["saldo_final"] - total, 2)

    except Exception as e:
        log(LOG_PREFIX, "ERROR", f"  Falha ao processar {filename}: {e}")
        result["notas"].append(f"Erro no parsing: {e}")
        result["requires_llm_fallback"] = True

    log(LOG_PREFIX, "INFO", f"  → {len(result['transacoes'])} transações extraídas")
    return result


def _rico_brl(cell: Any) -> Optional[float]:
    if cell is None:
        return None
    return parse_brl(str(cell).replace("R$", "").replace("\xa0", " ").strip())


def _rico_qtd(cells: List[Any]) -> Optional[int]:
    for c in reversed(cells):
        s = str(c).strip()
        if _RICO_QTD_RE.fullmatch(s):
            return int(s.replace(".", ""))
    return None


def _rico_is_class_header(cells: List[Any]) -> bool:
    """Cabeçalho de classe = 2 células, a 2ª é o subtotal R$ da classe."""
    if len(cells) != 2:
        return False
    first, second = str(cells[0]).strip(), str(cells[1]).strip()
    return (
        second.startswith("R$") and not first.startswith("R$") and not _RICO_TICKER_RE.match(first)
    )


def _rico_position(cells: List[Any], classe: str) -> Optional[Dict[str, Any]]:
    valor = _rico_brl(cells[1])
    if valor is None:
        return None
    nome = str(cells[0]).strip()
    pos: Dict[str, Any] = {"nome": nome, "valor_atual": valor, "classe": classe}
    if _RICO_TICKER_RE.match(nome):
        pos["ticker"] = nome
        qtd = _rico_qtd(cells)
        if qtd is not None:
            pos["quantidade"] = qtd
    return pos


def _rico_proventos_index(rows: List[List[Any]]) -> Optional[int]:
    for i, raw in enumerate(rows):
        cells = [str(c).strip() for c in raw if c is not None and str(c).strip()]
        if cells and _RICO_PROVENTOS_RE.search(cells[0]):
            return i
    return None


def _consume_rico_position_row(
    result: Dict[str, Any],
    cells: List[Any],
    subtotais: Dict[str, float],
    classe: Optional[str] = None,
) -> Optional[str]:
    if _rico_is_class_header(cells):
        nome = str(cells[0]).strip()
        subtotais[nome] = _rico_brl(cells[1]) or 0.0
        return nome
    if classe and len(cells) >= 3 and str(cells[1]).strip().startswith("R$"):
        pos = _rico_position(cells, classe)
        if pos is not None:
            result["posicoes"].append(pos)
    return classe


def _rico_carteira_rows(xlsx_path: Path) -> List[List[Any]]:
    import warnings

    import openpyxl

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Sua carteira"] if "Sua carteira" in wb.sheetnames else wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _fill_rico_positions(result: Dict[str, Any], rows: List[List[Any]]) -> Dict[str, float]:
    prov_idx = _rico_proventos_index(rows)
    body = rows[:prov_idx] if prov_idx is not None else rows
    tail = rows[prov_idx:] if prov_idx is not None else []
    subtotais: Dict[str, float] = {}
    classe: Optional[str] = None
    for raw in body:
        cells = [c for c in raw if c is not None and str(c).strip()]
        if cells:
            classe = _consume_rico_position_row(result, cells, subtotais, classe)
    n_prov = sum(1 for raw in tail if _rico_row_is_proventos_position(raw))
    if n_prov:
        result["notas"].append(
            f"{n_prov} proventos/JCP detectados — fora do PL (categoria própria, follow-up)"
        )
    return subtotais


def _rico_row_is_proventos_position(raw: List[Any]) -> bool:
    cells = [c for c in raw if c is not None and str(c).strip()]
    return len(cells) >= 2 and _RICO_TICKER_RE.match(str(cells[0]).strip()) is not None


def parse_rico_carteira(xlsx_path: Path, filename: str) -> Dict[str, Any]:
    """Carteira Rico/XP (XLSX): posições valoradas por classe; proventos fora do PL (ADR-346)."""
    log(LOG_PREFIX, "INFO", f"Parsing Rico carteira: {filename}")
    result = new_investment_position_result(BANCO_RICO)
    try:
        rows = _rico_carteira_rows(xlsx_path)
    except Exception as e:
        result["requires_llm_fallback"] = True
        result["notas"].append(f"Erro ao ler XLSX: {e}")
        return result
    subtotais = _fill_rico_positions(result, rows)
    apply_rv_carteira_checksum(result, subtotais)
    log(LOG_PREFIX, "INFO", f"  → {len(result['posicoes'])} posição(ões) RV")
    return result
