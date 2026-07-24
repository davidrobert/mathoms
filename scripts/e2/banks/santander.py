#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Santander — all parsers (extratos + faturas).

Extrato parsers:
  - parse_santander_xls       : extrato conta corrente XLS (xlrd)
  - parse_santander_cdb_xlsx  : CDB investment position XLSX (openpyxl)
  - parse_santander_conta     : extrato conta corrente PDF (pdfplumber)

Fatura parsers:
  - parse_santander_fatura_csv : fatura Unique CSV
  - parse_santander_unique     : fatura Unique PDF (pdfplumber)
"""

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

from scripts.e2.common import (
    BANCO_SANTANDER,
    CARTAO_UNIQUE,
    FAMILY,
    MESES_BR_STR,
    SANTANDER_XLS_LAYOUT,
    TITULAR,
    VENC_UNIQUE,
    count_candidate_rows,
    detect_member_from_text,
    infer_fatura_ref_from_filename,
    infer_periodo_from_filename,
    log,
    make_result_template,
    new_cdb_position_result,
    parse_brl,
    read_pdf_text,
    resolve_date_ddmm,
    safe_date,
)
from scripts.e2.validation import apply_cdb_checksum

LOG_PREFIX_EXTRATO = "E2-EXTRATO"
LOG_PREFIX_FATURA = "E2-FATURA"

PARSERS = [
    # Anchors subtipo-agnósticos (sem underscore terminador) — casam subtipos de
    # moeda do E0 (extratocontausd/brl/eur/...). Ordem preserva format-specific:
    # `.xls$` → parse_santander_xls antes do fallback any-ext. Ver bankofamerica.py.
    (r"^santander_extratoconta.*\.xls$", "parse_santander_xls"),
    (r"^santander_extratoconta", "parse_santander_conta"),
    (r"^santander_cdbresumo_.*\.xlsx$", "parse_santander_cdb_xlsx"),
    (r"^santander_cdbdetalhes_.*\.xlsx$", "parse_santander_cdb_xlsx"),
    (r"^santander_cdbresumo_.*\.pdf$", "parse_santander_cdb_pdf"),
    (r"^santander_cdbdetalhes_.*\.pdf$", "parse_santander_cdb_pdf"),
    (r"santander_faturaunique.*\.csv$", "parse_santander_fatura_csv"),
    (r"santander_faturaunique", "parse_santander_unique"),
]


# ---------------------------------------------------------------------------
# Extrato: XLS (xlrd)
# ---------------------------------------------------------------------------


def parse_santander_xls(xls_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Santander XLS bank statement exported from internet banking.

    Format: real XLS (xlrd-compatible) with 7 columns.
    Transactions listed newest-first. Values as Brazilian-formatted strings.
    """
    try:
        import xlrd
    except ImportError:
        log(LOG_PREFIX_EXTRATO, "ERROR", "xlrd not installed. Run: pip install xlrd")
        result = make_result_template(BANCO_SANTANDER, "extratoconta", "BRL")
        result["notas"].append("xlrd not installed — cannot parse XLS")
        result["requires_llm_fallback"] = True
        return result

    log(LOG_PREFIX_EXTRATO, "INFO", f"Parsing Santander XLS: {filename}")
    result = make_result_template(BANCO_SANTANDER, "extratoconta", "BRL")
    result["tipo_conta"] = "corrente"

    periodo_inicio, periodo_fim = infer_periodo_from_filename(filename)
    result["periodo"]["inicio"] = periodo_inicio
    result["periodo"]["fim"] = periodo_fim

    try:
        wb = xlrd.open_workbook(xls_path)
        sh = wb.sheet_by_index(0)  # "Plan1"

        # --- Header info — layout from config ---
        _sh = SANTANDER_XLS_LAYOUT.get("header", {})
        _st_r = _sh.get("titular_row", 2)
        _st_c = _sh.get("titular_col", 0)
        _sc_r = _sh.get("conta_row", 2)
        _sc_c = _sh.get("conta_col", 4)
        _sp_r = _sh.get("periodo_row", 4)
        _sp_c = _sh.get("periodo_col", 4)

        if sh.nrows >= _st_r + 1:
            # Titular
            nome_raw = str(sh.cell(_st_r, _st_c).value).strip()
            result["titular"] = detect_member_from_text(nome_raw)

            # Conta — "Conta: 1652-01.001341.6" (agencia é o prefixo antes do primeiro hífen)
            conta_raw = str(sh.cell(_sc_r, _sc_c).value).strip()
            m = re.search(r"Conta:\s*([\d\-\.]+)", conta_raw)
            if m:
                result["numero_conta"] = m.group(1)
                ag_m = re.match(r"(\d+)", m.group(1))
                if ag_m:
                    result["agencia"] = ag_m.group(1)

        # --- Period ---
        if sh.nrows >= _sp_r + 1:
            periodo_raw = str(sh.cell(_sp_r, _sp_c).value).strip()
            pm = re.search(
                r"Extrato de\s+(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})", periodo_raw
            )
            if pm:
                p1 = pm.group(1).split("/")
                p2 = pm.group(2).split("/")
                result["periodo"]["inicio"] = f"{p1[2]}-{p1[1]}-{p1[0]}"
                result["periodo"]["fim"] = f"{p2[2]}-{p2[1]}-{p2[0]}"

        # --- Parse transactions — columns from config ---
        saldo_anterior = None
        saldo_values = []

        _sdata_start = SANTANDER_XLS_LAYOUT.get("data_start_row", 6)
        _scols = SANTANDER_XLS_LAYOUT.get("columns", {})
        _scd = _scols.get("data", 0)
        _scde = _scols.get("descricao", 1)
        _sccr = _scols.get("credito", 4)
        _scdb = _scols.get("debito", 5)
        _scss = _scols.get("saldo", 6)

        for r in range(_sdata_start, sh.nrows):
            cell_date = str(sh.cell(r, _scd).value).strip()
            cell_desc = str(sh.cell(r, _scde).value).strip()
            cell_credito = str(sh.cell(r, _sccr).value).strip()
            cell_debito = str(sh.cell(r, _scdb).value).strip()
            cell_saldo = str(sh.cell(r, _scss).value).strip()

            # Stop at TOTAL row or empty section
            if cell_date.upper().startswith("TOTAL"):
                break

            # Skip footer/metadata rows
            if cell_date.startswith("Saldo de Conta") or cell_date.startswith("Juros acum"):
                break
            if cell_date.startswith("IOF acum"):
                break

            # Skip empty rows
            if not cell_date or cell_date == " ":
                continue

            # Parse date
            date_match = re.match(r"(\d{2})/(\d{2})/(\d{4})", cell_date)
            if not date_match:
                continue

            dd, mm, yyyy = date_match.group(1), date_match.group(2), date_match.group(3)
            iso_date = f"{yyyy}-{mm}-{dd}"

            # SALDO ANTERIOR row
            if "SALDO ANTERIOR" in cell_desc.upper():
                saldo_val = parse_brl(cell_saldo)
                if saldo_val is not None:
                    saldo_anterior = saldo_val
                continue

            # Regular transaction — determine value from Crédito or Débito
            credito = parse_brl(cell_credito)
            debito = parse_brl(cell_debito)
            saldo = parse_brl(cell_saldo)

            if credito is not None and credito != 0:
                valor = abs(credito)  # Credits are positive
            elif debito is not None and debito != 0:
                valor = (
                    -abs(debito) if debito > 0 else debito
                )  # Debits shown as negative or need to be negated
            else:
                continue

            result["transacoes"].append(
                {
                    "data": iso_date,
                    "descricao": cell_desc,
                    "valor": valor,
                }
            )

            if saldo is not None:
                saldo_values.append((iso_date, saldo))

        # Santander lists newest first — reverse to chronological order
        result["transacoes"].reverse()
        saldo_values.reverse()

        # Set saldos
        if saldo_anterior is not None:
            result["saldo_inicial"] = saldo_anterior
        if saldo_values:
            result["saldo_final"] = saldo_values[-1][1]
        elif saldo_anterior is not None and not result["transacoes"]:
            # Period with no transactions — saldo final = saldo anterior
            result["saldo_final"] = saldo_anterior
            result["notas"].append(
                "Conta sem movimentação no período (apenas saldo anterior registrado)"
            )

        # A39.l7 · ADR-342: saldo_inicial (SALDO ANTERIOR) e saldo_final vêm de
        # células independentes (não derivados) → declara verificabilidade p/ o
        # gate HARD graduar conservação. Wise/Rico ficam fora (saldo derivado).
        if (
            result.get("saldo_inicial") is not None
            and result.get("saldo_final") is not None
            and result["transacoes"]
        ):
            result["conservacao_verificavel"] = True

    except Exception as e:
        log(LOG_PREFIX_EXTRATO, "ERROR", f"  Falha ao processar Santander XLS {filename}: {e}")
        result["notas"].append(f"Erro no parsing XLS: {e}")
        result["requires_llm_fallback"] = True

    log(LOG_PREFIX_EXTRATO, "INFO", f"  → {len(result['transacoes'])} transações extraídas do XLS")
    return result


# ---------------------------------------------------------------------------
# Extrato: CDB XLSX (openpyxl)
# ---------------------------------------------------------------------------


def parse_santander_cdb_xlsx(xlsx_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Santander CDB investment position from XLSX export.

    Output is compatible with E4's build_investimentos_unified().
    """
    try:
        import openpyxl
    except ImportError:
        log(LOG_PREFIX_EXTRATO, "ERROR", "openpyxl not installed. Run: pip install openpyxl")
        return {"requires_llm_fallback": True, "tipo": "cdbresumo"}

    log(LOG_PREFIX_EXTRATO, "INFO", f"Parsing Santander CDB XLSX: {filename}")

    result = {
        # ADR-284/A24.l7: `banco` aditivo ao lado de `instituicao` (valor idêntico)
        # — satisfaz required do e2_extract.schema.json; E4 lê `instituicao or banco`.
        "banco": BANCO_SANTANDER,
        "instituicao": BANCO_SANTANDER,
        "tipo": "cdbresumo",
        "tipo_produto": "CDB",
        "tipo_conta": "investimento",
        "membro": None,
        "moeda": "BRL",
        "numero_conta": None,
        "agencia": None,
        "data_referencia": None,
        "periodo": {"inicio": None, "fim": None},
        "saldo_anterior": None,
        "saldo_atual": None,
        "resumo": {},
        "posicoes": [],
        "notas": [],
    }

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        sh = wb[wb.sheetnames[0]]

        # Row 1: total header — "CDB" | "Valor Total: R$300.444,46" | "Valores Referentes a: DD/MM/YYYY"
        row1 = [str(sh.cell(1, c).value or "").strip() for c in range(1, sh.max_column + 1)]
        if len(row1) >= 2:
            total_m = re.search(r"Valor Total:\s*R\$\s*([\d.,]+)", row1[1])
            if total_m:
                result["saldo_atual"] = parse_brl(total_m.group(1))
        if len(row1) >= 3:
            date_m = re.search(r"Valores Referentes a:\s*(\d{2}/\d{2}/\d{4})", row1[2])
            if date_m:
                parts = date_m.group(1).split("/")
                result["data_referencia"] = f"{parts[2]}-{parts[1]}-{parts[0]}"
                result["periodo"]["fim"] = result["data_referencia"]
                result["periodo"]["inicio"] = result["data_referencia"]

        # Try to detect member from filename; Santander CDB XLSX doesn't include
        # the holder name, so also check account number from filename
        result["membro"] = detect_member_from_text(filename)
        if not result["membro"]:
            result["membro"] = FAMILY.get("titular", "david")

        # Parse product groups — scan rows for product headers
        current_product = None
        current_valor_total = None
        current_resgate = None

        r = 2
        while r <= sh.max_row:
            vals = [str(sh.cell(r, c).value or "").strip() for c in range(1, sh.max_column + 1)]

            if not any(vals):
                r += 1
                continue

            # Product header: "CDB DI SANTANDER" | "Valor Total: R$..." | "Disponível para Resgate: R$..."
            if (
                vals[0]
                and "CDB" in vals[0].upper()
                and "Valor Total:" in (vals[1] if len(vals) > 1 else "")
            ):
                current_product = vals[0]
                vt_m = (
                    re.search(r"Valor Total:\s*R\$\s*([\d.,]+)", vals[1]) if len(vals) > 1 else None
                )
                current_valor_total = parse_brl(vt_m.group(1)) if vt_m else None
                dr_m = (
                    re.search(r"Dispon[ií]vel para Resgate:\s*R\$\s*([\d.,]+)", vals[2])
                    if len(vals) > 2
                    else None
                )
                current_resgate = parse_brl(dr_m.group(1)) if dr_m else None
                r += 1
                continue

            # Column header row — skip
            if vals[0] == "Operação":
                r += 1
                continue

            # Data row: operation number | "R$NNN.NNN,NN" | "R$NNN.NNN,NN"
            if current_product and vals[0] and re.match(r"^\d{15,}$", vals[0]):
                n_operacao = vals[0]
                valor_str = vals[1] if len(vals) > 1 else ""
                resgate_str = vals[2] if len(vals) > 2 else ""

                valor_m = re.search(r"R\$\s*([\d.,]+)", valor_str)
                resgate_m = re.search(r"R\$\s*([\d.,]+)", resgate_str)

                valor = parse_brl(valor_m.group(1)) if valor_m else current_valor_total
                resgate = parse_brl(resgate_m.group(1)) if resgate_m else current_resgate

                posicao = {
                    "nome": f"{current_product} - Op. {n_operacao}",
                    "tipo": current_product,
                    "n_operacao": n_operacao,
                    "valor_total": valor,
                    "valor_atual": valor,
                    "valor_resgate_disponivel": resgate,
                }
                result["posicoes"].append(posicao)

            r += 1

    except Exception as e:
        log(LOG_PREFIX_EXTRATO, "ERROR", f"  Falha ao processar Santander CDB XLSX {filename}: {e}")
        result["notas"].append(f"Erro no parsing: {e}")
        result["requires_llm_fallback"] = True

    n_pos = len(result["posicoes"])
    saldo = result.get("saldo_atual", 0) or 0
    log(LOG_PREFIX_EXTRATO, "INFO", f"  → {n_pos} posições CDB Santander, total R$ {saldo:,.2f}")
    apply_cdb_checksum(result, result.get("saldo_atual"))
    return result


# ---------------------------------------------------------------------------
# Extrato: Conta corrente PDF (pdfplumber)
# ---------------------------------------------------------------------------


def parse_santander_conta(pdf_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Santander bank account statement."""
    log(LOG_PREFIX_EXTRATO, "INFO", f"Parsing Santander Conta: {filename}")
    result = make_result_template(BANCO_SANTANDER, "extratoconta", "BRL")
    result["tipo_conta"] = "corrente"

    periodo_inicio, periodo_fim = infer_periodo_from_filename(filename)
    result["periodo"]["inicio"] = periodo_inicio
    result["periodo"]["fim"] = periodo_fim

    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_text = ""
            for page in pdf.pages:
                text = page.extract_text() or ""
                all_text += text + "\n"

            result["raw_rows_detected"] = count_candidate_rows(all_text)
            result["titular"] = detect_member_from_text(all_text)

            # Account: "Agência e Conta: 1652 / 01001341-6"
            m = re.search(r"Ag[êe]ncia\s+e\s+Conta[:\s]+([\d\s/\-]+)", all_text)
            if m:
                raw_conta = m.group(1).strip()
                result["numero_conta"] = raw_conta
                ag_m = re.match(r"(\d+)", raw_conta)
                if ag_m:
                    result["agencia"] = ag_m.group(1)

            # Periodo: "Período: DD/MM/YYYY a DD/MM/YYYY"
            pm = re.search(
                r"Per[ií]odo[:\s]+(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})", all_text
            )
            if pm:
                p1 = pm.group(1).split("/")
                p2 = pm.group(2).split("/")
                result["periodo"]["inicio"] = f"{p1[2]}-{p1[1]}-{p1[0]}"
                result["periodo"]["fim"] = f"{p2[2]}-{p2[1]}-{p2[0]}"

            # Check for "SALDO ANTERIOR" only with no transactions
            saldo_ant_match = re.search(
                r"(\d{2}/\d{2}/\d{4})\s+SALDO ANTERIOR\s+(-?[\d.,]+)", all_text
            )
            if saldo_ant_match:
                result["saldo_inicial"] = parse_brl(saldo_ant_match.group(2))

            # Santander format: each line is a complete transaction
            # "DD/MM/YYYY DESCRIÇÃO DOCTO SITUAÇÃO CRÉDITO DÉBITO SALDO"
            # Credit appears as positive number, Debit as negative with "-"
            # Example: "06/02/2026 PIX RECEBIDO DAVID... 000000 5.000,00 118,34"
            # Example: "06/02/2026 DEBITO AUT. TELEFONE... 000000 -338,00 -4.881,66"

            tx_pattern = re.compile(
                r"^(\d{2}/\d{2}/\d{4})\s+"  # Date
                r"(.+?)\s+"  # Description
                r"(\d{6})\s*"  # Docto (6 digits)
                r"(-?[\d.,]+)\s+"  # Value (credit positive, debit negative)
                r"(-?[\d.,]+)\s*$",  # Saldo
                re.MULTILINE,
            )

            saldo_values: List[Tuple[str, float]] = []

            for m in tx_pattern.finditer(all_text):
                date_parts = m.group(1).split("/")
                iso_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
                descricao = m.group(2).strip()
                valor = parse_brl(m.group(4))
                saldo = parse_brl(m.group(5))

                if valor is None:
                    continue

                result["transacoes"].append(
                    {
                        "data": iso_date,
                        "descricao": descricao,
                        "valor": valor,
                    }
                )

                if saldo is not None:
                    saldo_values.append((iso_date, saldo))

            # Santander lists newest first; reverse to chronological
            result["transacoes"].reverse()
            saldo_values.reverse()

            # Note if no transactions found (legitimate zero-activity period)
            if not result["transacoes"] and result["saldo_inicial"] is not None:
                result["notas"].append(
                    "Conta sem movimentação no período (apenas saldo anterior registrado)"
                )

            # Saldo anterior from previous search
            sa_pattern = re.search(
                r"Saldo anterior.*?Saldo \(R\$\)\s*\n\s*(\d{2}/\d{2}/\d{4})\s+(-?[\d.,]+)",
                all_text,
                re.DOTALL,
            )
            if sa_pattern:
                result["saldo_inicial"] = parse_brl(sa_pattern.group(2))
            elif saldo_values:
                # Earliest saldo minus earliest transaction = initial
                result["saldo_inicial"] = saldo_values[0][1] - (
                    result["transacoes"][0]["valor"] if result["transacoes"] else 0
                )

            if saldo_values:
                result["saldo_final"] = saldo_values[-1][1]

    except Exception as e:
        log(LOG_PREFIX_EXTRATO, "ERROR", f"  Falha ao processar {filename}: {e}")
        result["notas"].append(f"Erro no parsing: {e}")
        result["requires_llm_fallback"] = True

    log(LOG_PREFIX_EXTRATO, "INFO", f"  → {len(result['transacoes'])} transações extraídas")
    return result


# ---------------------------------------------------------------------------
# Fatura: Unique CSV
# ---------------------------------------------------------------------------


def parse_santander_fatura_csv(csv_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Santander Unique credit card invoice from CSV export.

    CSV structure:
      - UTF-8 BOM header
      - Separator: comma (,)
      - 3 columns: data, lançamento, valor
      - Date format: YYYY-MM-DD
      - Valor: float with dot as decimal separator
      - Negative values = payments (PAGAMENTO EFETUADO) or credits (ESTORNO)
      - Positive values = purchases/charges
    """
    import csv as csv_mod

    log(LOG_PREFIX_FATURA, "INFO", f"Parsing Santander Unique CSV: {filename}")

    result = {
        "banco": BANCO_SANTANDER,
        "tipo": "faturaunique",
        "cartao": CARTAO_UNIQUE,
        "titular": None,
        "moeda": "BRL",
        "data_vencimento": None,
        "saldo_anterior": None,
        "total_compras": None,
        "pagamentos": None,
        "saldo_atual": None,
        "transacoes": [],
        "cartoes": [],
    }

    # Infer vencimento from filename: santander_faturaunique_YYYYMM-0_original.csv
    # Token ancorado ao fim do stem (documents.period via routing) — busca livre
    # de 6 dígitos casava o prefixo sha256[:12] e gerava 2100/1899 (A32.l3).
    ref_year, ref_month = infer_fatura_ref_from_filename(filename)
    if ref_year and ref_month:
        result["data_vencimento"] = safe_date(ref_year, ref_month, VENC_UNIQUE)

    # Read CSV (handle BOM)
    raw_text = csv_path.read_text(encoding="utf-8-sig")
    reader = csv_mod.reader(raw_text.splitlines(), delimiter=",")

    # Consume header
    header = next(reader, None)
    if not header or "data" not in header[0].lower():
        log(LOG_PREFIX_FATURA, "WARN", f"  Header CSV inesperado: {header}")
        return result

    total_compras = 0.0
    total_pagamentos = 0.0

    for row in reader:
        if len(row) < 3:
            continue

        data_str = row[0].strip()
        descricao = row[1].strip()
        valor_str = row[2].strip()

        if not data_str or not valor_str:
            continue

        # Parse date (YYYY-MM-DD format)
        date_match = re.match(r"(\d{4})-(\d{2})-(\d{2})", data_str)
        if not date_match:
            # Try DD/MM/YYYY fallback
            date_match2 = re.match(r"(\d{2})/(\d{2})/(\d{4})", data_str)
            if date_match2:
                iso_date = f"{date_match2.group(3)}-{date_match2.group(2)}-{date_match2.group(1)}"
            else:
                log(LOG_PREFIX_FATURA, "WARN", f"  Data não reconhecida: {data_str}")
                continue
        else:
            iso_date = data_str

        # Parse valor (dot as decimal separator in CSV)
        try:
            valor = float(valor_str)
        except ValueError:
            # Try Brazilian format (1.234,56)
            val = parse_brl(valor_str)
            if val is not None:
                valor = val
            else:
                log(LOG_PREFIX_FATURA, "WARN", f"  Valor não reconhecido: {valor_str}")
                continue

        # Classify: negative = payment/credit, positive = purchase/debit
        if valor < 0:
            total_pagamentos += valor
            tipo_txn = "pagamento" if "PAGAMENTO" in descricao.upper() else "estorno"
        else:
            total_compras += valor
            tipo_txn = "compra"

        # Detect IOF
        if "IOF" in descricao.upper():
            tipo_txn = "iof"

        txn = {
            "data": iso_date,
            "descricao": descricao,
            "valor": round(abs(valor), 2),
            "tipo": tipo_txn,
        }
        # Keep sign convention: purchases positive, payments negative
        if valor < 0:
            txn["valor"] = -round(abs(valor), 2)

        result["transacoes"].append(txn)

    # Summary. `total_compras` = Σ das próprias linhas → NÃO emite
    # `total_lancamentos_conferivel` (checksum tautológico: Σtx == Σtx sempre passa,
    # daria selo falso). O CSV não tem subtotal independente p/ reconciliar — uma
    # linha malformada some com o total junto. Opt-in só onde há total impresso
    # independente (parse_santander_unique PDF, ADR-342 §Emenda 2026-07-24).
    result["total_compras"] = round(total_compras, 2) if total_compras else None
    result["pagamentos"] = round(total_pagamentos, 2) if total_pagamentos else None
    saldo = total_compras + total_pagamentos
    result["saldo_atual"] = round(saldo, 2)

    n_txns = len(result["transacoes"])
    log(LOG_PREFIX_FATURA, "INFO", f"  → {n_txns} transações, saldo R$ {saldo:,.2f}")

    # Parse quality
    if n_txns > 0:
        result["parse_quality"] = "ok"
    else:
        result["parse_quality"] = "empty_csv"
        log(LOG_PREFIX_FATURA, "WARN", f"  CSV vazio (0 transações): {filename}")

    return result


# ---------------------------------------------------------------------------
# Fatura: Unique PDF (pdfplumber)
# ---------------------------------------------------------------------------

# pdfplumber funde a coluna direita (Resumo/CET) sobre as linhas de tx no
# layout lado-a-lado (#3e41/#4bb2). A tx real vem SEMPRE primeiro (coluna
# esquerda); o lixo começa num destes marcadores. Estripar ANTES do
# tx_pattern — senão o `$`-âncora captura o número poluído à direita (o
# pagamento -119,21 virava +119,21, corrupção silenciosa de sinal).
_RESUMO_POLLUTION = (" (+)", " (-)", " (=)", " Saldo", " Total", " COTAÇÃO")


def _strip_resumo_pollution(line: str) -> str:
    hits = [i for i in (line.find(m) for m in _RESUMO_POLLUTION) if i > 0]
    return line[: min(hits)].rstrip() if hits else line


def parse_santander_unique(pdf_path: Path, filename: str) -> Dict[str, Any]:
    """Parse Santander Unique credit card invoice."""
    log(LOG_PREFIX_FATURA, "INFO", f"Parsing Santander Unique: {filename}")

    # Token ancorado ao fim do stem (documents.period via routing) — busca livre
    # de 6 dígitos casava o prefixo sha256[:12] e gerava 2100/1899 (A32.l3).
    ref_year, ref_month = infer_fatura_ref_from_filename(filename)

    result = {
        "banco": BANCO_SANTANDER,
        "tipo": "faturaunique",
        "cartao": CARTAO_UNIQUE,
        "titular": None,
        "moeda": "BRL",
        "data_vencimento": None,
        "saldo_anterior": None,
        "total_compras": None,
        "pagamentos": None,
        "saldo_atual": None,
        "transacoes": [],
        "cartoes": [],
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_text = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_text.append(text)

            full_text = "\n".join(all_text)

            # --- Header ---
            _unique_regex = TITULAR.get("regex_nome_fatura", {}).get("santander_unique", "")
            m = re.search(_unique_regex, full_text) if _unique_regex else None
            if m:
                result["titular"] = m.group(0).strip()

            # Total a Pagar + Vencimento: Santander layout has headers on one line,
            # values on another: "Total a Pagar  Vencimento  Melhor Data..."
            #                    "R$ 372,85      15/01/2026  10/02/2026"
            m = re.search(r"R\$\s*([\d.,]+)\s+(\d{2}/\d{2}/\d{4})\s+\d{2}/\d{2}/\d{4}", full_text)
        if m:
            result["saldo_atual"] = parse_brl(m.group(1))
            venc_parts = m.group(2).split("/")
            if len(venc_parts) == 3:
                result["data_vencimento"] = f"{venc_parts[2]}-{venc_parts[1]}-{venc_parts[0]}"

        # Fallback: try simpler patterns
        if result["saldo_atual"] is None:
            m = re.search(r"R\$\s*([\d.,]+)\s+\d{2}/\d{2}/\d{4}", full_text)
            if m:
                result["saldo_atual"] = parse_brl(m.group(1))
        if result["saldo_atual"] is None:
            m = re.search(r"Total a Pagar\s*\n?\s*R\$\s*([\d.,]+)", full_text)
            if m:
                result["saldo_atual"] = parse_brl(m.group(1))
        if result["data_vencimento"] is None and ref_year and ref_month:
            # Fallback: buscar data DD/MM/YYYY que seja coerente com ref_year/ref_month
            candidates = re.findall(r"(\d{2})/(\d{2})/(\d{4})", full_text)
            for dd, mm, yyyy in candidates:
                if int(yyyy) == ref_year and int(mm) == ref_month:
                    result["data_vencimento"] = f"{yyyy}-{mm}-{dd}"
                    break
            # Se nenhuma data do mês correto, usar ref_year-ref_month-15 como estimativa
            if result["data_vencimento"] is None:
                result["data_vencimento"] = f"{ref_year}-{ref_month:02d}-15"
                log(
                    LOG_PREFIX_FATURA,
                    "WARN",
                    f"  Vencimento estimado (sem match exato): {result['data_vencimento']}",
                )

        # Saldo Anterior
        m = re.search(r"Saldo Anterior\s+([\d.,]+)", full_text)
        if m:
            result["saldo_anterior"] = parse_brl(m.group(1))

        # Total Despesas (Brasil + Exterior — dois subtotais separados no doc)
        m = re.search(r"Total Despesas/Débitos no Brasil\s+([\d.,]+)", full_text)
        if m:
            result["total_compras"] = parse_brl(m.group(1))
        m = re.search(r"Total Despesas/Débitos no Exterior\s+([\d.,]+)", full_text)
        if m:
            result["total_exterior"] = parse_brl(m.group(1))

        # Total pagamentos (sempre negativo por convenção — reduzem saldo da fatura)
        m = re.search(r"Total de pagamentos\s+([\d.,]+)", full_text)
        if m:
            val = parse_brl(m.group(1))
            result["pagamentos"] = -abs(val) if val else None

        # --- Transactions by card holder ---
        # pdfplumber merges left+right columns into single lines, so we need
        # to handle "polluted" lines where right-column text is appended.
        #
        # Card sections: "CARDHOLDER NAME - XXXX XXXX XXXX XXXX"
        # Transactions: "[prefix] DD/MM DESCRIPTION VALUE [USD_VALUE]"
        # Prefix can be "1 " or "□ " from checkbox icons

        card_section_pattern = re.compile(
            r"(?:@ )?([A-ZÇÃÕÉ][A-ZÇÃÕÉ\s]+?)\s*-\s*(\d{4}\s+XXXX\s+XXXX\s+\d{4})",
        )

        # Transaction pattern — allows optional leading "1 " or similar prefix,
        # optional negative sign, and trailing junk from right column
        tx_pattern = re.compile(
            r"^\s*(?:\d\s+)?(\d{2}/\d{2})\s+(.+?)\s+(-?[\d.,]+)(?:\s+([\d.,]+))?\s*(?:\s+\(\+\).*|\s+\(\-\).*|\s+\(=\).*)?$"
        )

        detail_start = full_text.find("Detalhamento da Fatura")
        if detail_start < 0:
            detail_start = 0
        detail_text = full_text[detail_start:]

        current_card = None
        current_section_type = None

        for raw_line in detail_text.split("\n"):
            # Card header — cada cartão reinicia suas subseções.
            card_m = card_section_pattern.search(raw_line)
            if card_m:
                current_card = f"{card_m.group(1).strip()} - {card_m.group(2).strip()}"
                current_section_type = None

            # Seção por header específico. "Total de pagamentos" (poluição do
            # Resumo, fundido na linha) contém "pagamento" e re-disparava a
            # seção → só o header literal "Pagamento e Demais Créditos" conta.
            if "Pagamento e Demais" in raw_line:
                current_section_type = "pagamento"
            elif re.match(r"^\s*Despesas\b", raw_line):
                current_section_type = "despesas"

            if not current_card:
                continue

            # IOF DESPESA NO EXTERIOR (linha sem data). O emissor conta o IOF em
            # "Total Despesas/Débitos no Brasil" (verificado no corpus) → escopo
            # despesa_brasil, não exterior.
            if "IOF DESPESA NO EXTERIOR" in raw_line and not re.match(r"\s*\d{2}/\d{2}", raw_line):
                iof_m = re.search(r"IOF DESPESA NO EXTERIOR\s+([\d.,]+)", raw_line)
                if iof_m:
                    iof_date = result.get("data_vencimento")
                    if result.get("transacoes"):
                        last_date = result["transacoes"][-1].get("data")
                        if last_date:
                            iof_date = last_date
                    result["transacoes"].append(
                        {
                            "data": iof_date,
                            "descricao": "IOF DESPESA NO EXTERIOR",
                            "valor": parse_brl(iof_m.group(1)),
                            "cartao": current_card,
                            "escopo": "despesa_brasil",
                        }
                    )
                continue

            line = _strip_resumo_pollution(raw_line)
            tx_m = tx_pattern.match(line)
            if tx_m:
                date_parts = tx_m.group(1).split("/")
                dd = int(date_parts[0])
                mm = int(date_parts[1])

                raw_desc = tx_m.group(2).strip()
                valor_brl = parse_brl(tx_m.group(3).strip())
                valor_usd = (
                    parse_brl(tx_m.group(4)) if tx_m.group(4) and tx_m.group(4).strip() else None
                )

                if valor_brl is None:
                    continue

                date_str = resolve_date_ddmm(dd, mm, ref_year, ref_month)

                # Escopo = balde de completude (checksum por seção). Pagamento
                # (débito da fatura anterior) sai como tipo=pagamento p/ E3/E4
                # tratar como transferência interna, nunca despesa. Sinal em E2 =
                # espaço-do-doc (magnitude impressa); normalização de fluxo é E3/E4.
                if current_section_type == "pagamento":
                    escopo, tipo = "pagamento", "pagamento"
                elif valor_usd:
                    escopo, tipo = "exterior", None
                else:
                    escopo, tipo = "despesa_brasil", None

                tx = {
                    "data": date_str,
                    "descricao": raw_desc,
                    "valor": valor_brl,
                    "cartao": current_card,
                    "escopo": escopo,
                }
                if tipo:
                    tx["tipo"] = tipo
                if valor_usd:
                    tx["forex"] = {"moeda_original": "USD", "valor_original": valor_usd}

                result["transacoes"].append(tx)

        # Opt-in do checksum de completude (ADR-342 §Emenda 2026-07-24): um signal
        # por seção impressa (despesa_brasil + exterior). O gate confere Σ(tx do
        # escopo) == subtotal; pagamento fica fora (escopo próprio). Lista quando
        # há exterior; objeto único quando só Brasil (retrocompat).
        signals = []
        if result.get("total_compras") is not None:
            signals.append(
                {"valor_cents": round(result["total_compras"] * 100), "escopo": "despesa_brasil"}
            )
        if result.get("total_exterior") is not None:
            signals.append(
                {"valor_cents": round(result["total_exterior"] * 100), "escopo": "exterior"}
            )
        if signals:
            result["total_lancamentos_conferivel"] = signals if len(signals) > 1 else signals[0]

        log(LOG_PREFIX_FATURA, "INFO", f"  → {len(result['transacoes'])} transações extraídas")
    except Exception as e:
        log(LOG_PREFIX_FATURA, "ERROR", f"  Falha ao abrir PDF {pdf_path.name}: {e}")
        return {"erro": str(e), "requires_llm_fallback": True, "tipo": "fatura"}

    return result


# ---------------------------------------------------------------------------
# CDB — posição em PDF ("Detalhes do Investimento") — A38.l12
# ---------------------------------------------------------------------------

# "CDB DI SANTANDER Valor Total : R$ 143.248,51 Disponível para Resgate : R$\n138.304,04"
# O valor de resgate pode quebrar para a linha seguinte (DOTALL).
_SANT_CDB_PRODUTO_RE = re.compile(
    r"([A-Z][A-Za-zÀ-Ú0-9 ]+?)\s+Valor\s+Total\s*:\s*R\$\s*([\d.,]+)"
    r"\s+Dispon[íi]vel\s+para\s+Resgate\s*:\s*R\$\s*([\d.,]+)",
    re.IGNORECASE | re.DOTALL,
)
_SANT_CDB_TOTAL_RE = re.compile(r"CDB\s+Valor\s+total\s*\(R\$\)\s*:\s*([\d.,]+)", re.IGNORECASE)


def _santander_cdb_posicoes(full_text: str) -> List[Dict[str, Any]]:
    return [
        {
            "nome": re.sub(r"\s+", " ", m.group(1)).strip(),
            "valor_atual": parse_brl(m.group(2)),
            "valor_resgate_disponivel": parse_brl(m.group(3)),
        }
        for m in _SANT_CDB_PRODUTO_RE.finditer(full_text)
    ]


def parse_santander_cdb_pdf(pdf_path: Path, filename: str) -> Dict[str, Any]:
    """Posição de CDB Santander em PDF; emite cdbresumo + posicoes, checksum Σ==total (ADR-342)."""
    log(LOG_PREFIX_EXTRATO, "INFO", f"Parsing Santander CDB PDF: {filename}")
    result = new_cdb_position_result(BANCO_SANTANDER)
    full_text = read_pdf_text(pdf_path)
    if not full_text:
        result["requires_llm_fallback"] = True
        return result

    result["titular"] = detect_member_from_text(full_text)
    result["posicoes"] = _santander_cdb_posicoes(full_text)
    total_m = _SANT_CDB_TOTAL_RE.search(full_text)
    total_declarado = parse_brl(total_m.group(1)) if total_m else None
    apply_cdb_checksum(result, total_declarado)
    log(LOG_PREFIX_EXTRATO, "INFO", f"  → {len(result['posicoes'])} posições de CDB")
    return result
