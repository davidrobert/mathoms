#!/usr/bin/env python3
"""
Extract patrimonial data from IRPF declarations, receipts, and real estate records.
Stage E1.5 of Financas Familia pipeline.
"""

import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional
import logging

# PDF libraries
import pdfplumber
import openpyxl

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

BASE_DIR = Path("/sessions/sharp-cool-shannon/mnt/Financas Familia/financas-familia")
DATA_DIR = BASE_DIR / "data"
EXTRACT_DIR = BASE_DIR / "processed/E2_extracts"


class IRPFDeclarationExtractor:
    """Extract data from IRPF declarations (forms)."""

    def __init__(self, pdf_path: Path):
        self.pdf_path = pdf_path
        self.data = {
            "source": str(pdf_path.name),
            "membro": None,
            "ano_base": None,
            "bens_direitos": [],
            "total_bens": 0,
            "rendimentos": [],
            "total_rendimentos": 0,
            "deducoes": [],
            "total_deducoes": 0,
            "debitos": [],
            "total_debitos": 0,
        }

    def extract(self) -> Dict:
        """Extract all relevant data from IRPF declaration."""
        try:
            with pdfplumber.open(self.pdf_path) as pdf:
                logger.info(f"Processing {self.pdf_path.name} ({len(pdf.pages)} pages)")

                # Extract from all pages
                text = ""
                for page in pdf.pages:
                    text += page.extract_text() or ""

                # Parse the text
                self._parse_text(text)

        except Exception as e:
            logger.error(f"Failed to read {self.pdf_path.name}: {e}")

        return self.data

    def _parse_text(self, text: str):
        """Parse IRPF declaration text."""
        lines = text.split('\n')

        # Identify member name and ano_base
        self._extract_declarant_info(lines)

        # Extract assets section (Bens e Direitos)
        self._extract_assets(lines)

        # Extract income section (Rendimentos)
        self._extract_income(lines)

        # Extract deductions (Deducoes)
        self._extract_deductions(lines)

        # Extract debts
        self._extract_debts(lines)

    def _extract_declarant_info(self, lines: List[str]):
        """Extract declarant name and tax year."""
        text = '\n'.join(lines)

        # Look for CPF and name patterns
        cpf_match = re.search(r'CPF\s*[:=]?\s*(\d{3}\.\d{3}\.\d{3}-\d{2})', text)

        # Extract year - IRPF declarations have 'Ano-calendário' or 'ano de referência'
        year_match = re.search(r'(?:Ano-calendário|ano\s+de\s+referência|ano\s+base)\s*[:=]?\s*(\d{4})', text, re.IGNORECASE)
        if year_match:
            self.data["ano_base"] = int(year_match.group(1))

        # Look for name at beginning or in standard patterns
        for line in lines[:50]:
            if 'declara' in line.lower() and len(line) > 10:
                # Try to extract name from lines containing declarant info
                name_match = re.search(r'(?:Declarante|Nome|Nome do Declarante)\s*[:=]?\s*([A-Z][A-Z\s]+)', line)
                if name_match:
                    self.data["membro"] = name_match.group(1).strip()
                    break

    def _extract_assets(self, lines: List[str]):
        """Extract bens e direitos (assets)."""
        in_assets_section = False

        for i, line in enumerate(lines):
            line_lower = line.lower()

            # Detect assets section
            if 'bens e direitos' in line_lower or 'ativo imobilizado' in line_lower:
                in_assets_section = True
                continue

            # Exit assets section when hitting next section
            if in_assets_section and any(section in line_lower for section in
                    ['rendimento', 'deduçao', 'débito', 'cálculo do imposto']):
                in_assets_section = False
                break

            # Extract asset data
            if in_assets_section:
                asset = self._parse_asset_line(line)
                if asset:
                    self.data["bens_direitos"].append(asset)
                    self.data["total_bens"] += asset.get("valor_31_12", 0)

    def _parse_asset_line(self, line: str) -> Optional[Dict]:
        """Parse a single asset line."""
        line = line.strip()
        if not line or len(line) < 5:
            return None

        # Look for patterns with description and values
        # Format typically: [Code] Description ... Value at 31/12 ... Prior year value

        # Extract numeric values (values are typically formatted with dots or commas)
        values = re.findall(r'[\d.,]+', line)

        if len(values) >= 1:
            try:
                # Try to extract the main value
                main_value = self._parse_currency(values[-1] if len(values) > 0 else "0")
                prior_value = self._parse_currency(values[-2] if len(values) > 1 else "0")

                return {
                    "descricao": line[:100],
                    "valor_31_12": main_value,
                    "valor_ano_anterior": prior_value if prior_value != main_value else 0,
                }
            except:
                return None

        return None

    def _extract_income(self, lines: List[str]):
        """Extract rendimentos (income sources)."""
        in_income_section = False

        for line in lines:
            line_lower = line.lower()

            if 'rendimento' in line_lower and ('pj' in line_lower or 'clt' in line_lower or
                    'aluguel' in line_lower or 'aplicação' in line_lower):
                in_income_section = True

                # Parse income line
                income = self._parse_income_line(line)
                if income:
                    self.data["rendimentos"].append(income)
                    self.data["total_rendimentos"] += income.get("valor", 0)

            if in_income_section and any(s in line_lower for s in
                    ['deduçao', 'débito', 'cálculo', 'bens e direitos']):
                in_income_section = False

    def _parse_income_line(self, line: str) -> Optional[Dict]:
        """Parse a single income line."""
        line = line.strip()
        if not line:
            return None

        # Look for income type and value
        values = re.findall(r'[\d.,]+', line)
        if values:
            try:
                valor = self._parse_currency(values[-1])
                return {
                    "tipo": "income",
                    "descricao": line[:100],
                    "valor": valor,
                }
            except:
                return None

        return None

    def _extract_deductions(self, lines: List[str]):
        """Extract deductions (pagamentos que reduzem imposto)."""
        for i, line in enumerate(lines):
            line_lower = line.lower()

            if any(d in line_lower for d in ['previdência', 'educação', 'saúde', 'pensão']):
                values = re.findall(r'[\d.,]+', line)
                if values:
                    try:
                        valor = self._parse_currency(values[-1])
                        self.data["deducoes"].append({
                            "tipo": "deduction",
                            "descricao": line[:100],
                            "valor": valor,
                        })
                        self.data["total_deducoes"] += valor
                    except:
                        pass

    def _extract_debts(self, lines: List[str]):
        """Extract debts (financing, mortgages, etc)."""
        for line in lines:
            line_lower = line.lower()

            if any(d in line_lower for d in ['financiamento', 'empréstimo', 'hipoteca', 'débito']):
                values = re.findall(r'[\d.,]+', line)
                if values:
                    try:
                        valor = self._parse_currency(values[-1])
                        self.data["debitos"].append({
                            "tipo": "debt",
                            "descricao": line[:100],
                            "valor": valor,
                        })
                        self.data["total_debitos"] += valor
                    except:
                        pass

    def _parse_currency(self, value_str: str) -> float:
        """Parse Brazilian currency format to float."""
        # Remove spaces
        value_str = value_str.strip()
        # Handle Brazilian format: 1.234,56 (dot as thousands, comma as decimal)
        value_str = value_str.replace('.', '').replace(',', '.')
        return float(value_str)


class IRPFReceiptExtractor:
    """Extract data from IRPF receipts (recibos)."""

    def __init__(self, pdf_path: Path):
        self.pdf_path = pdf_path
        self.data = {
            "source": str(pdf_path.name),
            "membro": None,
            "imposto_total": 0,
            "data_envio": None,
            "status": None,
            "numero_recibo": None,
        }

    def extract(self) -> Dict:
        """Extract receipt data."""
        try:
            with pdfplumber.open(self.pdf_path) as pdf:
                logger.info(f"Processing receipt {self.pdf_path.name}")

                text = ""
                for page in pdf.pages:
                    text += page.extract_text() or ""

                self._parse_text(text)

        except Exception as e:
            logger.error(f"Failed to read {self.pdf_path.name}: {e}")

        return self.data

    def _parse_text(self, text: str):
        """Parse receipt text."""
        lines = text.split('\n')

        # Extract total tax
        for line in lines:
            if 'imposto' in line.lower() and 'total' in line.lower():
                values = re.findall(r'[\d.,]+', line)
                if values:
                    try:
                        self.data["imposto_total"] = self._parse_currency(values[-1])
                    except:
                        pass

            # Extract filing date
            if 'data' in line.lower() and 'envio' in line.lower():
                date_match = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', line)
                if date_match:
                    self.data["data_envio"] = f"{date_match.group(3)}-{date_match.group(2).zfill(2)}-{date_match.group(1).zfill(2)}"

            # Extract receipt number
            if 'número' in line.lower() and 'recibo' in line.lower():
                num_match = re.search(r'(\d+)', line)
                if num_match:
                    self.data["numero_recibo"] = num_match.group(1)

    def _parse_currency(self, value_str: str) -> float:
        """Parse Brazilian currency format to float."""
        value_str = value_str.strip()
        value_str = value_str.replace('.', '').replace(',', '.')
        return float(value_str)


class RealEstateExtractor:
    """Extract data from real estate spreadsheet."""

    def __init__(self, xlsx_path: Path):
        self.xlsx_path = xlsx_path
        self.data = []

    def extract(self) -> List[Dict]:
        """Extract real estate data from XLSX."""
        try:
            wb = openpyxl.load_workbook(self.xlsx_path)
            ws = wb.active

            logger.info(f"Processing real estate file: {self.xlsx_path.name}")

            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            # Extract data rows
            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    cell_value = ws.cell(row_idx, col_idx).value
                    if header:
                        row_data[header] = cell_value

                if row_data.get('Endereço') or row_data.get('endereço'):
                    self.data.append(self._normalize_property(row_data))

            logger.info(f"Extracted {len(self.data)} properties")

        except Exception as e:
            logger.error(f"Failed to read {self.xlsx_path.name}: {e}")

        return self.data

    def _normalize_property(self, row: Dict) -> Dict:
        """Normalize property data."""
        return {
            "endereco": row.get('Endereço') or row.get('endereço'),
            "data_compra": row.get('Data Compra') or row.get('data_compra'),
            "valor_compra": self._parse_currency(row.get('Valor Compra') or row.get('valor_compra') or 0),
            "vendedor": row.get('Vendedor') or row.get('vendedor'),
            "financiamento": row.get('Financiamento') or row.get('financiamento'),
            "status": row.get('Status') or row.get('status') or 'Ativo',
        }

    def _parse_currency(self, value) -> float:
        """Parse currency value."""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        value_str = str(value).strip()
        value_str = value_str.replace('.', '').replace(',', '.')
        try:
            return float(value_str)
        except:
            return 0.0


def main():
    """Main extraction pipeline."""

    # Create output directory
    EXTRACT_DIR.mkdir(parents=True, exist_ok=True)

    extracted_data = {
        "declarations": [],
        "receipts": [],
        "properties": [],
        "members": set(),
        "anos_base": set(),
    }

    # Process IRPF declarations
    declaration_files = [
        DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracao_2023-0_original.pdf",
        DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracao_2024-0_original.pdf",
        DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracaomariana_2024-0_original.pdf",
    ]

    for decl_file in declaration_files:
        if decl_file.exists():
            extractor = IRPFDeclarationExtractor(decl_file)
            data = extractor.extract()
            extracted_data["declarations"].append(data)

            if data["membro"]:
                extracted_data["members"].add(data["membro"])
            if data["ano_base"]:
                extracted_data["anos_base"].add(data["ano_base"])

            # Save individual extract
            output_file = EXTRACT_DIR / decl_file.name.replace("-0_original.pdf", "-2_extract.json")
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logger.info(f"Saved declaration extract: {output_file.name}")

    # Process IRPF receipts
    receipt_files = [
        DATA_DIR / "income_tax_br" / "receitafederal_irpfrecibo_2024-0_original.pdf",
        DATA_DIR / "income_tax_br" / "receitafederal_irpfrecibomariana_2024-0_original.pdf",
    ]

    for receipt_file in receipt_files:
        if receipt_file.exists():
            extractor = IRPFReceiptExtractor(receipt_file)
            data = extractor.extract()
            extracted_data["receipts"].append(data)

            # Save individual extract
            output_file = EXTRACT_DIR / receipt_file.name.replace("-0_original.pdf", "-2_extract.json")
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logger.info(f"Saved receipt extract: {output_file.name}")

    # Process real estate
    xlsx_file = DATA_DIR / "real_estate" / "dados_imoveis-0_original.xlsx"
    if xlsx_file.exists():
        extractor = RealEstateExtractor(xlsx_file)
        properties = extractor.extract()
        extracted_data["properties"] = properties

        # Save real estate extract
        output_file = EXTRACT_DIR / "dados_imoveis-2_extract.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(properties, f, ensure_ascii=False, indent=2)
        logger.info(f"Saved real estate extract: {output_file.name}")

    # Create consolidated baseline
    consolidated = {
        "pipeline_stage": "E1.5_Baseline_Patrimonial",
        "data_processamento": "2026-04-08",
        "membros": list(extracted_data["members"]),
        "anos_base": sorted(list(extracted_data["anos_base"])),
        "declarations": extracted_data["declarations"],
        "receipts": extracted_data["receipts"],
        "properties": extracted_data["properties"],
        "resumo": {
            "total_membros": len(extracted_data["members"]),
            "total_declarations": len(extracted_data["declarations"]),
            "total_receipts": len(extracted_data["receipts"]),
            "total_properties": len(extracted_data["properties"]),
        }
    }

    # Save consolidated file
    output_file = EXTRACT_DIR / "baseline_patrimonial-1.5_consolidated.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(consolidated, f, ensure_ascii=False, indent=2)
    logger.info(f"Saved consolidated baseline: {output_file.name}")

    logger.info("Extraction pipeline completed successfully!")
    return 0


if __name__ == "__main__":
    sys.exit(main())
