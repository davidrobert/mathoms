#!/usr/bin/env python3
"""
Extract patrimônio baseline from IRPF declarations, receipts, QuintoAndar rent reports, and real estate XLSX.
Step 6b: E1.5 - Baseline Patrimonial extraction - Enhanced version
"""

import json
import re
from datetime import datetime
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    print("Installing pdfplumber...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'pdfplumber', '--break-system-packages'])
    import pdfplumber

try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl', '--break-system-packages'])
    from openpyxl import load_workbook

BASE_DIR = Path("/sessions/sharp-cool-shannon/mnt/Financas Familia/financas-familia")
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "processed" / "E2_extracts"

# Ensure output directory exists
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def parse_valor_brasileiro(valor_str):
    """Parse Brazilian format values (1.234.567,89)."""
    if not valor_str:
        return None
    try:
        # Remove spaces and convert Brazilian format to float
        valor_str = str(valor_str).strip()
        # Replace . with empty (thousands) and , with . (decimal)
        valor_str = valor_str.replace('.', '').replace(',', '.')
        return float(valor_str)
    except:
        return None

def extract_irpf_declaration(pdf_path):
    """Extract patrimônio data from IRPF declaration PDF."""
    result = {
        "source_file": str(pdf_path),
        "extraction_type": "irpf_declaracao",
        "bens_direitos": [],
        "total_bens": None,
        "rendimentos": {
            "pj_salario": None,
            "pj_total": None,
            "alugueis": None,
            "financeiros": None,
            "outros": None
        },
        "deductions": {
            "contribuicao_previdenciaria": None,
            "pagamentos_deductiveis": None
        },
        "extraction_date": datetime.now().isoformat()
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            pages_text = []

            for page_num, page in enumerate(pdf.pages):
                page_text = page.extract_text() or ""
                full_text += page_text + "\n"
                pages_text.append((page_num + 1, page_text))

            # Extract nome e CPF
            match = re.search(r'Nome:?\s*([^\n]+)\s+CPF:?\s*(\d+\.?\d+\.?\d+[\/-]\d+)', full_text)
            if match:
                result["declarante"] = {
                    "nome": match.group(1).strip(),
                    "cpf": match.group(2).strip()
                }

            # Extract BENS E DIREITOS section
            bens_match = re.search(
                r'BENS E DIREITOS.*?\n(.*?)(?:\nDÍVIDAS|Total de|$)',
                full_text,
                re.IGNORECASE | re.DOTALL
            )

            if bens_match:
                bens_text = bens_match.group(1)

                # Parse lines with pattern: GRUPO CÓDIGO DISCRIMINAÇÃO VALOR_ANT VALOR_AT
                # Values are in the format "123.456,78" (Brazilian format)
                lines = bens_text.split('\n')

                i = 0
                while i < len(lines):
                    line = lines[i].strip()

                    # Try to match pattern: "01 11 APARTAMENTO..." with values at end
                    match = re.match(r'^(\d{2})\s+(\d{2})\s+(.+?)(?:\s+([\d.,]+)\s+([\d.,]+))?\s*$', line)
                    if match:
                        grupo = match.group(1)
                        codigo = match.group(2)
                        descricao = match.group(3).strip()

                        # Try to get values from the same line or next lines
                        valor_ant = None
                        valor_atu = None

                        # If values are at the end of current line
                        if match.group(4) and match.group(5):
                            valor_ant = parse_valor_brasileiro(match.group(4))
                            valor_atu = parse_valor_brasileiro(match.group(5))
                        else:
                            # Look for values in the line with regex
                            valores = re.findall(r'([\d.]+,\d{2})', line)
                            if len(valores) >= 2:
                                valor_ant = parse_valor_brasileiro(valores[-2])
                                valor_atu = parse_valor_brasileiro(valores[-1])
                            elif len(valores) == 1:
                                valor_atu = parse_valor_brasileiro(valores[0])

                            # If still not found, look in next lines
                            if not valor_atu and i + 1 < len(lines):
                                next_line = lines[i + 1].strip()
                                valores = re.findall(r'([\d.]+,\d{2})', next_line)
                                if len(valores) >= 2:
                                    valor_ant = parse_valor_brasileiro(valores[0])
                                    valor_atu = parse_valor_brasileiro(valores[1])
                                elif len(valores) == 1:
                                    valor_atu = parse_valor_brasileiro(valores[0])

                        bem = {
                            "grupo": grupo,
                            "codigo": codigo,
                            "descricao": descricao,
                            "valor_31_12_anterior": valor_ant,
                            "valor_31_12_atual": valor_atu
                        }
                        result["bens_direitos"].append(bem)

                    i += 1

            # Extract TOTAL DE BENS - try multiple patterns
            match = re.search(
                r'Total\s+de\s+Bens\s+e\s+Direitos\s*[:\-]?\s*(\d{1,3}(?:\.\d{3})*,\d{2})',
                full_text,
                re.IGNORECASE
            )
            if match:
                result["total_bens"] = parse_valor_brasileiro(match.group(1))
            else:
                # Calculate from bens_direitos if not found
                total = sum(b.get("valor_31_12_atual") or 0 for b in result["bens_direitos"])
                if total > 0:
                    result["total_bens"] = total

            # Extract rendimentos
            # PJ - Salário
            match = re.search(
                r'Total\s+(?:de\s+)?(?:Rendimentos\s+)?Recebidos?.*?PJ.*?\n.*?(\d{1,3}(?:\.\d{3})*,\d{2})',
                full_text,
                re.IGNORECASE | re.DOTALL
            )
            if match:
                result["rendimentos"]["pj_total"] = parse_valor_brasileiro(match.group(1))

            # Aluguéis
            match = re.search(
                r'Aluguéis?.*?\n.*?(\d{1,3}(?:\.\d{3})*,\d{2})',
                full_text,
                re.IGNORECASE | re.DOTALL
            )
            if match:
                result["rendimentos"]["alugueis"] = parse_valor_brasileiro(match.group(1))

            # Rendimentos financeiros
            match = re.search(
                r'(?:Rendimentos?|Juros)\s+Financeiros?.*?\n.*?(\d{1,3}(?:\.\d{3})*,\d{2})',
                full_text,
                re.IGNORECASE | re.DOTALL
            )
            if match:
                result["rendimentos"]["financeiros"] = parse_valor_brasileiro(match.group(1))

            # Contribuição previdenciária
            match = re.search(
                r'Contribuição\s+(?:à\s+)?Seguridade\s+Social\s+(?:do\s+Contribuinte\s+Individual)?.*?\n.*?(\d{1,3}(?:\.\d{3})*,\d{2})',
                full_text,
                re.IGNORECASE | re.DOTALL
            )
            if match:
                result["deductions"]["contribuicao_previdenciaria"] = parse_valor_brasileiro(match.group(1))

    except Exception as e:
        result["error"] = str(e)

    return result

def extract_irpf_receipt(pdf_path):
    """Extract data from IRPF receipt PDF."""
    result = {
        "source_file": str(pdf_path),
        "extraction_type": "irpf_recibo",
        "numero_recibo": None,
        "data_processamento": None,
        "imposto_total": None,
        "imposto_devido": None,
        "extraction_date": datetime.now().isoformat()
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            for page in pdf.pages:
                full_text += page.extract_text() or ""

            # Extract número do recibo
            match = re.search(r'Recibo\s+(?:Número|nº)?\s*[:\-]?\s*(\d+[\.\-]\d+[\.\-]\d+[\.\-]\d+[\.\-]\d+[\.\-]\d+)',
                            full_text, re.IGNORECASE)
            if match:
                result["numero_recibo"] = match.group(1)

            # Extract data de processamento
            match = re.search(r'Data\s+(?:de\s+)?Processamento\s*[:\-]?\s*(\d{2}[/-]\d{2}[/-]\d{4})',
                            full_text, re.IGNORECASE)
            if match:
                result["data_processamento"] = match.group(1)

            # Extract imposto total
            match = re.search(r'Imposto\s+Total\s*[:\-]?\s*(\d{1,3}(?:\.\d{3})*,\d{2})',
                            full_text, re.IGNORECASE)
            if match:
                result["imposto_total"] = parse_valor_brasileiro(match.group(1))

            # Extract imposto devido
            match = re.search(r'Imposto\s+(?:Devido|a\s+Pagar|Restituição)\s*[:\-]?\s*(\d{1,3}(?:\.\d{3})*,\d{2})',
                            full_text, re.IGNORECASE)
            if match:
                result["imposto_devido"] = parse_valor_brasileiro(match.group(1))

    except Exception as e:
        result["error"] = str(e)

    return result

def extract_quintoandar_rent(pdf_path):
    """Extract data from QuintoAndar rent report PDF."""
    result = {
        "source_file": str(pdf_path),
        "extraction_type": "quintoandar_aluguel",
        "properties": [],
        "ano_referencia": None,
        "extraction_date": datetime.now().isoformat()
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            for page in pdf.pages:
                full_text += page.extract_text() or ""

            # Extract year
            match = re.search(r'(?:Ano|Year|2024|2025|2026)', full_text)
            if match:
                match_ano = re.search(r'(202[0-9])', full_text)
                if match_ano:
                    result["ano_referencia"] = int(match_ano.group(1))

            # Look for property sections
            # Pattern: property address followed by values
            prop_matches = re.finditer(
                r'(?:Imóvel|Property|Endereço).*?\n(.*?)(?:\n\n|\nImóvel|\n$)',
                full_text,
                re.IGNORECASE | re.DOTALL
            )

            for match in prop_matches:
                prop_section = match.group(1)

                # Extract values
                valores = re.findall(r'(\d{1,3}(?:\.\d{3})*,\d{2})', prop_section)

                if valores:
                    prop = {
                        "descricao": prop_section.split('\n')[0].strip(),
                        "renda_bruta": parse_valor_brasileiro(valores[0]) if len(valores) > 0 else None,
                        "descontos": parse_valor_brasileiro(valores[1]) if len(valores) > 1 else None,
                        "renda_liquida": parse_valor_brasileiro(valores[2]) if len(valores) > 2 else None,
                        "ano": result["ano_referencia"] or 2025
                    }
                    result["properties"].append(prop)

            # If no properties found, try simpler pattern
            if not result["properties"]:
                valores = re.findall(r'(\d{1,3}(?:\.\d{3})*,\d{2})', full_text)
                if len(valores) >= 3:
                    result["properties"].append({
                        "descricao": "Propriedade não identificada",
                        "renda_bruta": parse_valor_brasileiro(valores[0]),
                        "descontos": parse_valor_brasileiro(valores[1]) if len(valores) > 1 else None,
                        "renda_liquida": parse_valor_brasileiro(valores[2]) if len(valores) > 2 else None,
                        "ano": result["ano_referencia"] or 2025
                    })

    except Exception as e:
        result["error"] = str(e)

    return result

def extract_real_estate_xlsx(xlsx_path):
    """Extract data from real estate XLSX file."""
    result = {
        "source_file": str(xlsx_path),
        "extraction_type": "real_estate",
        "properties": [],
        "extraction_date": datetime.now().isoformat()
    }

    try:
        wb = load_workbook(xlsx_path)

        # Try to find the main data sheet
        sheet = None
        for sheet_name in wb.sheetnames:
            if any(keyword in sheet_name.lower() for keyword in ['dados', 'imovel', 'property', 'properties', 'real estate']):
                sheet = wb[sheet_name]
                break

        if not sheet:
            sheet = wb.active

        # Extract headers
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower())

        # Extract rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if not any(cell.value for cell in row[:3]):  # Skip empty rows
                continue

            prop = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    header = headers[col_idx]
                    value = cell.value

                    # Convert datetime objects to string
                    if isinstance(value, datetime):
                        value = value.isoformat()

                    # Map common headers
                    if 'endereco' in header or 'address' in header or 'local' in header:
                        prop["endereco"] = value
                    elif 'data' in header and 'aquisicao' in header:
                        prop["data_aquisicao"] = str(value) if value else None
                    elif 'valor' in header and 'compra' in header:
                        try:
                            prop["valor_compra"] = float(value) if value else None
                        except:
                            prop["valor_compra"] = value
                    elif 'financiamento' in header:
                        try:
                            prop["financiamento"] = float(value) if value else None
                        except:
                            prop["financiamento"] = value
                    elif 'status' in header:
                        prop["status"] = value
                    else:
                        # Store all other columns
                        if header:
                            prop[header] = value

            if prop:
                result["properties"].append(prop)

    except Exception as e:
        result["error"] = str(e)

    return result

def main():
    """Main extraction function."""

    print("=" * 80)
    print("STEP 6b: E1.5 - Baseline Patrimonial Extraction (Enhanced)")
    print("=" * 80)

    # Define input files and their output names
    files_to_process = [
        (DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracao_2023-0_original.pdf",
         "receitafederal_irpfdeclaracao_2023-2_extract.json",
         "irpf_declaracao"),

        (DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracao_2024-0_original.pdf",
         "receitafederal_irpfdeclaracao_2024-2_extract.json",
         "irpf_declaracao"),

        (DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracaomariana_2024-0_original.pdf",
         "receitafederal_irpfdeclaracaomariana_2024-2_extract.json",
         "irpf_declaracao"),

        (DATA_DIR / "income_tax_br" / "receitafederal_irpfrecibo_2024-0_original.pdf",
         "receitafederal_irpfrecibo_2024-2_extract.json",
         "irpf_recibo"),

        (DATA_DIR / "income_tax_br" / "receitafederal_irpfrecibomariana_2024-0_original.pdf",
         "receitafederal_irpfrecibomariana_2024-2_extract.json",
         "irpf_recibo"),

        (DATA_DIR / "financial_statements" / "quintoandar_informerendimentosaluguel_2025-0_original.pdf",
         "quintoandar_informerendimentosaluguel_2025-2_extract.json",
         "quintoandar_aluguel"),

        (DATA_DIR / "financial_statements" / "quintoandar_informerendimentosaluguelmariana_2025-0_original.pdf",
         "quintoandar_informerendimentosaluguelmariana_2025-2_extract.json",
         "quintoandar_aluguel"),

        (DATA_DIR / "real_estate" / "dados_imoveis-0_original.xlsx",
         "dados_imoveis-2_extract.json",
         "real_estate"),
    ]

    extracted_data = {}
    declarations = []
    receipts = []
    all_properties = []

    # Process each file
    for input_path, output_name, file_type in files_to_process:
        print(f"\nProcessing: {input_path.name}")
        print(f"  Type: {file_type}")

        if not input_path.exists():
            print(f"  ERROR: File not found!")
            continue

        try:
            if file_type == "irpf_declaracao":
                data = extract_irpf_declaration(input_path)
                declarations.append(data)
                if data.get("bens_direitos"):
                    print(f"  Extracted {len(data['bens_direitos'])} bens e direitos")
                if data.get("total_bens"):
                    print(f"  Total bens: R$ {data['total_bens']:,.2f}")

            elif file_type == "irpf_recibo":
                data = extract_irpf_receipt(input_path)
                receipts.append(data)
                if data.get("imposto_total"):
                    print(f"  Imposto total: R$ {data['imposto_total']:,.2f}")

            elif file_type == "quintoandar_aluguel":
                data = extract_quintoandar_rent(input_path)
                if data["properties"]:
                    all_properties.extend(data["properties"])
                    print(f"  Extracted {len(data['properties'])} properties")

            elif file_type == "real_estate":
                data = extract_real_estate_xlsx(input_path)
                if data["properties"]:
                    all_properties.extend(data["properties"])
                    print(f"  Extracted {len(data['properties'])} properties")

            # Save individual extract
            output_path = OUTPUT_DIR / output_name
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            print(f"  Saved: {output_path.name}")
            extracted_data[output_name] = data

        except Exception as e:
            print(f"  ERROR during extraction: {e}")
            import traceback
            traceback.print_exc()

    # Create consolidated baseline
    consolidated = {
        "pipeline_stage": "E1.5_Baseline_Patrimonial",
        "data_processamento": datetime.now().isoformat(),
        "membros": ["David", "Mariana"],
        "anos_base": [2023, 2024, 2025],
        "declarations": declarations,
        "receipts": receipts,
        "properties": all_properties,
        "summary": {
            "total_declarations": len(declarations),
            "total_receipts": len(receipts),
            "total_properties": len(all_properties),
            "bens_direitos_count": sum(len(d.get("bens_direitos", [])) for d in declarations),
            "total_bens_value": sum(d.get("total_bens") or 0 for d in declarations if d.get("total_bens"))
        }
    }

    # Save consolidated baseline
    consolidated_path = OUTPUT_DIR / "baseline_patrimonial-1.5_consolidated.json"
    with open(consolidated_path, 'w', encoding='utf-8') as f:
        json.dump(consolidated, f, ensure_ascii=False, indent=2)

    print("\n" + "=" * 80)
    print("CONSOLIDATION SUMMARY")
    print("=" * 80)
    print(f"Total IRPF Declarations: {len(declarations)}")
    print(f"Total IRPF Receipts: {len(receipts)}")
    print(f"Total Properties: {len(all_properties)}")
    print(f"Total Bens e Direitos: {consolidated['summary']['bens_direitos_count']}")
    print(f"Total Bens Value: R$ {consolidated['summary']['total_bens_value']:,.2f}")
    print(f"\nConsolidated output: {consolidated_path.name}")
    print("=" * 80)

if __name__ == "__main__":
    main()
