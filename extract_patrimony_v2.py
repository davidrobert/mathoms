#!/usr/bin/env python3
"""
Extract patrimonial data from IRPF declarations, receipts, and real estate records.
Stage E1.5 of Financas Familia pipeline - Detailed extraction.
"""

import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import logging

import pdfplumber
import openpyxl

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

BASE_DIR = Path("/sessions/sharp-cool-shannon/mnt/Financas Familia/financas-familia")
DATA_DIR = BASE_DIR / "data"
EXTRACT_DIR = BASE_DIR / "processed/E2_extracts"


def parse_currency(value_str: str) -> float:
    """Parse Brazilian currency format to float."""
    if not value_str:
        return 0.0
    value_str = str(value_str).strip()
    # Handle Brazilian format: 1.234,56 (dot as thousands, comma as decimal)
    value_str = value_str.replace('.', '').replace(',', '.')
    try:
        return float(value_str)
    except:
        return 0.0


def extract_irpf_declaration(pdf_path: Path) -> Dict:
    """Extract detailed data from IRPF declaration."""
    data = {
        "source": str(pdf_path.name),
        "membro": None,
        "cpf": None,
        "ano_base": None,
        "tipo_declaracao": None,
        "bens_direitos": [],
        "total_bens": 0.0,
        "rendimentos_pj": [],
        "rendimentos_pf": [],
        "rendimentos_isentos": [],
        "rendimentos_aplicacoes": [],
        "total_rendimentos": 0.0,
        "pagamentos_deductibles": [],
        "total_deducoes": 0.0,
        "debitos": [],
        "total_debitos": 0.0,
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            logger.info(f"Processing {pdf_path.name} ({len(pdf.pages)} pages)")

            # Concatenate all text
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"

            # Extract header information
            _extract_header_info(full_text, data)

            # Extract assets (bens e direitos)
            _extract_assets_section(full_text, data)

            # Extract rendimentos (income)
            _extract_income_sections(full_text, data)

            # Extract deductions and payments
            _extract_deductions_section(full_text, data)

    except Exception as e:
        logger.error(f"Failed to process {pdf_path.name}: {e}")

    return data


def _extract_header_info(text: str, data: Dict):
    """Extract name, CPF, year from header."""
    # Extract name and CPF
    name_match = re.search(r'NOME:\s*([A-Z\s]+?)(?:\n|CPF)', text)
    if name_match:
        data["membro"] = name_match.group(1).strip()

    cpf_match = re.search(r'CPF:\s*(\d{3}\.\d{3}\.\d{3}-\d{2})', text)
    if cpf_match:
        data["cpf"] = cpf_match.group(1)

    # Extract year (Ano-calendário)
    year_match = re.search(r'ANO-CALENDÁRIO\s+(\d{4})', text)
    if year_match:
        data["ano_base"] = int(year_match.group(1))

    # Extract declaration type
    if 'Original' in text:
        data["tipo_declaracao"] = "Original"
    elif 'Retificadora' in text:
        data["tipo_declaracao"] = "Retificadora"


def _extract_assets_section(text: str, data: Dict):
    """Extract bens e direitos section."""
    # Find the assets section
    assets_start = text.lower().find('declaração de bens e direitos')
    if assets_start < 0:
        return

    # Find where this section ends (next major section or end of doc)
    assets_end_markers = [
        'resumo das atividades da pes. jurídica',
        'perdas em operações de renda variável',
        'livro caixa',
    ]

    assets_end = len(text)
    for marker in assets_end_markers:
        pos = text.lower().find(marker, assets_start)
        if pos > 0:
            assets_end = pos
            break

    assets_section = text[assets_start:assets_end]

    # Parse asset lines - pattern is:
    # GRUPO CÓDIGO DISCRIMINAÇÃO ... SITUAÇÃO EM 31/12/XXXX  31/12/YYYY
    # followed by details

    lines = assets_section.split('\n')

    i = 0
    while i < len(lines):
        line = lines[i].strip()
        i += 1

        # Look for asset entries: start with group and code (e.g., "01 11")
        match = re.match(r'^(\d{2})\s+(\d{2})\s+(.+)', line)
        if match:
            group = match.group(1)
            code = match.group(2)
            description = match.group(3)

            # Extract values from this and following lines
            asset_data = {
                "grupo": group,
                "codigo": code,
                "descricao": description,
                "detalhes": {},
                "valor_31_12_anterior": 0.0,
                "valor_31_12_atual": 0.0,
            }

            # Collect following lines for this asset
            asset_lines = [line]
            while i < len(lines):
                next_line = lines[i]

                # Check if this is a new asset (starts with 2-digit group + space + 2-digit code)
                if re.match(r'^\d{2}\s+\d{2}\s+', next_line.strip()):
                    break

                # Check if line contains the values we need
                if 'SITUAÇÃO EM' in next_line or '31/12/' in next_line:
                    # Extract values from header
                    values = re.findall(r'([\d.]+,\d{2})\s*\n?\s*([\d.]+,\d{2})?', next_line)
                    if values:
                        for val_pair in values:
                            if val_pair[0]:
                                asset_data["valor_31_12_anterior"] = parse_currency(val_pair[0])
                            if val_pair[1]:
                                asset_data["valor_31_12_atual"] = parse_currency(val_pair[1])

                # Collect detail lines
                if next_line.strip() and not re.match(r'^NOME:|^CPF:', next_line):
                    asset_lines.append(next_line)
                    asset_data["detalhes"][next_line.split(':')[0].strip()] = next_line.split(':')[1].strip() if ':' in next_line else ""

                i += 1

            # Add to data if we have values
            if asset_data["valor_31_12_atual"] > 0 or asset_data["valor_31_12_anterior"] > 0:
                data["bens_direitos"].append(asset_data)
                data["total_bens"] += asset_data["valor_31_12_atual"]


def _extract_income_sections(text: str, data: Dict):
    """Extract all income sources."""

    # 1. Rendimentos de Pessoa Jurídica (CLT/PJ employment)
    pj_section_match = re.search(
        r'RENDIMENTOS TRIBUTÁVEIS RECEBIDOS DE PESSOA JURÍDICA PELO TITULAR.*?(?=RENDIMENTOS|DEDUÇÕES|$)',
        text, re.DOTALL | re.IGNORECASE
    )
    if pj_section_match:
        pj_section = pj_section_match.group(0)
        pj_lines = pj_section.split('\n')

        for line in pj_lines:
            # Look for company name and values
            if any(c.isalpha() for c in line) and any(c.isdigit() for c in line):
                values = re.findall(r'([\d.]+,\d{2})', line)
                if values:
                    # Typically: REND RECEBIDOS, CONTRIB PREVID, IMPOSTO RETIDO, 13º, IRRF 13º
                    try:
                        entry = {
                            "tipo": "PJ/CLT",
                            "descricao": re.sub(r'[\d.,]+', '', line)[:100].strip(),
                            "rend_recebidos": parse_currency(values[0]) if len(values) > 0 else 0,
                            "contrib_previd": parse_currency(values[1]) if len(values) > 1 else 0,
                            "imposto_retido": parse_currency(values[2]) if len(values) > 2 else 0,
                        }
                        if entry["rend_recebidos"] > 0:
                            data["rendimentos_pj"].append(entry)
                            data["total_rendimentos"] += entry["rend_recebidos"]
                    except:
                        pass

    # 2. Rendimentos Isentos e Não Tributáveis
    exempt_section_match = re.search(
        r'RENDIMENTOS ISENTOS E NÃO TRIBUTÁVEIS.*?TOTAL\s+([\d.]+,\d{2})',
        text, re.DOTALL | re.IGNORECASE
    )
    if exempt_section_match:
        total_exempt = parse_currency(exempt_section_match.group(1))
        data["rendimentos_isentos"].append({
            "tipo": "Isento",
            "total": total_exempt,
        })

    # 3. Rendimentos de Aplicações Financeiras
    app_section_match = re.search(
        r'13º salário.*?TOTAL\s+([\d.]+,\d{2})',
        text, re.DOTALL | re.IGNORECASE
    )
    if app_section_match:
        pass  # Covered by other sections


def _extract_deductions_section(text: str, data: Dict):
    """Extract deductible payments."""
    # Find section with payments
    payments_match = re.search(
        r'PAGAMENTOS EFETUADOS.*?(?=DOAÇÕES|DECLARAÇÃO|$)',
        text, re.DOTALL | re.IGNORECASE
    )

    if payments_match:
        payments_text = payments_match.group(0)

        # Look for payment entries (código + beneficiary + CNPJ/CPF + value)
        # Pattern: codigo NOME CNPJ/CPF VALOR
        payment_lines = payments_text.split('\n')

        for line in payment_lines:
            if any(c.isdigit() for c in line) and (',' in line or '.' in line):
                values = re.findall(r'([\d.]+,\d{2})', line)
                if values:
                    try:
                        amount = parse_currency(values[0])
                        if amount > 0:
                            data["pagamentos_deductibles"].append({
                                "descricao": line[:100].strip(),
                                "valor": amount,
                            })
                            data["total_deducoes"] += amount
                    except:
                        pass


def extract_irpf_receipt(pdf_path: Path) -> Dict:
    """Extract data from IRPF receipt (recibo)."""
    data = {
        "source": str(pdf_path.name),
        "membro": None,
        "cpf": None,
        "imposto_total": 0.0,
        "data_envio": None,
        "status": None,
        "numero_recibo": None,
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            logger.info(f"Processing receipt {pdf_path.name}")

            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"

            # Extract name
            name_match = re.search(r'(?:Contribuinte|Declarante|Nome):\s*([A-Z\s]+?)(?:\n|CPF)', full_text)
            if name_match:
                data["membro"] = name_match.group(1).strip()

            # Extract CPF
            cpf_match = re.search(r'CPF:\s*(\d{3}\.\d{3}\.\d{3}-\d{2})', full_text)
            if cpf_match:
                data["cpf"] = cpf_match.group(1)

            # Extract receipt number
            recibo_match = re.search(r'(?:Número|Nº).*?Recibo\s*[:=]?\s*(\d+)', full_text, re.IGNORECASE)
            if recibo_match:
                data["numero_recibo"] = recibo_match.group(1)

            # Extract filing date
            date_match = re.search(r'(?:Data|data).*?(?:de\s+)?(?:Envio|envio|Entrega|entrega)\s*[:=]?\s*(\d{1,2})[/-](\d{1,2})[/-](\d{4})', full_text, re.IGNORECASE)
            if date_match:
                data["data_envio"] = f"{date_match.group(3)}-{date_match.group(2).zfill(2)}-{date_match.group(1).zfill(2)}"

            # Extract total tax
            tax_patterns = [
                r'Imposto\s+Total\s*[:=]?\s*R?\s*([\d.]+,\d{2})',
                r'Valor\s+Total\s+(?:a\s+pagar|pago)\s*[:=]?\s*R?\s*([\d.]+,\d{2})',
                r'Total\s+(?:do\s+)?Imposto\s*[:=]?\s*R?\s*([\d.]+,\d{2})',
            ]

            for pattern in tax_patterns:
                tax_match = re.search(pattern, full_text, re.IGNORECASE)
                if tax_match:
                    data["imposto_total"] = parse_currency(tax_match.group(1))
                    break

            # Extract status
            if 'aceita' in full_text.lower():
                data["status"] = "Aceita"
            elif 'pendente' in full_text.lower():
                data["status"] = "Pendente"
            elif 'rejeitada' in full_text.lower():
                data["status"] = "Rejeitada"

    except Exception as e:
        logger.error(f"Failed to process receipt {pdf_path.name}: {e}")

    return data


def extract_real_estate(xlsx_path: Path) -> List[Dict]:
    """Extract data from real estate spreadsheet."""
    properties = []

    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        logger.info(f"Processing real estate file: {xlsx_path.name}")

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        # Extract data rows
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = ws.cell(row_idx, col_idx).value
                if header:
                    row_data[header] = cell_value

            # Check if row has meaningful data
            has_address = row_data.get('Endereço') or row_data.get('endereço')
            if has_address:
                # Convert datetime to string if needed
                data_compra = row_data.get('Data Compra') or row_data.get('data_compra')
                if hasattr(data_compra, 'isoformat'):
                    data_compra = data_compra.isoformat()
                else:
                    data_compra = str(data_compra) if data_compra else None

                prop = {
                    "endereco": has_address,
                    "data_compra": data_compra,
                    "valor_compra": parse_currency(row_data.get('Valor Compra') or row_data.get('valor_compra') or 0),
                    "vendedor": row_data.get('Vendedor') or row_data.get('vendedor'),
                    "financiamento": row_data.get('Financiamento') or row_data.get('financiamento'),
                    "status": row_data.get('Status') or row_data.get('status') or 'Ativo',
                }
                properties.append(prop)

        logger.info(f"Extracted {len(properties)} properties")

    except Exception as e:
        logger.error(f"Failed to process {xlsx_path.name}: {e}")

    return properties


def main():
    """Main extraction pipeline."""
    EXTRACT_DIR.mkdir(parents=True, exist_ok=True)

    # Dictionary to track all data
    all_data = {
        "declarations": [],
        "receipts": [],
        "properties": [],
        "members": set(),
        "anos_base": set(),
        "errors": [],
    }

    # Process IRPF declarations
    declaration_files = [
        DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracao_2023-0_original.pdf",
        DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracao_2024-0_original.pdf",
        DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracaomariana_2024-0_original.pdf",
    ]

    for decl_file in declaration_files:
        if decl_file.exists():
            data = extract_irpf_declaration(decl_file)
            all_data["declarations"].append(data)

            if data["membro"]:
                all_data["members"].add(data["membro"])
            if data["ano_base"]:
                all_data["anos_base"].add(data["ano_base"])

            # Save individual extract
            output_file = EXTRACT_DIR / decl_file.name.replace("-0_original.pdf", "-2_extract.json")
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logger.info(f"✓ Saved {output_file.name}")

    # Process IRPF receipts
    receipt_files = [
        DATA_DIR / "income_tax_br" / "receitafederal_irpfrecibo_2024-0_original.pdf",
        DATA_DIR / "income_tax_br" / "receitafederal_irpfrecibomariana_2024-0_original.pdf",
    ]

    for receipt_file in receipt_files:
        if receipt_file.exists():
            data = extract_irpf_receipt(receipt_file)
            all_data["receipts"].append(data)

            # Save individual extract
            output_file = EXTRACT_DIR / receipt_file.name.replace("-0_original.pdf", "-2_extract.json")
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logger.info(f"✓ Saved {output_file.name}")

    # Process real estate
    xlsx_file = DATA_DIR / "real_estate" / "dados_imoveis-0_original.xlsx"
    if xlsx_file.exists():
        properties = extract_real_estate(xlsx_file)
        all_data["properties"] = properties

        # Save real estate extract
        output_file = EXTRACT_DIR / "dados_imoveis-2_extract.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(properties, f, ensure_ascii=False, indent=2)
        logger.info(f"✓ Saved {output_file.name}")

    # Create consolidated baseline
    consolidated = {
        "pipeline_stage": "E1.5_Baseline_Patrimonial",
        "data_processamento": "2026-04-08",
        "membros": sorted(list(all_data["members"])),
        "anos_base": sorted(list(all_data["anos_base"])),
        "declarations": all_data["declarations"],
        "receipts": all_data["receipts"],
        "properties": all_data["properties"],
        "summary": {
            "total_members": len(all_data["members"]),
            "total_declarations": len(all_data["declarations"]),
            "total_receipts": len(all_data["receipts"]),
            "total_properties": len(all_data["properties"]),
            "total_assets": sum(d.get("total_bens", 0) for d in all_data["declarations"]),
            "total_income": sum(d.get("total_rendimentos", 0) for d in all_data["declarations"]),
        }
    }

    # Save consolidated file
    output_file = EXTRACT_DIR / "baseline_patrimonial-1.5_consolidated.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(consolidated, f, ensure_ascii=False, indent=2)
    logger.info(f"✓ Saved {output_file.name}")

    logger.info("\n" + "="*60)
    logger.info("EXTRACTION COMPLETE")
    logger.info("="*60)
    logger.info(f"Members: {consolidated['summary']['total_members']}")
    logger.info(f"Declarations: {consolidated['summary']['total_declarations']}")
    logger.info(f"Receipts: {consolidated['summary']['total_receipts']}")
    logger.info(f"Properties: {consolidated['summary']['total_properties']}")
    logger.info(f"Total Assets: R$ {consolidated['summary']['total_assets']:,.2f}")
    logger.info(f"Total Income: R$ {consolidated['summary']['total_income']:,.2f}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
