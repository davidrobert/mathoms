#!/usr/bin/env python3
"""
Extract patrimonial data from IRPF declarations, receipts, and real estate records.
Stage E1.5 of Financas Familia pipeline - Advanced extraction with detailed parsing.
"""

import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional
import logging

import pdfplumber
import openpyxl

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

BASE_DIR = Path("/sessions/sharp-cool-shannon/mnt/Financas Familia/financas-familia")
DATA_DIR = BASE_DIR / "data"
EXTRACT_DIR = BASE_DIR / "processed/E2_extracts"


def parse_currency(value_str: str) -> float:
    """Parse Brazilian currency format to float."""
    if not value_str:
        return 0.0
    value_str = str(value_str).strip()
    value_str = value_str.replace('.', '').replace(',', '.')
    try:
        return float(value_str)
    except:
        return 0.0


def extract_irpf_declaration(pdf_path: Path) -> Dict:
    """Extract detailed data from IRPF declaration using advanced parsing."""
    data = {
        "source": str(pdf_path.name),
        "membro": None,
        "cpf": None,
        "ano_base": None,
        "tipo_declaracao": None,
        "bens_direitos": [],
        "total_bens": 0.0,
        "rendimentos_pj": [],
        "rendimentos_pf": [],
        "rendimentos_isentos": [],
        "rendimentos_aplicacoes": [],
        "total_rendimentos": 0.0,
        "pagamentos_deductibles": [],
        "total_deducoes": 0.0,
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            logger.info(f"Processing {pdf_path.name} ({len(pdf.pages)} pages)")

            # Concatenate all text
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"

            # Extract header information
            _extract_header_info(full_text, data)

            # Extract assets (bens e direitos) - main section
            _extract_assets_detailed(full_text, data)

            # Extract rendimentos (income)
            _extract_income_detailed(full_text, data)

            # Extract deductions and payments
            _extract_deductions_detailed(full_text, data)

    except Exception as e:
        logger.error(f"Failed to process {pdf_path.name}: {e}")

    return data


def _extract_header_info(text: str, data: Dict):
    """Extract name, CPF, year from header."""
    # Extract name - appears right after NOME:
    name_match = re.search(r'NOME:\s*([A-Z\s]+?)(?:\n|CPF|\s{2,})', text)
    if name_match:
        data["membro"] = name_match.group(1).strip()

    # Extract CPF
    cpf_match = re.search(r'CPF:\s*(\d{3}\.\d{3}\.\d{3}-\d{2})', text)
    if cpf_match:
        data["cpf"] = cpf_match.group(1)

    # Extract year (Ano-calendário or ANO-CALENDÁRIO)
    year_match = re.search(r'(?:ANO-CALENDÁRIO|Ano-calendário)\s+(\d{4})', text)
    if year_match:
        data["ano_base"] = int(year_match.group(1))

    # Extract declaration type
    if 'Original' in text:
        data["tipo_declaracao"] = "Original"
    elif 'Retificadora' in text:
        data["tipo_declaracao"] = "Retificadora"


def _extract_assets_detailed(text: str, data: Dict):
    """Extract bens e direitos with detailed parsing."""
    # Find the assets section start
    assets_marker = 'DECLARAÇÃO DE BENS E DIREITOS'
    assets_start = text.find(assets_marker)

    if assets_start < 0:
        return

    # Find where this section ends
    end_markers = [
        'RESUMO DAS ATIVIDADES',
        'PERDAS EM OPERAÇÕES',
        'LIVRO CAIXA',
    ]

    assets_end = len(text)
    for marker in end_markers:
        pos = text.find(marker, assets_start)
        if pos > 0:
            assets_end = pos
            break

    assets_section = text[assets_start:assets_end]

    # Parse asset entries
    # Pattern: GRUPO CODE DESC ... VALUES
    # Find entries with group and code (e.g., "01 11")

    lines = assets_section.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        i += 1

        # Look for asset header line with group, code, description, and two values
        # Pattern: GRUPO CÓDIGO DISCRIMINAÇÃO ... SITUAÇÃO EM 31/12/XXXX 31/12/YYYY
        # Then: 01 11 DESCRIPTION VALUE1 VALUE2

        # Match: 2-digit group, space, 2-digit code, space, description
        match = re.match(r'^(\d{2})\s+(\d{2})\s+(.+)$', line)
        if match:
            group = match.group(1)
            code = match.group(2)
            desc_line = match.group(3)

            # Try to extract values from this line
            values = re.findall(r'([\d.]+,\d{2})', desc_line)

            asset = {
                "grupo": group,
                "codigo": code,
                "descricao": desc_line,
                "valor_31_12_anterior": 0.0,
                "valor_31_12_atual": 0.0,
                "detalhes": {}
            }

            # Extract values
            if len(values) >= 2:
                asset["valor_31_12_anterior"] = parse_currency(values[-2])
                asset["valor_31_12_atual"] = parse_currency(values[-1])
            elif len(values) == 1:
                asset["valor_31_12_atual"] = parse_currency(values[0])

            # Collect detail lines (following lines until next asset)
            detail_start = i
            while i < len(lines):
                next_line = lines[i].strip()

                # Stop if we hit a new asset (matches pattern "##\s+##")
                if re.match(r'^\d{2}\s+\d{2}\s+', next_line):
                    break

                # Stop if we hit a header line (GRUPO CÓDIGO DISCRIMINAÇÃO)
                if 'GRUPO' in next_line and 'CÓDIGO' in next_line:
                    i += 1
                    break

                # Stop if blank line after details
                if not next_line and (i - detail_start) > 5:
                    break

                # Collect detail
                if next_line and ':' in next_line:
                    key, val = next_line.split(':', 1)
                    asset["detalhes"][key.strip()] = val.strip()

                i += 1

            # Add asset if has values
            if asset["valor_31_12_atual"] > 0 or asset["valor_31_12_anterior"] > 0:
                data["bens_direitos"].append(asset)
                data["total_bens"] += asset["valor_31_12_atual"]


def _extract_income_detailed(text: str, data: Dict):
    """Extract income with detailed parsing."""

    # 1. Extract PJ/CLT income (RENDIMENTOS TRIBUTÁVEIS RECEBIDOS DE PESSOA JURÍDICA PELO TITULAR)
    pj_pattern = r'RENDIMENTOS TRIBUTÁVEIS RECEBIDOS DE PESSOA JURÍDICA PELO TITULAR.*?(?=RENDIMENTOS|DEDUÇÕES|$)'
    pj_match = re.search(pj_pattern, text, re.DOTALL | re.IGNORECASE)

    if pj_match:
        pj_section = pj_match.group(0)
        # Look for company names with amounts
        # Pattern: company name line followed by amounts
        lines = pj_section.split('\n')

        for line in lines:
            # Skip headers and empty lines
            if any(x in line.upper() for x in ['REND RECEBIDOS', 'CONTR PREVID', 'TOTAL GERAL', 'CNPJ']):
                continue

            # Extract monetary values
            values = re.findall(r'([\d.]+,\d{2})', line)

            if len(values) >= 1 and any(c.isalpha() for c in line):
                try:
                    desc = re.sub(r'[\d.,]+', '', line)[:100].strip()
                    rend = parse_currency(values[0])

                    if rend > 0:
                        entry = {
                            "tipo": "PJ/CLT",
                            "descricao": desc,
                            "rend_recebidos": rend,
                            "contrib_previd": parse_currency(values[1]) if len(values) > 1 else 0,
                            "imposto_retido": parse_currency(values[2]) if len(values) > 2 else 0,
                        }
                        data["rendimentos_pj"].append(entry)
                        data["total_rendimentos"] += rend
                except:
                    pass

    # 2. Extract rental income (PF)
    pf_pattern = r'RENDIMENTOS.*?RECEBIDOS DE PESSOA FÍSICA.*?(?=RENDIMENTOS ISENTOS|$)'
    pf_match = re.search(pf_pattern, text, re.DOTALL | re.IGNORECASE)

    if pf_match:
        pf_section = pf_match.group(0)
        lines = pf_section.split('\n')

        for line in lines:
            # Look for rental income (ALUGUÉIS or rentals)
            if 'aluguel' in line.lower():
                values = re.findall(r'([\d.]+,\d{2})', line)
                if values:
                    rend = parse_currency(values[0])
                    if rend > 0:
                        data["rendimentos_pf"].append({
                            "tipo": "Aluguel",
                            "descricao": line[:100],
                            "valor": rend,
                        })
                        data["total_rendimentos"] += rend

    # 3. Extract exempt income (RENDIMENTOS ISENTOS E NÃO TRIBUTÁVEIS)
    exempt_pattern = r'RENDIMENTOS ISENTOS E NÃO TRIBUTÁVEIS.*?TOTAL\s+([\d.]+,\d{2})'
    exempt_match = re.search(exempt_pattern, text, re.DOTALL | re.IGNORECASE)

    if exempt_match:
        total_exempt = parse_currency(exempt_match.group(1))
        if total_exempt > 0:
            data["rendimentos_isentos"].append({
                "tipo": "Isento",
                "total": total_exempt,
            })

    # 4. Extract financial income (aplicações financeiras)
    app_pattern = r'(?:13º salário|Rendimentos de aplicações).*?(\d+\.\d{3},\d{2})'
    app_matches = re.finditer(app_pattern, text, re.IGNORECASE)

    for match in app_matches:
        valor = parse_currency(match.group(1))
        if valor > 0:
            data["rendimentos_aplicacoes"].append({
                "tipo": "Aplicação Financeira",
                "valor": valor,
            })


def _extract_deductions_detailed(text: str, data: Dict):
    """Extract deductible payments."""

    # Find PAGAMENTOS EFETUADOS section
    payments_pattern = r'PAGAMENTOS EFETUADOS.*?(?=DOAÇÕES|DECLARAÇÃO|$)'
    payments_match = re.search(payments_pattern, text, re.DOTALL | re.IGNORECASE)

    if payments_match:
        payments_section = payments_match.group(0)
        lines = payments_section.split('\n')

        for line in lines:
            # Skip headers and empty lines
            if any(x in line.upper() for x in ['CÓD NOME', 'BENEFICIÁRIO', 'PARC']):
                continue

            # Look for lines with amounts
            values = re.findall(r'([\d.]+,\d{2})', line)

            if values and any(c.isalpha() for c in line):
                try:
                    amount = parse_currency(values[0])
                    if amount > 0:
                        data["pagamentos_deductibles"].append({
                            "descricao": line[:100].strip(),
                            "valor": amount,
                        })
                        data["total_deducoes"] += amount
                except:
                    pass


def extract_irpf_receipt(pdf_path: Path) -> Dict:
    """Extract data from IRPF receipt (recibo)."""
    data = {
        "source": str(pdf_path.name),
        "membro": None,
        "cpf": None,
        "imposto_total": 0.0,
        "imposto_devido": 0.0,
        "data_envio": None,
        "status": None,
        "numero_recibo": None,
        "total_rendimentos": 0.0,
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            logger.info(f"Processing receipt {pdf_path.name}")

            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"

            # Extract CPF and name from receipt header
            # Pattern: "CPF do declarante Nome do declarante" followed by actual values
            cpf_pattern = r'(\d{3}\.\d{3}\.\d{3}-\d{2})\s+([A-Z][A-Z\s]+?)(?:\s+\(?\d{2}\)|\n|Endereço)'
            cpf_match = re.search(cpf_pattern, full_text)
            if cpf_match:
                data["cpf"] = cpf_match.group(1)
                data["membro"] = cpf_match.group(2).strip()

            # Extract receipt number (NÚMERO DO RECIBO)
            recibo_patterns = [
                r'NÚMERO DO RECIBO.*?(\d{2}\.\d{2}\.\d{2}\.\d{2}\.\d{2}\s*-\s*\d{2})',
                r'é:\s*(\d{2}\.\d{2}\.\d{2}\.\d{2}\.\d{2}\s*-\s*\d{2})',
            ]
            for pattern in recibo_patterns:
                recibo_match = re.search(pattern, full_text, re.IGNORECASE)
                if recibo_match:
                    data["numero_recibo"] = recibo_match.group(1).strip()
                    break

            # Extract filing date (delivered date)
            date_pattern = r'em\s+(\d{1,2})/(\d{1,2})/(\d{4})\s+às'
            date_match = re.search(date_pattern, full_text)
            if date_match:
                data["data_envio"] = f"{date_match.group(3)}-{date_match.group(2).zfill(2)}-{date_match.group(1).zfill(2)}"

            # Extract total rendimentos
            rend_pattern = r'TOTAL RENDIMENTOS TRIBUTÁVEIS\s+([\d.]+,\d{2})'
            rend_match = re.search(rend_pattern, full_text)
            if rend_match:
                data["total_rendimentos"] = parse_currency(rend_match.group(1))

            # Extract total tax (IMPOSTO DEVIDO)
            tax_patterns = [
                r'IMPOSTO DEVIDO\s+([\d.]+,\d{2})',
                r'IMPOSTO A PAGAR\s+([\d.]+,\d{2})',
            ]
            for pattern in tax_patterns:
                tax_match = re.search(pattern, full_text)
                if tax_match:
                    data["imposto_total"] = parse_currency(tax_match.group(1))
                    break

            # Extract imposto_devido if different
            devido_match = re.search(r'IMPOSTO DEVIDO\s+([\d.]+,\d{2})', full_text)
            if devido_match:
                data["imposto_devido"] = parse_currency(devido_match.group(1))

            # Status is typically "Aceita" for receipts
            data["status"] = "Aceita"

    except Exception as e:
        logger.error(f"Failed to process receipt {pdf_path.name}: {e}")

    return data


def extract_real_estate(xlsx_path: Path) -> List[Dict]:
    """Extract data from real estate spreadsheet."""
    properties = []

    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        logger.info(f"Processing real estate file: {xlsx_path.name}")

        # Get headers from first row (strip whitespace/newlines)
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(1, col_idx).value
            if header:
                # Clean up header (remove newlines, extra spaces)
                clean_header = re.sub(r'\s+', ' ', str(header)).strip()
                headers[col_idx] = clean_header

        # Extract data rows
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, clean_header in headers.items():
                cell_value = ws.cell(row_idx, col_idx).value
                row_data[clean_header] = cell_value

            # Check if row has meaningful data (has address)
            has_address = None
            for key in row_data:
                if 'endereço' in key.lower():
                    has_address = row_data[key]
                    break

            if has_address:
                # Find the right columns for data extraction
                data_aquisicao = None
                valor_compra_val = 0.0
                tipo_imovel = None
                detalhes_compra = None
                banco_financiamento = None

                for key, val in row_data.items():
                    if 'data de aquisição' in key.lower() or 'data de assinatura do registro' in key.lower():
                        data_aquisicao = val
                    elif 'valor de compra' in key.lower():
                        valor_compra_val = parse_currency(val) if val else 0.0
                    elif 'tipo' in key.lower() and val:
                        tipo_imovel = val
                    elif 'detalhes da compra' in key.lower():
                        detalhes_compra = val
                    elif 'banco' in key.lower() and 'financiamento' in key.lower():
                        banco_financiamento = val

                # Convert datetime to string if needed
                if hasattr(data_aquisicao, 'isoformat'):
                    data_aquisicao = data_aquisicao.isoformat()
                elif data_aquisicao:
                    data_aquisicao = str(data_aquisicao)

                prop = {
                    "tipo_imovel": tipo_imovel,
                    "endereco": str(has_address),
                    "data_aquisicao": data_aquisicao,
                    "valor_compra": valor_compra_val,
                    "detalhes_compra": detalhes_compra,
                    "banco_financiamento": banco_financiamento,
                    "status": "Ativo",
                }
                properties.append(prop)

        logger.info(f"Extracted {len(properties)} properties")

    except Exception as e:
        logger.error(f"Failed to process {xlsx_path.name}: {e}")

    return properties


def main():
    """Main extraction pipeline."""
    EXTRACT_DIR.mkdir(parents=True, exist_ok=True)

    all_data = {
        "declarations": [],
        "receipts": [],
        "properties": [],
        "members": set(),
        "anos_base": set(),
    }

    # Process IRPF declarations
    declaration_files = [
        DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracao_2023-0_original.pdf",
        DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracao_2024-0_original.pdf",
        DATA_DIR / "income_tax_br" / "receitafederal_irpfdeclaracaomariana_2024-0_original.pdf",
    ]

    for decl_file in declaration_files:
        if decl_file.exists():
            data = extract_irpf_declaration(decl_file)
            all_data["declarations"].append(data)

            if data["membro"]:
                all_data["members"].add(data["membro"])
            if data["ano_base"]:
                all_data["anos_base"].add(data["ano_base"])

            # Save individual extract
            output_file = EXTRACT_DIR / decl_file.name.replace("-0_original.pdf", "-2_extract.json")
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logger.info(f"✓ Saved {output_file.name}")

    # Process IRPF receipts
    receipt_files = [
        DATA_DIR / "income_tax_br" / "receitafederal_irpfrecibo_2024-0_original.pdf",
        DATA_DIR / "income_tax_br" / "receitafederal_irpfrecibomariana_2024-0_original.pdf",
    ]

    for receipt_file in receipt_files:
        if receipt_file.exists():
            data = extract_irpf_receipt(receipt_file)
            all_data["receipts"].append(data)

            # Save individual extract
            output_file = EXTRACT_DIR / receipt_file.name.replace("-0_original.pdf", "-2_extract.json")
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logger.info(f"✓ Saved {output_file.name}")

    # Process real estate
    xlsx_file = DATA_DIR / "real_estate" / "dados_imoveis-0_original.xlsx"
    if xlsx_file.exists():
        properties = extract_real_estate(xlsx_file)
        all_data["properties"] = properties

        # Save real estate extract
        output_file = EXTRACT_DIR / "dados_imoveis-2_extract.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(properties, f, ensure_ascii=False, indent=2)
        logger.info(f"✓ Saved {output_file.name}")

    # Create consolidated baseline
    consolidated = {
        "pipeline_stage": "E1.5_Baseline_Patrimonial",
        "data_processamento": "2026-04-08",
        "membros": sorted(list(all_data["members"])),
        "anos_base": sorted(list(all_data["anos_base"])),
        "declarations": all_data["declarations"],
        "receipts": all_data["receipts"],
        "properties": all_data["properties"],
        "summary": {
            "total_members": len(all_data["members"]),
            "total_declarations": len(all_data["declarations"]),
            "total_receipts": len(all_data["receipts"]),
            "total_properties": len(all_data["properties"]),
            "total_assets_bens_direitos": sum(d.get("total_bens", 0) for d in all_data["declarations"]),
            "total_income": sum(d.get("total_rendimentos", 0) for d in all_data["declarations"]),
            "total_deductions": sum(d.get("total_deducoes", 0) for d in all_data["declarations"]),
        }
    }

    # Save consolidated file
    output_file = EXTRACT_DIR / "baseline_patrimonial-1.5_consolidated.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(consolidated, f, ensure_ascii=False, indent=2)
    logger.info(f"✓ Saved {output_file.name}")

    logger.info("\n" + "="*70)
    logger.info("STAGE E1.5 - BASELINE PATRIMONIAL EXTRACTION COMPLETE")
    logger.info("="*70)
    logger.info(f"Members: {consolidated['summary']['total_members']}")
    logger.info(f"Declarations: {consolidated['summary']['total_declarations']}")
    logger.info(f"Receipts: {consolidated['summary']['total_receipts']}")
    logger.info(f"Properties: {consolidated['summary']['total_properties']}")
    logger.info(f"Total Assets (Bens): R$ {consolidated['summary']['total_assets_bens_direitos']:,.2f}")
    logger.info(f"Total Income: R$ {consolidated['summary']['total_income']:,.2f}")
    logger.info(f"Total Deductions: R$ {consolidated['summary']['total_deductions']:,.2f}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
